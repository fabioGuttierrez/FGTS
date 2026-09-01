import time
import threading
from datetime import datetime, date
from decimal import Decimal
from dateutil.relativedelta import relativedelta
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.views.generic import FormView, CreateView, UpdateView, ListView, View, DetailView
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from billing.services.features import can_use_feature, feature_block_context
from empresas.models import Empresa
from .models import Lancamento, ImportacaoLancamento
from .models import ImportacaoResponsabilidade
from .models_conferencia import ConferenciaLancamento
from .forms import (
    RelatorioCompetenciaForm,
    LancamentoForm,
    LegacyImportForm,
    SefipExportForm,
    ConferenciaLancamentoForm,
    RejeicaoLancamentoForm,
    FiltroConferenciaForm,
    SefipImportForm,
    RelatorioRecolhimentoFuncionarioForm,
    ImportacaoUploadForm,
    ImportacaoConfirmacaoForm,
)
from .services.calculo import (
    calcular_fgts_atualizado,
    gerar_memoria_calculo,
    get_config_numeric,
    get_config_str,
    aplicar_plano_economico_legacy,
    calcular_jam_ate_pagamento,
    buscar_coef_jam,
)
from .services.importacao import LancamentoImportService
from .services.competencia_13 import Competencia13Service
from .services.sefip_legacy import SefipLegacyFilters, SefipExportError, gerar_sefip_legacy
from empresas.models_feature import empresa_tem_recurso
from django.conf import settings
from indices.services.indice_service import IndiceFGTSService
from funcionarios.models import Funcionario
from fgtsweb.mixins import get_allowed_empresa_ids, is_empresa_allowed, EmpresaScopeMixin
from django.http import HttpResponseForbidden, HttpResponse
from django.db.models.functions import Substr
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required


def _process_importacao(importacao_id):
    """Processa importação de lançamentos em background thread."""
    from django.db import connection
    importacao = None
    try:
        importacao = ImportacaoLancamento.objects.get(id=importacao_id)
        importacao.status = 'processing'
        importacao.save(update_fields=['status', 'atualizado_em'])

        # Callback de progresso com throttle de 2 segundos para reduzir writes no DB.
        last_update = [0.0]

        def on_progress(linhas_processadas, linhas_total):
            now = time.time()
            if linhas_processadas == 0:
                # Primeira chamada: salva o total de linhas
                importacao.linhas_total = linhas_total
                importacao.save(update_fields=['linhas_total', 'atualizado_em'])
                last_update[0] = now
            elif now - last_update[0] >= 2.0:
                # Demais chamadas: throttle de 2s
                importacao.linhas_processadas = linhas_processadas
                importacao.save(update_fields=['linhas_processadas', 'atualizado_em'])
                last_update[0] = now

        with open(importacao.arquivo.path, 'rb') as f:
            result = LancamentoImportService.import_lancamentos_from_file(
                f, importacao.empresa, importacao.usuario, progress_callback=on_progress,
                recalcular_fgts=importacao.recalcular_fgts,
                aplicar_jam=importacao.aplicar_jam,
                data_referencia_jam=importacao.data_referencia_jam,
                extrato_analitico=importacao.extrato_analitico,
            )

        importacao.status = 'done'
        importacao.resultado_json = result
        importacao.save(update_fields=['status', 'resultado_json', 'atualizado_em'])

        # Atualizar contadores de responsabilidade
        try:
            resp = importacao.responsabilidade
            resp.linhas_valor_do_arquivo = result.get('linhas_valor_do_arquivo', 0)
            resp.linhas_jam_aplicado = result.get('linhas_jam_aplicado', 0)
            resp.save(update_fields=['linhas_valor_do_arquivo', 'linhas_jam_aplicado'])
        except Exception:
            pass
    except Exception as e:
        try:
            if importacao is None:
                importacao = ImportacaoLancamento.objects.get(id=importacao_id)
            importacao.status = 'error'
            importacao.mensagem_erro = str(e)
            importacao.save(update_fields=['status', 'mensagem_erro', 'atualizado_em'])
        except Exception:
            pass
    finally:
        connection.close()


class LancamentoDeleteView(LoginRequiredMixin, EmpresaScopeMixin, View):
    """Exclui um lançamento específico via POST."""

    def post(self, request, pk):
        try:
            lancamento = get_object_or_404(Lancamento, pk=pk)

            # Permissão por empresa
            if not is_empresa_allowed(request.user, lancamento.empresa.codigo):
                messages.error(request, '❌ Você não tem permissão para excluir este lançamento.')
                return redirect('lancamento-list')

            lancamento.delete()
            messages.success(request, '🗑️ Lançamento excluído com sucesso.')
        except Exception as e:
            messages.error(request, f'❌ Erro ao excluir lançamento: {str(e)}')

        return redirect('lancamento-list')


class LancamentoCreateView(LoginRequiredMixin, EmpresaScopeMixin, CreateView):
    """Criar novo lançamento mensal (base FGTS)"""
    model = Lancamento
    form_class = LancamentoForm
    template_name = 'lancamentos/lancamento_form.html'
    success_url = reverse_lazy('lancamento-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        empresa = form.cleaned_data.get('empresa')
        if empresa and not is_empresa_allowed(self.request.user, empresa.codigo):
            return HttpResponseForbidden('Empresa não permitida para este usuário.')
        
        lancamento = form.save()  # Já calcula valor_fgts no save() do formulário
        vinculo_label = f" (matrícula {lancamento.vinculo.matricula})" if getattr(lancamento, 'vinculo', None) and lancamento.vinculo.matricula else ""
        messages.success(self.request, f'✅ Lançamento para {lancamento.funcionario.nome}{vinculo_label} ({lancamento.competencia}) registrado com sucesso!')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Foram encontradas inconsistencias nas informacoes enviadas. Revise as informacoes.')
        return super().form_invalid(form)


class LancamentoUpdateView(LoginRequiredMixin, EmpresaScopeMixin, UpdateView):
    """Editar lançamento mensal"""
    model = Lancamento
    form_class = LancamentoForm
    template_name = 'lancamentos/lancamento_form.html'
    success_url = reverse_lazy('lancamento-list')
    pk_url_kwarg = 'pk'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        empresa = form.cleaned_data.get('empresa')
        if empresa and not is_empresa_allowed(self.request.user, empresa.codigo):
            return HttpResponseForbidden('Empresa não permitida para este usuário.')
        
        lancamento = form.save()
        vinculo_label = f" (matrícula {lancamento.vinculo.matricula})" if getattr(lancamento, 'vinculo', None) and lancamento.vinculo.matricula else ""
        messages.success(self.request, f'✅ Lançamento para {lancamento.funcionario.nome}{vinculo_label} ({lancamento.competencia}) atualizado com sucesso!')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Foram encontradas inconsistencias nas informacoes enviadas. Revise as informacoes.')
        return super().form_invalid(form)


class LancamentoListView(LoginRequiredMixin, EmpresaScopeMixin, ListView):
    """Listar lançamentos cadastrados"""
    model = Lancamento
    template_name = 'lancamentos/lancamento_list.html'
    context_object_name = 'lancamentos'
    paginate_by = 20


    def get_queryset(self):
        from empresas.models_grupo import FuncionarioVinculo
        from django.db import models
        from django.db.models import OuterRef, Exists, Q, DateField, Value, F, Func, Case, When, BooleanField
        from django.db.models.functions import Substr, Cast, TruncMonth
        import datetime

        qs = super().get_queryset().select_related('empresa', 'funcionario', 'vinculo').prefetch_related('funcionario__vinculos')
        qs = qs.annotate(
            ano_comp=Case(
                When(
                    competencia__regex=r'^\d{2}/\d{4}$',
                    then=Substr('competencia', 4, 4)
                ),
                When(
                    competencia__regex=r'^\d{4}-\d{2}$',
                    then=Substr('competencia', 1, 4)
                ),
                default=Substr('competencia', 1, 4),
                output_field=models.CharField()
            )
        )

        allowed_ids = get_allowed_empresa_ids(self.request.user)
        if allowed_ids is not None:
            qs = qs.filter(empresa__codigo__in=allowed_ids)

        competencia = self.request.GET.get('competencia', '').strip()
        funcionario_id = self.request.GET.get('funcionario', '').strip()
        empresa_id = self.request.GET.get('empresa', '').strip()
        matricula = self.request.GET.get('matricula', '').strip()
        vinculo_id = self.request.GET.get('vinculo', '').strip()
        ano = self.request.GET.get('ano', '').strip()
        fonte_confirmacao = self.request.GET.get('fonte_confirmacao', '').strip()

        # Usar sempre MM/YYYY (string) para busca
        if competencia:
            qs = qs.filter(competencia=competencia)

        if funcionario_id:
            qs = qs.filter(funcionario_id=funcionario_id)

        if empresa_id:
            qs = qs.filter(empresa_id=empresa_id)

        if vinculo_id:
            qs = qs.filter(vinculo_id=vinculo_id)

        if matricula:
            qs = qs.filter(vinculo__matricula__icontains=matricula)

        if fonte_confirmacao == 'nao_pago':
            qs = qs.filter(pago=False)
        elif fonte_confirmacao == 'manual':
            qs = qs.filter(pago=True, fonte_confirmacao_pagamento='manual')
        elif fonte_confirmacao == 'extrato_analitico':
            qs = qs.filter(pago=True, fonte_confirmacao_pagamento='extrato_analitico')

        if ano:
            qs = qs.filter(ano_comp=ano)

        # Refatoração: filtro de vínculo ativo via Exists (SQL)
        # Considera vínculo ativo se:
        # - empresa igual
        # - funcionario igual
        # - data_admissao <= competência
        # - (data_demissao é nula ou >= competência)

        # Converter competencia (str) para data (primeiro dia do mês)

        # Anotar a data da competência (primeiro dia do mês) para cada lançamento

        from django.db.models.functions import Concat, Substr
        from django.db.models import Case, When
        # Se competencia tem '/' (MM/YYYY), converte para YYYY-MM-01, senão assume YYYY-MM-01
        qs = qs.annotate(
            competencia_date=Cast(
                Case(
                    When(
                        competencia__regex=r'^\d{2}/\d{4}$',
                        then=Concat(
                            Substr('competencia', 4, 4),  # YYYY
                            Value('-'),
                            Substr('competencia', 1, 2),  # MM
                            Value('-01')
                        )
                    ),
                    default=Concat(F('competencia'), Value('-01')),
                    output_field=models.CharField()
                ),
                output_field=DateField()
            )
        )

        vinculo_exists = Exists(
            FuncionarioVinculo.objects.annotate(
                adm_mes=TruncMonth('data_admissao'),
                dem_mes=TruncMonth('data_demissao'),
            ).filter(
                funcionario_id=OuterRef('funcionario_id'),
                empresa_id=OuterRef('empresa_id'),
                adm_mes__lte=OuterRef('competencia_date'),
            ).filter(
                Q(dem_mes__isnull=True) | Q(dem_mes__gte=OuterRef('competencia_date'))
            )
        )

        dup_vinculo_qs = Lancamento.objects.filter(
            competencia=OuterRef('competencia'),
            parcela_13=OuterRef('parcela_13'),
            vinculo_id=OuterRef('vinculo_id'),
        ).exclude(pk=OuterRef('pk'))

        dup_funcionario_qs = Lancamento.objects.filter(
            competencia=OuterRef('competencia'),
            parcela_13=OuterRef('parcela_13'),
            vinculo__isnull=True,
            funcionario_id=OuterRef('funcionario_id'),
        ).exclude(pk=OuterRef('pk'))

        qs = qs.annotate(
            is_duplicate=Case(
                When(vinculo_id__isnull=False, then=Exists(dup_vinculo_qs)),
                default=Exists(dup_funcionario_qs),
                output_field=BooleanField(),
            )
        )

        qs = qs.annotate(vinculo_legado_ativo=vinculo_exists)

        qs = qs.annotate(
            vinculo_adm_mes=TruncMonth('vinculo__data_admissao'),
            vinculo_dem_mes=TruncMonth('vinculo__data_demissao'),
        )

        vinculo_ativo_explicito = Case(
            When(
                condition=(
                    Q(vinculo__isnull=False)
                    & Q(vinculo_adm_mes__lte=F('competencia_date'))
                    & (Q(vinculo_dem_mes__isnull=True) | Q(vinculo_dem_mes__gte=F('competencia_date')))
                ),
                then=Value(True),
            ),
            When(
                condition=Q(vinculo__isnull=True) & Q(vinculo_legado_ativo=True),
                then=Value(True),
            ),
            default=Value(False),
            output_field=models.BooleanField(),
        )
        qs = qs.annotate(vinculo_ativo=vinculo_ativo_explicito).filter(vinculo_ativo=True)

        # Ordenação
        ordem = self.request.GET.get('ordem', '-competencia').strip()
        if ordem == 'competencia_asc':
            qs = qs.order_by('competencia')
        elif ordem == 'competencia_desc':
            qs = qs.order_by('-competencia')
        elif ordem == 'ano_asc':
            qs = qs.order_by('ano_comp', 'competencia')
        elif ordem == 'ano_desc':
            qs = qs.order_by('-ano_comp', '-competencia')
        elif ordem == 'funcionario_asc':
            qs = qs.order_by('funcionario__nome')
        elif ordem == 'funcionario_desc':
            qs = qs.order_by('-funcionario__nome')
        else:
            qs = qs.order_by('-competencia')

        return qs
    
    def get_context_data(self, **kwargs):
        # Local import to guarantee binding even if globals change or circular imports occur
        from funcionarios.models import Funcionario as FuncionarioModel

        context = super().get_context_data(**kwargs)

        # Adicionar empresas e funcionários permitidos para o filtro
        allowed_ids = get_allowed_empresa_ids(self.request.user)
        if allowed_ids is not None:
            context['empresas'] = Empresa.objects.filter(codigo__in=allowed_ids)
            funcionario_ids = FuncionarioModel.objects.filter(
                vinculos__empresa__codigo__in=allowed_ids
            ).distinct().values_list('id', flat=True)
            context['funcionarios'] = FuncionarioModel.objects.filter(id__in=funcionario_ids).order_by('nome')
            base_qs = Lancamento.objects.filter(empresa__codigo__in=allowed_ids)
        else:
            context['empresas'] = Empresa.objects.all()
            context['funcionarios'] = FuncionarioModel.objects.all().order_by('nome')
            base_qs = Lancamento.objects.all()

        # Aplicar os mesmos filtros do get_queryset
        competencia = self.request.GET.get('competencia', '').strip()
        funcionario_id = self.request.GET.get('funcionario', '').strip()
        empresa_id = self.request.GET.get('empresa', '').strip()
        # Não aplicar filtro de competência para o filtro de ano, para garantir que todos os anos presentes sejam exibidos

        def extract_ano(comp):
            if not comp:
                return None
            if '/' in comp:
                candidate = comp.split('/')[-1]
            elif '-' in comp:
                parts = comp.split('-')
                candidate = parts[0] if len(parts[0]) == 4 else parts[-1]
            else:
                candidate = comp[:4]
            return candidate if len(candidate) == 4 and candidate.isdigit() else None

        competencias_para_anos = base_qs.values_list('competencia', flat=True).distinct()
        anos_todos = set()
        for comp in competencias_para_anos:
            ano_extraido = extract_ano(comp)
            if ano_extraido:
                anos_todos.add(ano_extraido)
        context['anos_todos'] = sorted(anos_todos, reverse=True)

        # Agora sim, aplicar o filtro de competência para a listagem
        if competencia:
            comp_parts = competencia.split('/')
            competencia_db = f"{comp_parts[1]}-{comp_parts[0]}" if len(comp_parts) == 2 else competencia
            base_qs = base_qs.filter(competencia=competencia_db)
        if funcionario_id:
            base_qs = base_qs.filter(funcionario_id=funcionario_id)
        if empresa_id:
            base_qs = base_qs.filter(empresa_id=empresa_id)

        competencias = base_qs.values_list('competencia', flat=True).distinct()
        anos_validos = set()
        for comp in competencias:
            ano_extraido = extract_ano(comp)
            if ano_extraido:
                anos_validos.add(ano_extraido)
        context['anos'] = sorted(anos_validos, reverse=True)

        # Passar parâmetros de filtro para o template
        context['competencia_filtro'] = self.request.GET.get('competencia', '')
        context['funcionario_filtro'] = self.request.GET.get('funcionario', '')
        context['empresa_filtro'] = self.request.GET.get('empresa', '')
        context['matricula_filtro'] = self.request.GET.get('matricula', '')
        context['vinculo_filtro'] = self.request.GET.get('vinculo', '')
        context['ano_filtro'] = self.request.GET.get('ano', '')
        context['fonte_confirmacao_filtro'] = self.request.GET.get('fonte_confirmacao', '')
        context['ordem_filtro'] = self.request.GET.get('ordem', '-competencia')

        # Empresa de referência para validar plano/trial (filtro, usuário ou única empresa permitida)
        empresa_contexto = None
        empresa_param = context['empresa_filtro']
        if empresa_param:
            try:
                empresa_contexto = Empresa.objects.filter(pk=empresa_param).first()
            except Exception:
                empresa_contexto = None
        if not empresa_contexto:
            try:
                empresa_contexto = self.request.user.empresa
            except Exception:
                empresa_contexto = None
        if not empresa_contexto and allowed_ids and len(allowed_ids) == 1:
            empresa_contexto = Empresa.objects.filter(codigo=allowed_ids[0]).first()

        bloqueio_ctx = feature_block_context('custom_reports', user=self.request.user, empresa=empresa_contexto)
        context['relatorio_bloqueado'] = bloqueio_ctx['feature_blocked']
        context['relatorio_bloqueio_motivo'] = bloqueio_ctx['feature_block_reason']
        context['empresa_contexto'] = empresa_contexto

        context['pode_importar_re_sefip'] = False
        context['pode_importar_extrato_cef'] = False

        # Buscar a data da última atualização da tabela indices_fgts (SupabaseIndice)
        try:
            from indices.models import SupabaseIndice
            ultima_atualizacao = SupabaseIndice.objects.order_by('-data_base').values_list('data_base', flat=True).first()
        except Exception:
            ultima_atualizacao = None
        context['ultima_atualizacao_indices_fgts'] = ultima_atualizacao

        return context


@login_required
def lancamento_ids(request):
    """Retorna todos os IDs de lançamentos aplicando os mesmos filtros da listagem."""
    view = LancamentoListView()
    view.request = request
    qs = view.get_queryset()
    ids = list(qs.values_list('id', flat=True))
    return JsonResponse({'ids': ids})



class RelatorioCompetenciaView(FormView):
    def normalizar_competencia(self, comp):
        """Aceita 'MM/YYYY', 'YYYY-MM' ou 'YYYY/MM' e retorna sempre 'MM/YYYY'."""
        if isinstance(comp, str):
            if '/' in comp:
                parts = comp.split('/')
                if len(parts) == 2 and len(parts[1]) == 4:
                    # MM/YYYY
                    return f'{parts[0].zfill(2)}/{parts[1]}'
                if len(parts) == 2 and len(parts[0]) == 4:
                    # YYYY/MM
                    return f'{parts[1].zfill(2)}/{parts[0]}'
            if '-' in comp:
                parts = comp.split('-')
                if len(parts) == 2 and len(parts[0]) == 4:
                    # YYYY-MM
                    return f'{parts[1].zfill(2)}/{parts[0]}'
        return comp
    template_name = 'lancamentos/relatorio_competencia.html'
    form_class = RelatorioCompetenciaForm
    success_url = reverse_lazy('relatorio-competencia')
    # Configurações de proteção contra loops
    MAX_ITERACOES_POR_COMPETENCIA = 10  # Máximo de vezes que a mesma competência pode ser reprocessada
    TIMEOUT_GLOBAL_SEGUNDOS = 60  # Timeout total de 60 segundos
    MAX_COMPETENCIAS = 12  # Limite para evitar timeouts em requisições grandes
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.tempo_inicio = None
        self.competencias_processadas = {}  # Track {competencia: count}
    
    def _verificar_loop(self, competencia_str):
        """Verifica se há risco de loop infinito"""
        import time
        
        # Verificar timeout global
        if self.tempo_inicio is None:
            self.tempo_inicio = time.time()
        
        tempo_decorrido = time.time() - self.tempo_inicio
        if tempo_decorrido > self.TIMEOUT_GLOBAL_SEGUNDOS:
            raise Exception(
                f"🛑 TIMEOUT: Processamento levou mais de {self.TIMEOUT_GLOBAL_SEGUNDOS}s. "
                f"Interrompendo para evitar loop infinito."
            )
        
        # Verificar iterações por competência
        if competencia_str not in self.competencias_processadas:
            self.competencias_processadas[competencia_str] = 0
        
        self.competencias_processadas[competencia_str] += 1
        contador = self.competencias_processadas[competencia_str]
        
        if contador > self.MAX_ITERACOES_POR_COMPETENCIA:
            raise Exception(
                f"🛑 LOOP DETECTADO: Competência {competencia_str} foi processada {contador} vezes. "
                f"Limite máximo de {self.MAX_ITERACOES_POR_COMPETENCIA} iterações excedido. "
                f"Há um loop infinito no processamento."
            )
        
        # Aviso quando aproximando do limite
        if contador > self.MAX_ITERACOES_POR_COMPETENCIA * 0.7:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(
                f"⚠️ AVISO DE LOOP: Competência {competencia_str} já foi processada {contador} vezes "
                f"({int((contador/self.MAX_ITERACOES_POR_COMPETENCIA)*100)}% do limite)."
            )
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        empresa_ctx = None
        form = ctx.get('form')
        if form:
            empresa_value = None
            try:
                empresa_value = form.data.get('empresa') or form.initial.get('empresa')
            except Exception:
                empresa_value = None

            if empresa_value:
                if isinstance(empresa_value, Empresa):
                    empresa_ctx = empresa_value
                else:
                    try:
                        empresa_ctx = Empresa.objects.filter(pk=empresa_value).first()
                    except Exception:
                        empresa_ctx = None

        if not empresa_ctx:
            try:
                empresa_ctx = self.request.user.empresa
            except Exception:
                empresa_ctx = None

        bloqueio_ctx = feature_block_context('custom_reports', user=self.request.user, empresa=empresa_ctx)
        ctx['relatorio_bloqueado'] = bloqueio_ctx['feature_blocked']
        ctx['relatorio_bloqueio_motivo'] = bloqueio_ctx['feature_block_reason']
        ctx['empresa_contexto'] = empresa_ctx
        ctx['exibir_indice'] = self.request.session.get('exibir_indice', False)
        ctx['exibir_jam'] = self.request.session.get('exibir_jam', True)
        ctx['exibir_correcao'] = self.request.session.get('exibir_correcao', True)
        return ctx

    def form_invalid(self, form):
        messages.error(self.request, 'Foram encontradas inconsistencias nas informacoes enviadas. Revise as informacoes.')
        return super().form_invalid(form)

    def _agrupar_resultados(self, resultados, agrupamento):
        """Agrupa resultados por competência, ano, funcionário ou vínculo"""
        from collections import defaultdict
        
        grupos = defaultdict(lambda: {
            'items': [],
            'totais': {k: Decimal('0') for k in ['valor_fgts', 'valor_corrigido', 'valor_jam', 'valor_deposito_fgts', 'total']}
        })
        
        for resultado in resultados:
            lancamento = resultado['lancamento']
            calc = resultado['calc']
            competencia_raw = resultado.get('competencia')
            competencia_label = resultado.get('competencia_display', competencia_raw)
            parcela_13 = resultado.get('parcela_13') or 0
			
            # Determinar chave do grupo
            if agrupamento == 'ano':
                # Extrair ano da competência (MM/YYYY)
                ano = competencia_raw.split('/')[-1] if competencia_raw and '/' in competencia_raw else competencia_raw
                chave = ano
                label = f"Ano {ano}"
            elif agrupamento == 'funcionario':
                chave = lancamento.funcionario.pk
                label = f"{lancamento.funcionario.nome} - {lancamento.funcionario.cpf}"
            elif agrupamento == 'vinculo':
                if getattr(lancamento, 'vinculo_id', None):
                    chave = f"vinc_{lancamento.vinculo_id}"
                    matricula = (getattr(lancamento.vinculo, 'matricula', None) or '').strip()
                    matricula_label = matricula if matricula else str(lancamento.vinculo_id)
                    label = f"{lancamento.funcionario.nome} — matrícula {matricula_label}"
                else:
                    chave = f"func_{lancamento.funcionario.pk}"
                    label = f"{lancamento.funcionario.nome} - {lancamento.funcionario.cpf} (sem vínculo)"
            else:  # competencia
                chave = f"{competencia_raw}|{parcela_13}"
                label = competencia_label
            
            grupos[chave]['label'] = label
            grupos[chave]['items'].append(resultado)
            
            # Acumular totais do grupo
            for k in ['valor_fgts', 'valor_corrigido', 'valor_jam', 'valor_deposito_fgts', 'total']:
                if k in calc:
                    grupos[chave]['totais'][k] += calc[k]
            # Total exibido = depósito corrigido (sem somar JAM)
            grupos[chave]['totais']['total'] = grupos[chave]['totais']['valor_deposito_fgts']
        
        # Ordenar grupos
        if agrupamento == 'ano':
            grupos_ordenados = sorted(grupos.items(), key=lambda x: x[0])
        elif agrupamento == 'competencia':
            def parse_comp_key(key):
                try:
                    comp_part = key.split('|')[0]
                    return datetime.strptime(comp_part, '%m/%Y').date()
                except Exception:
                    return datetime(1900, 1, 1).date()
            grupos_ordenados = sorted(grupos.items(), key=lambda x: parse_comp_key(x[0]))
        else:  # funcionario/vinculo
            grupos_ordenados = sorted(grupos.items(), key=lambda x: grupos[x[0]]['label'])
        return grupos_ordenados

    def _compute_for(self, empresa, competencia_str, parcela_13, data_pagamento, funcionario=None, matricula=None, jam_state=None, config_juros=None):
        import time
        from django.db.models import Q
        from empresas.models_grupo import get_aliquota_fgts
        inicio_timestamp = time.time()
        inicio_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        try:
            if jam_state is None:
                jam_state = {}

            if parcela_13 in [0, '0', '', None]:
                parcela_13 = None

            competencia_norm = self.normalizar_competencia(competencia_str)

            # BLOQUEIO: Validar se data_pagamento está dentro do range de data_base disponível para a competência/tabela
            from indices.models import SupabaseIndice
            from indices.services.indice_service import IndiceFGTSService
            try:
                competencia_date = datetime.strptime(competencia_norm, '%m/%Y').date().replace(day=1)
            except ValueError:
                return None, None, 'Competência inválida. Use MM/YYYY.', jam_state, []

            tabela = int(IndiceFGTSService.determinar_tabela(competencia_date))
            from django.core.cache import cache as _cache
            _range_key = f'indice_fgts_range_{competencia_date.strftime("%Y%m")}_{tabela}'
            datas_base = _cache.get(_range_key)
            if datas_base is None:
                qs_indices = SupabaseIndice.objects.filter(competencia=competencia_date, tabela=tabela)
                datas_base = list(qs_indices.values_list('data_base', flat=True))
                _cache.set(_range_key, datas_base, timeout=86400)
            if not datas_base:
                return None, None, f'Nenhum índice cadastrado para competência {competencia_norm} (tabela {tabela}).', jam_state, []
            data_base_min = min(datas_base)
            data_base_max = max(datas_base)
            if data_pagamento < data_base_min or data_pagamento > data_base_max:
                msg = (
                    f'❌ Não é possível calcular para a data de pagamento {data_pagamento.strftime("%d/%m/%Y")}.'
                    f' O intervalo permitido para competência {competencia_norm} (tabela {tabela}) é de '
                    f'{data_base_min.strftime("%d/%m/%Y")} até {data_base_max.strftime("%d/%m/%Y")}.'
                    ' Geralmente, as datas vão do dia 21 ao dia 20 do mês seguinte.'
                )
                return None, None, msg, jam_state, []
            # ...existing code...
            # (todo o restante do método permanece igual)
        except Exception as e:
            return None, None, f'Erro inesperado ao processar relatório: {str(e)}', jam_state if jam_state else {}, []
        avisos = []
        
        # 🛡️ Verificar se há loop infinito
        try:
            loop_key = competencia_norm if parcela_13 is None else f"{competencia_norm}|{parcela_13}"
            self._verificar_loop(loop_key)
        except Exception as e:
            return None, None, str(e), jam_state, avisos
        
        try:
            competencia_date = datetime.strptime(competencia_norm, '%m/%Y').date().replace(day=1)
        except ValueError:
            return None, None, 'Competência inválida. Use MM/YYYY.', jam_state, avisos


        # Buscar lançamentos pela competência armazenada como string 'MM/YYYY'

        # Busca por competencia como string 'MM/YYYY'
        # Aceita competências armazenadas como MM/YYYY ou YYYY-MM
        comp_iso = None
        try:
            mes_norm, ano_norm = competencia_norm.split('/')
            comp_iso = f"{ano_norm}-{mes_norm.zfill(2)}"
        except Exception:
            comp_iso = None

        filtro_comp = Q(competencia=competencia_norm)
        if comp_iso:
            filtro_comp |= Q(competencia=comp_iso)

        lancs_qs = (Lancamento.objects
            .filter(empresa=empresa)
            .filter(filtro_comp)
            .filter(parcela_13=parcela_13, pago=False)
            .select_related('funcionario', 'vinculo', 'vinculo__tipo_vinculo')
            .prefetch_related('funcionario__vinculos')
            .order_by('funcionario_id', 'vinculo_id'))
        if funcionario:
            lancs_qs = lancs_qs.filter(funcionario=funcionario)
        if matricula:
            lancs_qs = lancs_qs.filter(vinculo__matricula__iexact=str(matricula).strip())

        # Filtrar por vínculo ativo na competência
        lancamentos_filtrados = []
        for l in lancs_qs:
            if getattr(l, 'vinculo_id', None):
                if l.vinculo and l.vinculo.empresa_id == empresa.pk and l.vinculo.is_ativo_em_competencia(competencia_norm):
                    lancamentos_filtrados.append(l)
                continue

            vinculos = getattr(l.funcionario, 'vinculos', None)
            if not vinculos:
                continue
            if any(v.empresa_id == empresa.pk and v.is_ativo_em_competencia(competencia_norm) for v in vinculos.all()):
                lancamentos_filtrados.append(l)

        # ⚡ Se não há lançamentos para esta competência, pular silenciosamente
        if not lancamentos_filtrados:
            return [], {k: Decimal('0') for k in ['valor_corrigido', 'valor_jam', 'total']}, None, jam_state, avisos

        # REGRA DE NEGÓCIO IMUTÁVEL: Buscar índice EXATO
        # competencia = competencia_date E data_base = data_pagamento E tabela automática
        # USAR APENAS IndiceFGTSService - NUNCA ALTERAR ESTA LÓGICA
        # Tabela é determinada AUTOMATICAMENTE: 6 (até 09/1989) ou 7 (10/1989+)
        indice_valor = IndiceFGTSService.buscar_indice(
            competencia=competencia_date,
            data_pagamento=data_pagamento
            # tabela determinada automaticamente pelo serviço
        )

        if indice_valor is None:
            # Diagnóstico: listar datas_base disponíveis para a mesma competência/tabela
            try:
                from indices.models import SupabaseIndice
                datas_disponiveis = list(
                    SupabaseIndice.objects.filter(
                        competencia=competencia_date,
                        tabela=tabela
                    ).order_by('data_base').values_list('data_base', flat=True)
                )
            except Exception:
                datas_disponiveis = []

            # ⚠️ AVISO: Índice não encontrado, pular a competência mas notificar o usuário com as datas disponíveis
            if datas_disponiveis:
                datas_str = ', '.join([d.strftime('%d/%m/%Y') for d in datas_disponiveis])
                aviso = (
                    f"⚠️ Nenhum índice FGTS encontrado para competência {competencia_norm} na data de pagamento {data_pagamento.strftime('%d/%m/%Y')} "
                    f"(tabela {tabela}). Datas disponíveis na indices_fgts: {datas_str}."
                )
            else:
                aviso = (
                    f"⚠️ Nenhum índice FGTS encontrado para competência {competencia_norm} na data de pagamento {data_pagamento.strftime('%d/%m/%Y')} "
                    f"(tabela {tabela})."
                )

            avisos.append(aviso)
            return [], {k: Decimal('0') for k in ['valor_corrigido', 'valor_jam', 'total']}, None, jam_state, avisos

        if config_juros:
            juros_tipo = config_juros['juros_tipo']
            juros_mensal = config_juros['juros_mensal']
            juros_diario = config_juros['juros_diario']
            multa_percent = config_juros['multa_percent']
        else:
            juros_tipo = get_config_str('JUROS_TIPO', 'MENSAL')
            juros_mensal = get_config_numeric('JUROS_MENSAL_PERCENT', Decimal('0.5'))
            juros_diario = get_config_numeric('JUROS_DIARIO_PERCENT', Decimal('0.033'))
            multa_percent = get_config_numeric('MULTA_PERCENT', Decimal('10.0'))

        resultados = []
        totais = {k: Decimal('0') for k in ['valor_fgts', 'valor_corrigido', 'valor_jam', 'valor_deposito_fgts', 'total']}
        meses_sem_coef_aviso: set[date] = set()
        coef_cache: dict[date, Decimal | None] = {}

        def _get_coef(comp_date: date) -> Decimal | None:
            comp_key = comp_date.replace(day=1)
            if comp_key not in coef_cache:
                coef_cache[comp_key] = buscar_coef_jam(comp_key)
            return coef_cache[comp_key]

        comp_display = competencia_norm
        if parcela_13 == 1:
            comp_display = f"{competencia_norm} (13º 1ª)"
        elif parcela_13 == 2:
            comp_display = f"{competencia_norm} (13º 2ª)"

        for l in lancamentos_filtrados:
            jam_key = None
            if getattr(l, 'vinculo_id', None):
                jam_key = f"vinc_{l.vinculo_id}"
            else:
                jam_key = f"func_{l.funcionario.pk}"

            if jam_key not in jam_state:
                jam_state[jam_key] = {'acumulado': Decimal('0.00')}

            # Ajuste de plano econômico legado (multiplica e divide conforme VB6)
            valor_fgts_ajustado, fator_mult, fator_div, fator_liquido = aplicar_plano_economico_legacy(
                l.valor_fgts,
                competencia_date,
            )

            valor_jam, _detalhes_jam, meses_sem_coef = calcular_jam_ate_pagamento(
                valor_fgts=valor_fgts_ajustado,
                competencia=competencia_date,
                data_pagamento=data_pagamento,
                coef_lookup=_get_coef,
            )

            if meses_sem_coef:
                meses_sem_coef_aviso.update(meses_sem_coef)

            jam_state[jam_key]['acumulado'] = jam_state[jam_key]['acumulado'] + valor_fgts_ajustado + valor_jam

            calc = calcular_fgts_atualizado(
                valor_fgts=valor_fgts_ajustado,
                competencia=competencia_date,
                pagamento=data_pagamento,
                indice=indice_valor,
                jam_coef=None,
                valor_jam_override=valor_jam,
                aplicar_plano_economico=False,
                fator_plano_info=(fator_mult, fator_div, fator_liquido),
                valor_fgts_base=l.base_fgts,
                aliquota=get_aliquota_fgts(l.vinculo),
                juros_tipo=juros_tipo,
                juros_mensal=juros_mensal,
                juros_diario=juros_diario,
                multa_percent=multa_percent,
            )
            resultados.append({
                'lancamento': l,
                'calc': calc,
                'competencia': competencia_norm,
                'parcela_13': parcela_13,
                'competencia_display': comp_display,
            })
            for k in totais.keys():
                if k in calc:
                    totais[k] += calc[k]

        if meses_sem_coef_aviso:
            meses_txt = ', '.join(sorted({m.strftime('%m/%Y') for m in meses_sem_coef_aviso}))
            avisos.append(
                f"⚠️ Coeficiente JAM ausente para: {meses_txt}. Cálculo desses meses considerado como 0."
            )

        return resultados, totais, None, jam_state, avisos

    def form_valid(self, form):
        import logging
        import time
        from datetime import datetime
        logger = logging.getLogger(__name__)

        # Inicializar métricas de tempo para todos os fluxos
        inicio_timestamp = time.time()
        inicio_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

        # Reset contadores de loop para cada nova requisição
        self.tempo_inicio = None
        self.competencias_processadas = {}

        try:
            empresa = form.cleaned_data['empresa']
            competencia_str = (form.cleaned_data.get('competencia') or '').strip()
            competencias_multi = (form.cleaned_data.get('competencias') or '').strip()
            # Normalizar competência única para o formato do banco
            if competencia_str:
                competencia_str = self.normalizar_competencia(competencia_str)
            # Normalizar múltiplas competências (uma por linha)
            if competencias_multi:
                competencias_multi = '\n'.join([
                    self.normalizar_competencia(c.strip())
                    for c in competencias_multi.splitlines() if c.strip()
                ])
            funcionario = form.cleaned_data.get('funcionario')
            matricula = (form.cleaned_data.get('matricula') or '').strip()
            agrupamento = form.cleaned_data.get('agrupamento', 'competencia')
            if form.cleaned_data['data_pagamento']:
                data_pagamento = form.cleaned_data['data_pagamento']
            else:
                from indices.services.indice_service import IndiceFGTSService
                data_pagamento = IndiceFGTSService.obter_ultima_data_base() or date.today()

            # Escopo multi-tenant: empresa deve estar autorizada
            if not is_empresa_allowed(self.request.user, empresa.codigo):
                return render(self.request, self.template_name, {
                    'form': form,
                    'erro': 'Empresa não permitida para este usuário.',
                    **feature_block_context('custom_reports', user=self.request.user, empresa=empresa),
                })

            allowed_report, motivo_bloqueio = can_use_feature('custom_reports', user=self.request.user, empresa=empresa)
            if not allowed_report:
                messages.error(
                    self.request,
                    motivo_bloqueio or 'Trial expirado e nenhum plano ativo. Assine um plano para gerar relatórios.'
                )
                contexto_bloqueio = feature_block_context('custom_reports', user=self.request.user, empresa=empresa)
                return render(self.request, self.template_name, {
                    'form': form,
                    'erro': contexto_bloqueio.get('feature_block_reason') or motivo_bloqueio,
                    **contexto_bloqueio,
                })

            resultados = []
            totais = {k: Decimal('0') for k in ['valor_fgts', 'valor_corrigido', 'valor_jam', 'valor_deposito_fgts', 'total']}
            avisos_total = []  # Coletar todos os avisos

            def format_comp_display(comp, parcela_13):
                if parcela_13 == 1:
                    return f"{comp} (13º 1ª)"
                if parcela_13 == 2:
                    return f"{comp} (13º 2ª)"
                return comp

            # Determinar lista de competências a processar (com parcela_13)
            if competencias_multi:
                competencias_list = [{'competencia': c.strip(), 'parcela_13': None} for c in competencias_multi.splitlines() if c.strip()]
            elif competencia_str:
                competencias_list = [{'competencia': competencia_str, 'parcela_13': None}]
            else:
                # Busca por competencia como string
                lancamentos_qs = Lancamento.objects.filter(
                    empresa=empresa,
                    pago=False
                )
                if funcionario:
                    lancamentos_qs = lancamentos_qs.filter(funcionario=funcionario)
                if matricula:
                    lancamentos_qs = lancamentos_qs.filter(vinculo__matricula__iexact=matricula)

                competencias_list = list(
                    lancamentos_qs.values('competencia', 'parcela_13')
                    .distinct()
                    .order_by('competencia', 'parcela_13')
                )

                if not competencias_list:
                    return render(self.request, self.template_name, {
                        'form': form,
                        'erro': 'Nenhum lançamento em aberto encontrado. Verifique se existem lançamentos com status "Não Pago".'
                    })

            # Ordena competências cronologicamente para respeitar o acumulado do JAM legado
            def _parse_comp(c: str):
                # Aceita 'MM/YYYY' ou 'YYYY-MM' como válido
                try:
                    if isinstance(c, str):
                        if '/' in c:
                            return datetime.strptime(c, '%m/%Y').date()
                        if '-' in c:
                            return datetime.strptime(c, '%Y-%m').date()
                    return None
                except Exception:
                    return None

            competencias_invalidas = [c['competencia'] for c in competencias_list if _parse_comp(c['competencia']) is None]
            competencias_list = [c for c in competencias_list if _parse_comp(c['competencia']) is not None]
            competencias_list.sort(key=lambda x: (_parse_comp(x['competencia']) or date(1900, 1, 1), x.get('parcela_13') or 0))

            # Limitar quantidade para evitar timeouts
            if len(competencias_list) > self.MAX_COMPETENCIAS:
                return render(self.request, self.template_name, {
                    'form': form,
                    'erro_limite': {
                        'tipo': 'competencias',
                        'total': len(competencias_list),
                        'limite': self.MAX_COMPETENCIAS,
                    },
                    **feature_block_context('custom_reports', user=self.request.user, empresa=empresa),
                })

            if not competencias_list:
                erro_msg = 'Nenhuma competência válida encontrada.'
                if competencias_invalidas:
                    erro_msg += (
                        f' Competências inválidas: {", ".join(competencias_invalidas[:5])}. '
                        'Aceitos: MM/YYYY ou YYYY-MM.'
                    )
                return render(self.request, self.template_name, {'form': form, 'erro': erro_msg})

            jam_state = {}
            total_lancamentos = 0
            competencias_com_erro = []

            # COUNT rápido para evitar processar volumes que causam timeout
            LIMITE_LANCAMENTOS = 2000
            filtro_count = dict(
                empresa=empresa,
                pago=False,
                competencia__in=[c['competencia'] for c in competencias_list],
            )
            if funcionario:
                filtro_count['funcionario'] = funcionario
            total_estimado = Lancamento.objects.filter(**filtro_count).count()

            LIMITE_SYNC_LANCAMENTOS = 1000
            LIMITE_SYNC_COMPETENCIAS = 6
            usar_async = (
                total_estimado > LIMITE_SYNC_LANCAMENTOS
                or len(competencias_list) > LIMITE_SYNC_COMPETENCIAS
            )

            if not usar_async and total_estimado > LIMITE_LANCAMENTOS:
                return render(self.request, self.template_name, {
                    'form': form,
                    'erro_limite': {
                        'tipo': 'lancamentos',
                        'total': total_estimado,
                        'limite': LIMITE_LANCAMENTOS,
                    },
                    **feature_block_context('custom_reports', user=self.request.user, empresa=empresa),
                })

            if usar_async:
                from .models_relatorio import RelatorioTask
                from .services.relatorio_service import processar_relatorio
                task = RelatorioTask.objects.create(
                    usuario=self.request.user,
                    empresa=empresa,
                    parametros_json={
                        'empresa_id': empresa.pk,
                        'funcionario_id': funcionario.pk if funcionario else None,
                        'matricula': matricula or '',
                        'competencias_list': competencias_list,
                        'agrupamento': agrupamento,
                        'data_pagamento': data_pagamento.isoformat(),
                        'competencias_display': [format_comp_display(c['competencia'], c.get('parcela_13')) for c in competencias_list],
                        'competencias_param': [f"{c['competencia']}|{c.get('parcela_13') or ''}" for c in competencias_list],
                        'competencia_primeira': competencias_list[0]['competencia'] if competencias_list else '',
                    },
                    total_lancamentos=total_estimado,
                )
                threading.Thread(target=processar_relatorio, args=(task.id,), daemon=True).start()
                return redirect('relatorio-task-status', pk=task.pk)

            # Carregar configurações uma vez para todos os _compute_for
            config_juros = {
                'juros_tipo': get_config_str('JUROS_TIPO', 'MENSAL'),
                'juros_mensal': get_config_numeric('JUROS_MENSAL_PERCENT', Decimal('0.5')),
                'juros_diario': get_config_numeric('JUROS_DIARIO_PERCENT', Decimal('0.033')),
                'multa_percent': get_config_numeric('MULTA_PERCENT', Decimal('10.0')),
            }

            for comp_data in competencias_list:
                comp = comp_data['competencia']
                parc = comp_data.get('parcela_13')
                res, tot, err, jam_state, avisos = self._compute_for(empresa, comp, parc, data_pagamento, funcionario, matricula or None, jam_state, config_juros=config_juros)
                # Se houver erro (ex: índice ausente), registrar aviso e seguir para próxima competência
                if err:
                    comp_display = f"{comp} (13º {parc})" if parc else comp
                    aviso_erro = f"⚠️ Competência {comp_display} pulada: {err}"
                    avisos_total.append(aviso_erro)
                    competencias_com_erro.append(comp)
                    continue
                # Coletar avisos
                if avisos:
                    avisos_total.extend(avisos)
                if res:
                    resultados.extend(res)
                    total_lancamentos += len(res)
                    # ⚠️ NÃO SOMAR AQUI - os subtotais serão calculados na agregação
            if not resultados:
                fim_timestamp = time.time()
                fim_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
                tempo_total = fim_timestamp - inicio_timestamp
                # Montar mensagem detalhada de filtros
                filtros_info = f"<br><strong>Filtros utilizados:</strong>"\
                    f"<br>Empresa: {empresa.nome} (ID: {empresa.pk})"\
                    f"<br>Funcionário: {funcionario.nome if funcionario else 'Todos'}"\
                    f"<br>Matrícula: {matricula or '—'}"\
                    f"<br>Competências: {', '.join([c['competencia'] for c in competencias_list])}"\
                    f"<br>Data de Pagamento: {data_pagamento.strftime('%d/%m/%Y')}"\
                    f"<br>Status: Não Pago (pago=False)"
                erro_msg = 'Nenhum lançamento encontrado com os filtros aplicados.' \
                    ' Verifique se há lançamentos com status "Não Pago" para as competências selecionadas.' \
                    + filtros_info
                return render(self.request, self.template_name, {
                    'form': form,
                    'erro': erro_msg,
                    'kpi_inicio': inicio_str,
                    'kpi_fim': fim_str,
                    'kpi_tempo': f'{tempo_total:.2f} segundos',
                    'kpi_lancamentos': total_lancamentos,
                    'kpi_competencias': len(competencias_list),
                })

            # Aplicar agrupamento
            resultados_agrupados = self._agrupar_resultados(resultados, agrupamento)
            
            # ✅ CORRIGIR: Recalcular totais gerais a partir dos grupos (evitar duplicação)
            totais = {k: Decimal('0') for k in ['valor_fgts', 'valor_corrigido', 'valor_jam', 'valor_deposito_fgts', 'total']}
            for chave, grupo_data in resultados_agrupados:
                for k in totais.keys():
                    totais[k] += grupo_data['totais'][k]

            competencias_display = [format_comp_display(c['competencia'], c.get('parcela_13')) for c in competencias_list]
            competencias_param = [f"{c['competencia']}|{c.get('parcela_13') or ''}" for c in competencias_list]
            competencia_primeira = competencias_list[0]['competencia'] if competencias_list else ''

            # Deduplica avisos (remove mensagens duplicadas)
            avisos_unicos = list(dict.fromkeys(avisos_total))  # Preserva ordem mantendo apenas primeira ocorrência

            fim_timestamp = time.time()
            fim_str = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            tempo_total = fim_timestamp - inicio_timestamp
            contexto = {
                'form': form,
                'empresa': empresa,
                'funcionario': funcionario,
                'matricula': matricula or None,
                'competencias': competencias_display,
                'competencias_param': competencias_param,
                'competencia_primeira': competencia_primeira,
                'data_pagamento': data_pagamento,
                'resultados': resultados,
                'resultados_agrupados': resultados_agrupados,
                'agrupamento': agrupamento,
                'totais': totais,
                'avisos': avisos_unicos,  # Usar avisos deduplificados
                'kpi_inicio': inicio_str,
                'kpi_fim': fim_str,
                'kpi_tempo': f'{tempo_total:.2f} segundos',
                'kpi_lancamentos': total_lancamentos,
                'kpi_competencias': len(competencias_list),
                'exibir_indice': self.request.session.get('exibir_indice', False),
                'exibir_jam': self.request.session.get('exibir_jam', True),
                'exibir_correcao': self.request.session.get('exibir_correcao', True),
            }
            contexto.update(feature_block_context('custom_reports', user=self.request.user, empresa=empresa))
            return render(self.request, self.template_name, contexto)
            
        except Exception as e:
            logger.error(f"🛑 Erro em RelatorioCompetenciaView.form_valid: {str(e)}")
            return render(self.request, self.template_name, {
                'form': form,
                'erro': f"🛑 Erro ao processar relatório: {str(e)}",
                **feature_block_context('custom_reports', user=self.request.user, empresa=form.cleaned_data.get('empresa')),
            })

@login_required
def relatorio_por_ids(request):
    from django.http import HttpResponse
    from django.shortcuts import render
    from collections import defaultdict
    from decimal import Decimal

    debug_detalhado = request.GET.get('debug', '') == '1'
    debug_lancamentos = []
    ids_str = request.POST.get('ids', '') or request.GET.get('ids', '')
    agrupamento = request.POST.get('agrupamento', '') or request.GET.get('agrupamento', 'competencia')
    if not ids_str:
        return HttpResponse('Nenhum lançamento selecionado.', status=400)
    try:
        ids = [int(id_str.strip()) for id_str in ids_str.split(',') if id_str.strip()]
    except ValueError:
        return HttpResponse('IDs inválidos.', status=400)
    if not ids:
        return HttpResponse('Nenhum lançamento selecionado.', status=400)

    # Buscar lançamentos pelos IDs e apenas não pagos
    lancamentos = Lancamento.objects.filter(id__in=ids, pago=False).select_related('empresa', 'funcionario', 'vinculo').prefetch_related('funcionario__vinculos')
    if not lancamentos.exists():
        return HttpResponse('Nenhum lançamento encontrado.', status=404)

    empresa_referencia = lancamentos.first().empresa if lancamentos else None
    allowed_report, motivo_bloqueio = can_use_feature('custom_reports', user=request.user, empresa=empresa_referencia)
    if not allowed_report:
        messages.error(request, motivo_bloqueio or 'Trial expirado e nenhum plano ativo. Assine um plano para gerar relatórios.')
        return redirect('lancamento-list')

    # Verificar permissões multi-tenant
    allowed_ids = get_allowed_empresa_ids(request.user)
    if allowed_ids is not None:
        lancamentos = lancamentos.filter(empresa__codigo__in=allowed_ids)
        if not lancamentos.exists():
            return HttpResponse('Você não tem permissão para acessar esses lançamentos.', status=403)

    # Acima de 1000 IDs autorizados → processamento assíncrono
    LIMITE_SYNC_IDS = 1000
    if len(ids) > LIMITE_SYNC_IDS:
        import threading
        from .models_relatorio import RelatorioTask
        from .services.relatorio_service import processar_relatorio_por_ids
        ids_autorizados = list(lancamentos.values_list('id', flat=True))
        task = RelatorioTask.objects.create(
            usuario=request.user,
            empresa=empresa_referencia,
            parametros_json={
                'ids': ids_autorizados,
                'agrupamento': agrupamento,
            },
            total_lancamentos=len(ids_autorizados),
        )
        threading.Thread(target=processar_relatorio_por_ids, args=(task.id,), daemon=True).start()
        return redirect('relatorio-task-status', pk=task.pk)

    # Filtrar por vínculo ativo na competência e não pago
    lancamentos_filtrados = []
    avisos_total = []
    empresa = empresa_referencia
    # Forçar a data_pagamento para a última data_base disponível (ignora qualquer data enviada)
    from indices.services.indice_service import IndiceFGTSService
    data_pagamento = IndiceFGTSService.obter_ultima_data_base() or date.today()

    from empresas.models_grupo import FuncionarioVinculo  # noqa: F401 (usado para is_ativo_em_competencia)
    import re

    for lanc in lancamentos:
        competencia = lanc.competencia
        empresa = lanc.empresa
        funcionario = lanc.funcionario
        vinculos = getattr(funcionario, 'vinculos', None)
        competencia_norm = competencia
        match = re.match(r'^(\d{2})/(\d{4})$', competencia)
        if match:
            mes, ano = match.groups()
            competencia_norm = f"{ano}-{mes}"
        elif re.match(r'^13/\d{4}$', competencia):
            ano = competencia[-4:]
            competencia_norm = f"{ano}-12"
        motivo = []
        vinculo_ativo = False
        if vinculos:
            for v in vinculos.all():
                empresa_id = getattr(empresa, 'id', None) or getattr(empresa, 'codigo', None)
                if (str(v.empresa_id) == str(empresa_id)) and v.is_ativo_em_competencia(competencia_norm):
                    vinculo_ativo = True
                    break
        if not vinculos:
            motivo.append('Sem vínculos cadastrados')
        elif vinculo_ativo:
            motivo.append('Vínculo ativo OK')
        else:
            motivo.append('Sem vínculo ativo na competência')
        if not lanc.pago:
            motivo.append('Não pago')
        else:
            motivo.append('Já pago')

        if (vinculo_ativo or not vinculos) and not lanc.pago:
            lancamentos_filtrados.append(lanc)
            status = 'Incluído'
        else:
            status = 'Excluído'

        if debug_detalhado:
            debug_lancamentos.append({
                'id': lanc.id,
                'colaborador': getattr(funcionario, 'nome', str(funcionario)),
                'competencia': competencia,
                'empresa': getattr(empresa, 'nome', str(empresa)),
                'motivo': ', '.join(motivo),
                'status': status
            })

    # Usar o mesmo cálculo do relatório padrão (índice, correção, JAM)
    view = RelatorioCompetenciaView()
    view.request = request

    resultados = []
    totais = {k: Decimal('0') for k in ['valor_fgts', 'valor_corrigido', 'valor_jam', 'valor_deposito_fgts', 'total']}
    jam_state = {}
    avisos_total = []
    ids_set = {l.id for l in lancamentos_filtrados}

    grupos = defaultdict(list)
    for lanc in lancamentos_filtrados:
        if empresa is None:
            empresa = lanc.empresa
        comp_norm = view.normalizar_competencia(lanc.competencia)
        key = (lanc.empresa_id, comp_norm, lanc.parcela_13 or 0)
        grupos[key].append(lanc)

    import time as _time
    TIMEOUT_IDS_SEGUNDOS = 50
    inicio_ids = _time.time()

    for (empresa_id, comp_norm, parcela_13), _lancs in grupos.items():
        if _time.time() - inicio_ids > TIMEOUT_IDS_SEGUNDOS:
            aviso_timeout = (
                f'⏱️ Processamento interrompido após {TIMEOUT_IDS_SEGUNDOS}s para evitar timeout. '
                f'Foram calculadas {len(resultados)} entradas de {len(ids_set)} selecionadas. '
                'Para ver todos os dados, filtre por funcionário específico ou use exportação CSV/PDF.'
            )
            avisos_total.append(aviso_timeout)
            break
        empresa_grupo = _lancs[0].empresa
        res, _tot, err, jam_state, avisos = view._compute_for(
            empresa_grupo,
            comp_norm,
            parcela_13,
            data_pagamento,
            funcionario=None,
            jam_state=jam_state,
        )
        if avisos:
            avisos_total.extend(avisos)
        if err:
            continue

        # Manter apenas os lançamentos selecionados
        res_filtrados = [r for r in res if r.get('lancamento') and r['lancamento'].id in ids_set]
        resultados.extend(res_filtrados)

    for r in resultados:
        for k in totais.keys():
            totais[k] += r['calc'][k]
    totais['total'] = totais['valor_deposito_fgts']

    if not resultados:
        return HttpResponse('Nenhum resultado calculado para os lançamentos selecionados.', status=404)

    # Agrupar resultados por competência
    def parse_comp_key(key):
        try:
            return datetime.strptime(key[0], '%m/%Y').date()
        except Exception:
            return datetime(1900, 1, 1).date()

    resultados_agrupados = view._agrupar_resultados(resultados, agrupamento)

    def _format_comp_display(comp, parcela):
        if parcela == 1:
            return f"{comp} (13º 1ª)"
        if parcela == 2:
            return f"{comp} (13º 2ª)"
        return comp

    competencias_display = [_format_comp_display(k[1], k[2]) for k in grupos.keys()]
    competencias_param = [f"{k[1]}|{k[2] or ''}" for k in grupos.keys()]

    # Deduplica avisos (remove mensagens duplicadas)
    avisos_unicos = list(dict.fromkeys(avisos_total))  # Preserva ordem mantendo apenas primeira ocorrência

    contexto = {
        'empresa': empresa,
        'competencias': competencias_display,
        'competencias_param': competencias_param,
        'data_pagamento': data_pagamento,
        'resultados': resultados,
        'resultados_agrupados': resultados_agrupados,
        'agrupamento': agrupamento,
        'totais': totais,
        'avisos': avisos_unicos,  # Usar avisos deduplificados
        'from_selection': True,
        'ids_param': ','.join([str(i) for i in ids]),
        'debug_lancamentos': debug_lancamentos if debug_detalhado else None,
        'exibir_indice': request.session.get('exibir_indice', False),
        'exibir_jam': request.session.get('exibir_jam', True),
        'exibir_correcao': request.session.get('exibir_correcao', True),
    }
    return render(request, 'lancamentos/relatorio_competencia.html', contexto)

@login_required
def export_relatorio_competencia_csv(request):
    # Mantido para compatibilidade — redireciona para XLSX
    return export_relatorio_competencia_xlsx(request)


def _render_xlsx_relatorio(resultados_agrupados, empresa, totais, agrupamento):
    """Gera HttpResponse com XLSX a partir de resultados já agrupados."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Relatório FGTS'

    headers = ['Empresa', 'Competência', 'Funcionário', 'Matrícula', 'ID Vínculo',
               'Empresa do Vínculo', 'Admissão', 'Demissão', 'Base FGTS', 'FGTS',
               'Índice', 'Correção Monetária', 'Total']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    def _grupo_label(label):
        if agrupamento == 'funcionario':
            return f"Funcionário: {label}"
        if agrupamento == 'ano':
            return str(label)
        return f"Competência: {label}"

    empresa_nome = empresa.nome if empresa else ''

    for _chave, grupo in resultados_agrupados:
        ws.append([])
        ws.append([_grupo_label(grupo.get('label', ''))])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        for item in grupo['items']:
            l = item['lancamento']
            c = item['calc']
            comp_out = item.get('competencia_display', item.get('competencia', ''))
            func = l.funcionario
            vinculo = l.vinculo
            if vinculo is None:
                try:
                    vinculo = func.vinculos.filter(empresa=l.empresa).order_by('-data_admissao').first()
                except Exception:
                    vinculo = None
            empresa_vinculo = (
                getattr(getattr(vinculo, 'empresa', None), 'nome', None)
                or l.empresa.nome
            )
            data_admissao = vinculo.data_admissao.strftime('%d/%m/%Y') if vinculo and getattr(vinculo, 'data_admissao', None) else ''
            data_demissao = vinculo.data_demissao.strftime('%d/%m/%Y') if vinculo and getattr(vinculo, 'data_demissao', None) else ''
            try:
                base = float(l.base_fgts) if l.base_fgts is not None else ''
                fgts = float(c.get('valor_fgts', l.valor_fgts))
                correcao = float(c['valor_corrigido'])
                total = float(c['total'])
            except Exception:
                base = str(l.base_fgts) if l.base_fgts is not None else ''
                fgts = str(c.get('valor_fgts', l.valor_fgts))
                correcao = str(c['valor_corrigido'])
                total = str(c['total'])
            ws.append([
                empresa_nome or l.empresa.nome,
                comp_out,
                func.nome,
                getattr(vinculo, 'matricula', '') or '',
                getattr(vinculo, 'pk', '') or '',
                empresa_vinculo,
                data_admissao,
                data_demissao,
                base,
                fgts,
                str(c.get('indice', '')),
                correcao,
                total,
            ])

    ws.append([])
    try:
        totais_row = ['Totais', '', '', '', '', '',
                      float(totais['valor_fgts']), '', '',
                      float(totais['valor_corrigido']), '', '',
                      float(totais['total'])]
    except Exception:
        totais_row = ['Totais', '', '', '', '', '',
                      str(totais['valor_fgts']), '', '',
                      str(totais['valor_corrigido']), '', '',
                      str(totais['total'])]
    ws.append(totais_row)
    for cell in ws[ws.max_row]:
        if cell.value:
            cell.font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from django.http import HttpResponse
    resp = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = 'attachment; filename="relatorio_fgts.xlsx"'
    return resp


@login_required
def export_relatorio_competencia_xlsx(request):
    from django.http import HttpResponse
    from collections import defaultdict
    import urllib.parse

    # Fast path: relatório gerado via task assíncrona — usa dados já calculados
    task_id = request.GET.get('task_id')
    if task_id:
        from .models_relatorio import RelatorioTask
        from .services.relatorio_service import deserializar_resultado
        task = get_object_or_404(RelatorioTask, pk=task_id, usuario=request.user)
        if task.status != 'done':
            return HttpResponse('Relatório ainda não concluído.', status=400)
        empresa_obj = Empresa.objects.filter(pk=task.resultado_json.get('empresa_id')).first()
        allowed, motivo = can_use_feature('pdf_export', user=request.user, empresa=empresa_obj)
        if not allowed:
            return HttpResponseForbidden(motivo or 'Sem permissão para exportar.')
        agrupamento_task = request.GET.get('agrupamento') or task.resultado_json.get('agrupamento', 'competencia')
        contexto = deserializar_resultado(task.resultado_json, agrupamento_task)
        return _render_xlsx_relatorio(
            contexto['resultados_agrupados'],
            contexto.get('empresa'),
            contexto['totais'],
            agrupamento_task,
        )

    empresa_id = request.GET.get('empresa')
    competencias_multi = request.GET.get('competencias', '')
    competencia_unica = request.GET.get('competencia', '')
    funcionario_id = request.GET.get('funcionario')
    matricula = (request.GET.get('matricula') or '').strip()
    data_pagamento_str = request.GET.get('data_pagamento')
    agrupamento = request.GET.get('agrupamento', 'competencia')
    ids_str = request.GET.get('ids', '').strip()

    # Decodificar competências que podem vir URL-encoded
    competencias_multi = urllib.parse.unquote(competencias_multi)

    empresa = Empresa.objects.get(pk=empresa_id)
    allowed_csv, motivo_bloqueio = can_use_feature('pdf_export', user=request.user, empresa=empresa)
    if not allowed_csv:
        return HttpResponseForbidden(motivo_bloqueio or 'Seu plano não permite exportar relatórios (CSV/PDF).')
    if data_pagamento_str:
        data_pagamento = datetime.strptime(data_pagamento_str, '%Y-%m-%d').date()
    else:
        from indices.services.indice_service import IndiceFGTSService
        data_pagamento = IndiceFGTSService.obter_ultima_data_base() or date.today()
    funcionario = Funcionario.objects.get(pk=funcionario_id) if funcionario_id else None

    view = RelatorioCompetenciaView()
    view.request = request

    if ids_str:
        try:
            ids = [int(id_str.strip()) for id_str in ids_str.split(',') if id_str.strip()]
        except ValueError:
            return HttpResponse('IDs inválidos.', status=400)
        if not ids:
            return HttpResponse('Nenhum lançamento selecionado.', status=400)

        from indices.services.indice_service import IndiceFGTSService
        data_pagamento = IndiceFGTSService.obter_ultima_data_base() or date.today()

        lancamentos = Lancamento.objects.filter(id__in=ids, pago=False).select_related('empresa', 'funcionario', 'vinculo')
        if not lancamentos.exists():
            return HttpResponse('Nenhum lançamento encontrado.', status=404)

        allowed_ids = get_allowed_empresa_ids(request.user)
        if allowed_ids is not None:
            lancamentos = lancamentos.filter(empresa__codigo__in=allowed_ids)
            if not lancamentos.exists():
                return HttpResponse('Você não tem permissão para acessar esses lançamentos.', status=403)

        lancamentos_filtrados = []
        empresa = None
        from empresas.models_grupo import FuncionarioVinculo  # noqa: F401 (usado para is_ativo_em_competencia)
        import re

        for lanc in lancamentos:
            competencia = lanc.competencia
            empresa = lanc.empresa
            funcionario = lanc.funcionario
            vinculos = getattr(funcionario, 'vinculos', None)
            competencia_norm = competencia
            match = re.match(r'^(\d{2})/(\d{4})$', competencia)
            if match:
                mes, ano = match.groups()
                competencia_norm = f"{ano}-{mes}"
            elif re.match(r'^13/\d{4}$', competencia):
                ano = competencia[-4:]
                competencia_norm = f"{ano}-12"

            vinculo_ativo = False
            if vinculos:
                for v in vinculos.all():
                    empresa_id_comp = getattr(empresa, 'id', None) or getattr(empresa, 'codigo', None)
                    if (str(v.empresa_id) == str(empresa_id_comp)) and v.is_ativo_em_competencia(competencia_norm):
                        vinculo_ativo = True
                        break

            if (vinculo_ativo or not vinculos) and not lanc.pago:
                lancamentos_filtrados.append(lanc)

        if not lancamentos_filtrados:
            return HttpResponse('Nenhum lançamento encontrado para os filtros aplicados.', status=404)

        resultados = []
        totais = {k: Decimal('0') for k in ['valor_fgts', 'valor_corrigido', 'valor_jam', 'valor_deposito_fgts', 'total']}
        jam_state = {}
        ids_set = {l.id for l in lancamentos_filtrados}

        grupos = defaultdict(list)
        for lanc in lancamentos_filtrados:
            if empresa is None:
                empresa = lanc.empresa
            comp_norm = view.normalizar_competencia(lanc.competencia)
            key = (lanc.empresa_id, comp_norm, lanc.parcela_13 or 0)
            grupos[key].append(lanc)

        for (_empresa_id, comp_norm, parcela_13), _lancs in grupos.items():
            empresa_grupo = _lancs[0].empresa
            res, _tot, err, jam_state, _avisos = view._compute_for(
                empresa_grupo,
                comp_norm,
                parcela_13,
                data_pagamento,
                funcionario=None,
                jam_state=jam_state,
            )
            if err:
                continue

            res_filtrados = [r for r in res if r.get('lancamento') and r['lancamento'].id in ids_set]
            resultados.extend(res_filtrados)

        for r in resultados:
            for k in totais.keys():
                totais[k] += r['calc'][k]
        totais['total'] = totais['valor_deposito_fgts']

        if not resultados:
            return HttpResponse('Nenhum resultado calculado para os lançamentos selecionados.', status=404)

        agrupamento = 'funcionario'
        resultados_agrupados = view._agrupar_resultados(resultados, agrupamento)
        return _render_xlsx_relatorio(resultados_agrupados, empresa, totais, agrupamento)
    
    # Parse competências (separar por \n ou %0A) mantendo parcela_13 quando informada
    competencias_raw = [c.strip() for c in competencias_multi.replace('%0A', '\n').split('\n') if c.strip()]
    competencias_list = []
    for entry in competencias_raw:
        comp_str = entry
        parc_val = None
        if '|' in entry:
            comp_str, parc_part = entry.split('|', 1)
            parc_part = parc_part.strip()
            if parc_part:
                try:
                    parc_val = int(parc_part)
                except ValueError:
                    parc_val = None
        if parc_val == 0:
            parc_val = None
        competencias_list.append({'competencia': comp_str, 'parcela_13': parc_val})
    
    # Se não houver competências especificadas, buscar todas em aberto
    if not competencias_list:
        lancamentos_qs = Lancamento.objects.filter(empresa=empresa, pago=False)
        if funcionario:
            lancamentos_qs = lancamentos_qs.filter(funcionario=funcionario)
        if matricula:
            lancamentos_qs = lancamentos_qs.filter(vinculo__matricula__iexact=matricula)
        
        competencias_unicas = (
            lancamentos_qs.values('competencia', 'parcela_13')
            .distinct()
            .order_by('competencia', 'parcela_13')
        )
        competencias_list = [
            {'competencia': item['competencia'], 'parcela_13': item['parcela_13']}
            for item in competencias_unicas
        ]
    
    # Ordenar por competência
    def _parse_comp(c_dict):
        try:
            comp_str = c_dict['competencia']
            return datetime.strptime(comp_str, '%m/%Y').date()
        except Exception:
            return date(1900, 1, 1)
    
    competencias_list.sort(key=lambda x: (_parse_comp(x), x.get('parcela_13') or 0))

    resultados = []
    totais = {k: Decimal('0') for k in ['valor_fgts', 'valor_corrigido', 'valor_jam', 'valor_deposito_fgts', 'total']}
    jam_state = {}
    first_error = None
    first_error = None
    
    for comp_dict in competencias_list:
        comp = comp_dict['competencia']
        parcela_13 = comp_dict.get('parcela_13')
        
        res, tot, err, jam_state, _avisos = view._compute_for(empresa, comp, parcela_13, data_pagamento, funcionario, matricula or None, jam_state)
        if err:
            if first_error is None:
                first_error = err
            continue
        
        resultados.extend(res)
        
        for k in totais.keys():
            totais[k] += tot.get(k, Decimal('0'))


    if not resultados:
        mensagem = first_error or 'Nenhum lançamento encontrado para os filtros aplicados.'
        resp = HttpResponse(mensagem, status=400 if first_error else 404)
        resp['Content-Type'] = 'text/plain; charset=utf-8'
        return resp

    resultados_agrupados = view._agrupar_resultados(resultados, agrupamento)
    return _render_xlsx_relatorio(resultados_agrupados, empresa, totais, agrupamento)

@login_required
def export_relatorio_competencia_pdf(request):
    # Fast path: relatório gerado via task assíncrona — usa IDs já autorizados da task
    task_id = request.GET.get('task_id')
    if task_id:
        from django.http import HttpResponse
        from .models_relatorio import RelatorioTask
        task = get_object_or_404(RelatorioTask, pk=task_id, usuario=request.user)
        if task.status != 'done':
            return HttpResponse('Relatório ainda não concluído.', status=400)
        empresa_obj = Empresa.objects.filter(pk=task.resultado_json.get('empresa_id')).first()
        allowed, motivo = can_use_feature('pdf_export', user=request.user, empresa=empresa_obj)
        if not allowed:
            return HttpResponseForbidden(motivo or 'Sem permissão para exportar.')
        # Injeta ids_param no GET mutable para reutilizar o caminho de IDs abaixo
        request.GET = request.GET.copy()
        ids_param = task.resultado_json.get('ids_param', '')
        request.GET['ids'] = ids_param
        request.GET['empresa'] = str(task.resultado_json.get('empresa_id', ''))
        request.GET.setdefault('agrupamento', task.resultado_json.get('agrupamento', 'competencia'))

    # Função utilitária para converter competência em date
    def competencia_str_to_date(competencia):
        if isinstance(competencia, str):
            if '/' in competencia:
                mes, ano = competencia.split('/')
                return datetime(int(ano), int(mes), 1).date()
            elif '-' in competencia:
                ano, mes = competencia.split('-')
                return datetime(int(ano), int(mes), 1).date()
        elif isinstance(competencia, date):
            return competencia
        raise ValueError("Formato de competência inválido")

    from django.http import HttpResponse
    from collections import defaultdict
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak
    import urllib.parse

    empresa_id = request.GET.get('empresa')
    competencias_multi = request.GET.get('competencias', '')
    competencia_unica = request.GET.get('competencia', '')
    funcionario_id = request.GET.get('funcionario')
    matricula = (request.GET.get('matricula') or '').strip()
    data_pagamento_str = request.GET.get('data_pagamento')
    exibir_indice = request.session.get('exibir_indice', False)
    exibir_jam = request.session.get('exibir_jam', True)
    exibir_correcao = request.session.get('exibir_correcao', True)
    agrupamento = request.GET.get('agrupamento', 'competencia')
    ids_str = request.GET.get('ids', '').strip()

    # Decodificar competências que podem vir URL-encoded com %0A para \n
    competencias_multi = urllib.parse.unquote(competencias_multi)

    empresa = Empresa.objects.get(pk=empresa_id)
    allowed_pdf, motivo_bloqueio = can_use_feature('pdf_export', user=request.user, empresa=empresa)
    if not allowed_pdf:
        return HttpResponseForbidden(motivo_bloqueio or 'Seu plano não permite exportar relatórios (CSV/PDF).')
    if data_pagamento_str:
        data_pagamento = datetime.strptime(data_pagamento_str, '%Y-%m-%d').date()
    else:
        from indices.services.indice_service import IndiceFGTSService
        data_pagamento = IndiceFGTSService.obter_ultima_data_base() or date.today()
    funcionario = Funcionario.objects.get(pk=funcionario_id) if funcionario_id else None

    view = RelatorioCompetenciaView()
    view.request = request  # Necessário para EmpresaScopeMixin

    if ids_str:
        try:
            ids = [int(id_str.strip()) for id_str in ids_str.split(',') if id_str.strip()]
        except ValueError:
            return HttpResponse('IDs inválidos.', status=400)
        if not ids:
            return HttpResponse('Nenhum lançamento selecionado.', status=400)

        from indices.services.indice_service import IndiceFGTSService
        data_pagamento = IndiceFGTSService.obter_ultima_data_base() or date.today()

        lancamentos = Lancamento.objects.filter(id__in=ids, pago=False).select_related('empresa', 'funcionario', 'vinculo')
        if not lancamentos.exists():
            return HttpResponse('Nenhum lançamento encontrado.', status=404)

        allowed_ids = get_allowed_empresa_ids(request.user)
        if allowed_ids is not None:
            lancamentos = lancamentos.filter(empresa__codigo__in=allowed_ids)
            if not lancamentos.exists():
                return HttpResponse('Você não tem permissão para acessar esses lançamentos.', status=403)

        lancamentos_filtrados = []
        empresa = None
        from empresas.models_grupo import FuncionarioVinculo  # noqa: F401 (usado para is_ativo_em_competencia)
        import re

        for lanc in lancamentos:
            competencia = lanc.competencia
            empresa = lanc.empresa
            funcionario = lanc.funcionario
            vinculos = getattr(funcionario, 'vinculos', None)
            competencia_norm = competencia
            match = re.match(r'^(\d{2})/(\d{4})$', competencia)
            if match:
                mes, ano = match.groups()
                competencia_norm = f"{ano}-{mes}"
            elif re.match(r'^13/\d{4}$', competencia):
                ano = competencia[-4:]
                competencia_norm = f"{ano}-12"

            vinculo_ativo = False
            if vinculos:
                for v in vinculos.all():
                    empresa_id_comp = getattr(empresa, 'id', None) or getattr(empresa, 'codigo', None)
                    if (str(v.empresa_id) == str(empresa_id_comp)) and v.is_ativo_em_competencia(competencia_norm):
                        vinculo_ativo = True
                        break

            if (vinculo_ativo or not vinculos) and not lanc.pago:
                lancamentos_filtrados.append(lanc)

        if not lancamentos_filtrados:
            return HttpResponse('Nenhum lançamento encontrado para os filtros aplicados.', status=404)

        resultados = []
        totais = {k: Decimal('0') for k in ['valor_fgts', 'valor_corrigido', 'valor_jam', 'valor_deposito_fgts', 'total']}
        jam_state = {}
        ids_set = {l.id for l in lancamentos_filtrados}

        grupos = defaultdict(list)
        for lanc in lancamentos_filtrados:
            if empresa is None:
                empresa = lanc.empresa
            comp_norm = view.normalizar_competencia(lanc.competencia)
            key = (lanc.empresa_id, comp_norm, lanc.parcela_13 or 0)
            grupos[key].append(lanc)

        for (_empresa_id, comp_norm, parcela_13), _lancs in grupos.items():
            empresa_grupo = _lancs[0].empresa
            res, _tot, err, jam_state, _avisos = view._compute_for(
                empresa_grupo,
                comp_norm,
                parcela_13,
                data_pagamento,
                funcionario=None,
                jam_state=jam_state,
            )
            if err or not res:
                continue

            res_filtrados = [r for r in res if r.get('lancamento') and r['lancamento'].id in ids_set]
            resultados.extend(res_filtrados)

        for r in resultados:
            for k in totais.keys():
                totais[k] += r['calc'][k]
        totais['total'] = totais['valor_deposito_fgts']

        if not resultados:
            return HttpResponse('Nenhum resultado calculado para os lançamentos selecionados.', status=404)

        # Agrupar resultados por vínculo (quando disponível) para evitar ambiguidade
        from collections import defaultdict
        grupos_func = defaultdict(list)
        for item in resultados:
            l = item['lancamento']
            if getattr(l, 'vinculo_id', None):
                grupos_func[f"vinc_{l.vinculo_id}"].append(item)
            else:
                grupos_func[f"func_{l.funcionario_id}"].append(item)

        def _format_money(valor):
            try:
                return f"{valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            except Exception:
                return str(valor)

        def _format_indice(valor):
            try:
                return f"{valor:.9f}".replace('.', ',')
            except Exception:
                return str(valor)

        def _parse_comp(comp_str):
            try:
                return datetime.strptime(comp_str, '%m/%Y').date()
            except Exception:
                return date(1900, 1, 1)

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=15 * mm,
            rightMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = styles['Heading2']
        normal_style = styles['BodyText']
        small_style = styles['BodyText']
        small_style.fontSize = 8
        small_style.leading = 10

        now = datetime.now()
        usuario_label = getattr(request.user, 'username', 'Usuário')
        empresa_codigo = getattr(empresa, 'codigo_exibicao', None) or getattr(empresa, 'codigo', None)
        empresa_label = f"{empresa_codigo} {empresa.nome}" if empresa_codigo else empresa.nome
        cnpj_label = empresa.cnpj or ''

        story = []

        header_table = Table(
            [
                [
                    Paragraph("LISTAGEM DO RECOLHIMENTO FGTS", title_style),
                    Paragraph("Sistema FGTS em Atraso", normal_style),
                    Paragraph("FGTS WEB", normal_style),
                ],
                [
                    Paragraph(f"USUÁRIO: {usuario_label}", small_style),
                    Paragraph(f"{now.strftime('%d/%m/%Y')} - {now.strftime('%H:%M')}", small_style),
                    Paragraph("Página 1", small_style),
                ],
                [
                    Paragraph(empresa_label, normal_style),
                    Paragraph(f"CNPJ: {cnpj_label}", normal_style),
                    "",
                ],
            ],
            colWidths=[80*mm, 60*mm, 35*mm],
            hAlign='LEFT',
        )
        header_table.setStyle(
            TableStyle([
                ('LINEBELOW', (0, 2), (-1, 2), 0.75, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (2, 0), (2, 1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ])
        )
        story.append(header_table)
        story.append(Spacer(1, 6))

        for idx, (func_id, itens) in enumerate(grupos_func.items()):
            itens.sort(key=lambda x: _parse_comp(x.get('competencia') or x['lancamento'].competencia))
            funcionario = itens[0]['lancamento'].funcionario
            vinculo = itens[0]['lancamento'].vinculo

            data_adm = vinculo.data_admissao.strftime('%d/%m/%Y') if vinculo and vinculo.data_admissao else ''
            data_nasc = funcionario.data_nascimento.strftime('%d/%m/%Y') if funcionario.data_nascimento else ''
            ctps = funcionario.carteira_profissional or ''
            serie = funcionario.serie_carteira or ''
            cbo = funcionario.cbo or ''
            pis = funcionario.pis or ''
            matricula_label = (vinculo.matricula if vinculo and vinculo.matricula else '')

            funcionario_header = Table(
                [
                    [
                        Paragraph(f"{funcionario.id}  {funcionario.nome}" + (f" (Matrícula {matricula_label})" if matricula_label else ""), normal_style),
                        Paragraph(f"Data Adm {data_adm}", normal_style),
                        Paragraph(f"C.B.O. {cbo}", normal_style),
                        Paragraph(f"Data Nasc {data_nasc}", normal_style),
                    ],
                    [
                        Paragraph(f"PIS {pis}", normal_style),
                        Paragraph(f"C.T.P.S. {ctps} / {serie}", normal_style),
                        "",
                        "",
                    ],
                ],
                colWidths=[70*mm, 35*mm, 40*mm, 35*mm],
                hAlign='LEFT',
            )
            funcionario_header.setStyle(
                TableStyle([
                    ('LINEBELOW', (0, 1), (-1, 1), 0.75, colors.black),
                ])
            )
            story.append(funcionario_header)
            story.append(Spacer(1, 4))

            # JAM é um assunto distinto da correção/depósito FGTS.
            # No PDF, exibimos o JAM separado e não o somamos ao depósito.
            pdf_header = ["Comp.", "13º", "Base FGTS", "Valor FGTS"]
            if exibir_correcao:
                pdf_header.append("Correção")
            if exibir_jam:
                pdf_header.append("JAM")
            pdf_header.append("Total")
            if exibir_indice:
                pdf_header.append("Índice CEF")
            table_data = [pdf_header]

            total_fgts = Decimal('0')
            total_deposito_sem_jam = Decimal('0')
            total_jam = Decimal('0')
            total_recolher = Decimal('0')

            for item in itens:
                l = item['lancamento']
                c = item['calc']
                comp_label = l.competencia
                parcela_13 = l.parcela_13 or 0
                col_13 = "13º 1ª" if parcela_13 == 1 else "13º 2ª" if parcela_13 == 2 else ""
                valor_fgts = c.get('valor_fgts', l.valor_fgts)
                valor_deposito_sem_jam = c.get('valor_deposito_fgts')
                valor_correcao = c.get('valor_corrigido')
                valor_jam = c.get('valor_jam', Decimal('0'))
                valor_total = c.get('total')
                if valor_deposito_sem_jam is None:
                    # Fallback seguro (não deve acontecer se calc estiver completo)
                    valor_deposito_sem_jam = (Decimal(str(valor_fgts)) + Decimal(str(valor_correcao or 0))).quantize(Decimal('0.01'))
                if valor_correcao is None:
                    valor_correcao = (Decimal(str(valor_deposito_sem_jam)) - Decimal(str(valor_fgts))).quantize(Decimal('0.01'))
                if valor_total is None:
                    valor_total = (Decimal(str(valor_deposito_sem_jam)) + Decimal(str(valor_jam))).quantize(Decimal('0.01'))
                indice = c.get('indice', '')

                total_fgts += Decimal(str(valor_fgts))
                total_deposito_sem_jam += Decimal(str(valor_deposito_sem_jam))
                total_jam += Decimal(str(valor_jam))
                total_recolher += Decimal(str(valor_total))

                row = [
                    comp_label,
                    col_13,
                    _format_money(l.base_fgts),
                    _format_money(valor_fgts),
                ]
                if exibir_correcao:
                    row.append(_format_money(valor_correcao))
                if exibir_jam:
                    row.append(_format_money(valor_jam))
                row.append(_format_money(valor_total))
                if exibir_indice:
                    row.append(_format_indice(indice) if indice != '' else '')
                table_data.append(row)

            col_widths_pdf = [18*mm, 16*mm, 27*mm, 23*mm]
            if exibir_correcao:
                col_widths_pdf.append(22*mm)
            if exibir_jam:
                col_widths_pdf.append(18*mm)
            col_widths_pdf.append(22*mm)
            if exibir_indice:
                col_widths_pdf.append(30*mm)
            table = Table(
                table_data,
                colWidths=col_widths_pdf,
                hAlign='LEFT',
                repeatRows=1,
            )
            table.setStyle(
                TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f5f5f5')),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#999999')),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
                ])
            )
            story.append(table)

            box_rows = [
                    [
                        Paragraph("DATA DO CÁLCULO", normal_style),
                        Paragraph(f"{data_pagamento.strftime('%d/%m/%Y')}", normal_style),
                        "",
                        Paragraph("Total do F.G.T.S. Mensal", normal_style),
                        Paragraph(_format_money(total_fgts), normal_style),
                    ],
                ]
            if exibir_correcao:
                box_rows.append([
                        "",
                        "",
                        "",
                        Paragraph("Total Depósito (sem JAM)", normal_style),
                        Paragraph(_format_money(total_deposito_sem_jam), normal_style),
                ])
            if exibir_jam:
                box_rows.append([
                        "",
                        "",
                        "",
                        Paragraph("Total JAM (juros)", normal_style),
                        Paragraph(_format_money(total_jam), normal_style),
                ])
            box_rows.append([
                        "",
                        "",
                        "",
                        Paragraph("Valor da Multa Rescisória", normal_style),
                        Paragraph(_format_money(Decimal('0.00')), normal_style),
                    ])
            box_rows.append([
                        "",
                        "",
                        "",
                        Paragraph("TOTAL A RECOLHER", styles['Heading4']),
                        Paragraph(_format_money(total_recolher), styles['Heading4']),
                    ])
            box_table = Table(
                box_rows,
                colWidths=[35*mm, 30*mm, 10*mm, 55*mm, 30*mm],
                hAlign='LEFT',
            )
            num_box_rows = len(box_rows)
            box_style = [
                    ('GRID', (0, 0), (-1, -1), 0.75, colors.black),
                    ('SPAN', (0, 0), (1, 0)),
                    ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
            for i in range(1, num_box_rows):
                box_style.append(('SPAN', (0, i), (1, i)))
            box_table.setStyle(TableStyle(box_style))
            story.append(Spacer(1, 6))
            story.append(box_table)

            if idx < len(grupos_func) - 1:
                story.append(PageBreak())

        doc.build(story)
        pdf = buffer.getvalue()
        buffer.close()

        resp = HttpResponse(content_type='application/pdf')
        resp['Content-Disposition'] = 'attachment; filename="relatorio_fgts.pdf"'
        resp.write(pdf)
        return resp
    
    # Parse competências como dicionários com parcela_13
    # Separar por \n ou %0A (pode vir URL-encoded)
    competencias_raw = [c.strip() for c in competencias_multi.replace('%0A', '\n').split('\n') if c.strip()]

    if competencia_unica and not competencias_raw:
        competencias_raw = [competencia_unica.strip()]

    competencias_list = []
    for entry in competencias_raw:
        comp_str = entry
        parc_val = None
        if '|' in entry:
            comp_str, parc_part = entry.split('|', 1)
            parc_part = parc_part.strip()
            if parc_part:
                try:
                    parc_val = int(parc_part)
                except ValueError:
                    parc_val = None
        if parc_val == 0:
            parc_val = None
        competencias_list.append({'competencia': comp_str, 'parcela_13': parc_val})

    # Deduplicar competências (competencia, parcela_13)
    dedup = {}
    for c in competencias_list:
        key = (c['competencia'], c.get('parcela_13'))
        dedup[key] = c
    competencias_list = list(dedup.values())

    # Se não houver competências especificadas, não exportar para evitar incluir dados não solicitados
    if not competencias_list:
        resp = HttpResponse('Nenhuma competência informada para exportação.', status=400)
        resp['Content-Type'] = 'text/plain; charset=utf-8'
        return resp

    # Ordenar competências para respeitar o acumulado de JAM
    def _parse_comp(c: str):
        try:
            return datetime.strptime(c, '%m/%Y').date()
        except Exception:
            return None

    competencias_list = [c for c in competencias_list if _parse_comp(c['competencia']) is not None]
    competencias_list.sort(key=lambda x: (_parse_comp(x['competencia']) or date(1900, 1, 1), x.get('parcela_13') or 0))

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = styles['Title']
    subtitle_style = styles['Heading3']
    normal_style = styles['BodyText']
    story = [
        Paragraph(f"Relatório FGTS — {empresa.nome}", title_style),
        Paragraph(f"Pagamento: {data_pagamento.strftime('%d/%m/%Y')}", normal_style),
        Spacer(1, 8),
    ]

    resultados = []
    totais = {k: Decimal('0') for k in ['valor_fgts', 'valor_corrigido', 'valor_jam', 'valor_deposito_fgts', 'total']}
    jam_state = {}
    first_error = None

    for comp_dict in competencias_list:
        comp = comp_dict['competencia']
        parcela_13 = comp_dict.get('parcela_13')

        res, tot, err, jam_state, _avisos = view._compute_for(empresa, comp, parcela_13, data_pagamento, funcionario, jam_state)

        if err or not res:
            if err and first_error is None:
                first_error = err
            continue

        resultados.extend(res)

        for k in totais.keys():
            totais[k] += tot.get(k, Decimal('0'))

    totais['total'] = totais['valor_deposito_fgts']

    if not resultados:
        mensagem = first_error or 'Nenhum lançamento encontrado para os filtros aplicados.'
        resp = HttpResponse(mensagem, status=400 if first_error else 404)
        resp['Content-Type'] = 'text/plain; charset=utf-8'
        return resp

    resultados_agrupados = view._agrupar_resultados(resultados, agrupamento)

    def _grupo_label(label):
        if agrupamento == 'funcionario':
            return f"Funcionário: {label}"
        if agrupamento == 'ano':
            return f"{label}"
        return f"Competência: {label}"

    from reportlab.platypus import PageBreak
    for idx, (_chave, grupo) in enumerate(resultados_agrupados):
        story.append(Spacer(1, 6))
        story.append(Paragraph(_grupo_label(grupo.get('label')), subtitle_style))

        table_data = [
            [
                "Competência",
                "Funcionário",
                "Demissão",
                "Base FGTS",
                "FGTS Valor",
                "Correção",
                "JAM",
                "Total",
            ]
        ]

        for item in grupo['items']:
            l = item['lancamento']
            c = item['calc']
            comp_label = item.get('competencia_display', item.get('competencia'))
            funcionario = l.funcionario
            comp_date = competencia_str_to_date(l.competencia)
            vinculo = funcionario.vinculos.filter(
                empresa=l.empresa,
                data_admissao__lte=comp_date,
            ).order_by('-data_admissao').first()
            if vinculo and vinculo.data_demissao and vinculo.data_demissao < comp_date:
                vinculo = None
            empresa_vinculo = vinculo.empresa.nome if vinculo else l.empresa.nome
            data_admissao = vinculo.data_admissao.strftime('%d/%m/%Y') if vinculo and vinculo.data_admissao else ''
            data_demissao = vinculo.data_demissao.strftime('%d/%m/%Y') if vinculo and vinculo.data_demissao else ''

            # Envolver nome do funcionário em Paragraph para permitir quebra de linha
            nome_style = ParagraphStyle(
                'NomeFuncionario',
                parent=getSampleStyleSheet()['Normal'],
                fontSize=9,
                leading=11,
            )
            nome_paragraph = Paragraph(funcionario.nome, nome_style)

            table_data.append([
                comp_label,
                nome_paragraph,  # Usar Paragraph em vez de string simples
                data_demissao,
                f"{l.base_fgts}",
                f"{c.get('valor_fgts', l.valor_fgts)}",
                f"{c['valor_corrigido']}",
                f"{c.get('valor_jam', Decimal('0.00'))}",
                f"{c.get('total') or (c.get('valor_deposito_fgts', Decimal('0.00')) + c.get('valor_jam', Decimal('0.00')))}",
            ])

        # Ajustar colWidths para caber em 170mm (A4 útil)
        # Ajustar colWidths para 8 colunas: aumentada coluna Funcionário para permitir quebra de texto
        table = Table(
            table_data,
            colWidths=[22*mm, 50*mm, 20*mm, 18*mm, 18*mm, 18*mm, 16*mm, 18*mm],
            hAlign='LEFT',
            repeatRows=1,
            splitByRow=1,
        )
        table.setStyle(
            TableStyle(
                [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f5f5f5')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#333333')),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('ALIGN', (5, 1), (-1, -1), 'RIGHT'),
                    ('ALIGN', (0, 0), (4, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Alinhar ao topo para permitir quebra de linhas
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cccccc')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fbfbfb')]),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]
            )
        )

        story.append(table)

    story.append(Spacer(1, 10))
    story.append(Paragraph("Totais", subtitle_style))
    total_recolher = (totais['valor_deposito_fgts'] + totais.get('valor_jam', Decimal('0'))).quantize(Decimal('0.01'))
    totais_table = Table(
        [
            [
                "Valor sem juros",
                "Correção",
                "Depósito (sem JAM)",
                "JAM",
                "Total a recolher",
            ],
            [
                f"{totais['valor_fgts']}",
                f"{totais['valor_corrigido']}",
                f"{totais['valor_deposito_fgts']}",
                f"{totais.get('valor_jam', Decimal('0.00'))}",
                f"{total_recolher}",
            ],
        ],
        colWidths=[28 * mm, 26 * mm, 34 * mm, 22 * mm, 32 * mm],
        hAlign='LEFT',
    )
    totais_table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#eaeaea')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (0, 1), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#cccccc')),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(totais_table)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    resp = HttpResponse(content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="relatorio_fgts.pdf"'
    resp.write(pdf)
    return resp


def export_sefip(request):
    """Exporta arquivo SEFIP.RE seguindo a mesma lógica do sistema legado.

    Parâmetros via GET:
    - empresa: ID da empresa
    - competencia: MM/YYYY
    - funcionario_de: ID inicial do funcionário
    - funcionario_ate: ID final do funcionário
    """
    from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden

    empresa_id = request.GET.get('empresa')
    competencia = request.GET.get('competencia')
    func_de = request.GET.get('funcionario_de')
    func_ate = request.GET.get('funcionario_ate')

    if not all([empresa_id, competencia, func_de, func_ate]):
        return HttpResponseBadRequest('Parâmetros obrigatórios: empresa, competencia, funcionario_de, funcionario_ate')

    try:
        empresa = Empresa.objects.get(pk=empresa_id)
    except Empresa.DoesNotExist:
        return HttpResponseBadRequest('Empresa inválida')

    # Escopo multi-tenant: checar permissão da empresa
    if not is_empresa_allowed(request.user, empresa.codigo):
        return HttpResponseForbidden('Empresa não permitida para este usuário.')

    try:
        func_de_id = int(func_de)
        func_ate_id = int(func_ate)
    except ValueError:
        return HttpResponseBadRequest('IDs de funcionário inválidos')

    filtros = SefipFilters(
        empresa=empresa,
        competencia=competencia,
        funcionario_de=func_de_id,
        funcionario_ate=func_ate_id,
    )

    conteudo = gerar_sefip_conteudo(filtros)

    response = HttpResponse(conteudo, content_type='text/plain; charset=iso-8859-1')
    # Mesmo nome de arquivo do legado
    response['Content-Disposition'] = 'attachment; filename="SEFIP.RE"'
    return response


@login_required
def sefip_export_view(request):
    """Tela exclusiva para gerar o SEFIP.RE conforme script legado."""
    if request.method == 'POST':
        form = SefipExportForm(request.POST, user=request.user)
        if form.is_valid():
            empresa = form.cleaned_data['empresa']
            if not is_empresa_allowed(request.user, empresa.codigo):
                return HttpResponseForbidden('Empresa não permitida para este usuário.')
            if not (request.user.is_staff or request.user.is_superuser) and not empresa_tem_recurso(empresa, 'gerar_sefip'):
                messages.warning(request, 'Recurso SEFIP.RE não habilitado para esta empresa.')
                return render(request, 'lancamentos/sefip_export.html', {'form': form, 'feature_blocked': True})

            filtros = SefipLegacyFilters(
                empresa=empresa,
                competencia=form.cleaned_data['competencia'],
                funcionario_de_id=form.cleaned_data['funcionario_de'].id,
                funcionario_ate_id=form.cleaned_data['funcionario_ate'].id,
            )

            try:
                conteudo = gerar_sefip_legacy(filtros)
            except SefipExportError as exc:
                messages.warning(request, str(exc))
            else:
                response = HttpResponse(conteudo, content_type='text/plain; charset=iso-8859-1')
                response['Content-Disposition'] = 'attachment; filename="SEFIP.RE"'
                return response
    else:
        form = SefipExportForm(user=request.user)

    return render(request, 'lancamentos/sefip_export.html', {'form': form})


@login_required
def download_memoria_calculo(request):
    """Gera e baixa a memória de cálculo em formato .txt"""
    from django.http import HttpResponse
    
    empresa_id = request.GET.get('empresa')
    funcionario_id = request.GET.get('funcionario')
    vinculo_id = request.GET.get('vinculo')
    competencia_str = request.GET.get('competencia')
    data_pagamento_str = request.GET.get('data_pagamento')
    parcela_13_str = request.GET.get('parcela_13')

    if not all([empresa_id, funcionario_id, competencia_str, data_pagamento_str]):
        return HttpResponse('Parâmetros incompletos', status=400)

    empresa = Empresa.objects.get(pk=empresa_id)

    # Verificação multi-tenant: usuário só acessa empresas permitidas
    if not is_empresa_allowed(request.user, empresa.codigo):
        return HttpResponse('Acesso negado a esta empresa.', status=403)

    funcionario = Funcionario.objects.get(pk=funcionario_id)
    data_pagamento = datetime.strptime(data_pagamento_str, '%Y-%m-%d').date()
    competencia_date = datetime.strptime(competencia_str, '%m/%Y').date().replace(day=1)

    # Parsear parcela_13 (None = competência normal, 1 = 13° 1ª, 2 = 13° 2ª)
    parcela_13 = int(parcela_13_str) if parcela_13_str and parcela_13_str.isdigit() else None

    # Busca o lançamento (vínculo-first para evitar ambiguidade)
    base_qs = Lancamento.objects.filter(
        empresa=empresa,
        funcionario=funcionario,
        competencia=competencia_str,
        parcela_13=parcela_13
    )
    if vinculo_id:
        base_qs = base_qs.filter(vinculo_id=vinculo_id)

    if not vinculo_id and base_qs.count() > 1:
        return HttpResponse('Lançamento ambíguo: informe o VÍNCULO (ID) ou MATRÍCULA para baixar a memória de cálculo.', status=400)

    lancamento = base_qs.select_related('vinculo__tipo_vinculo').first()
    
    if not lancamento:
        return HttpResponse('Lançamento não encontrado', status=404)
    
    # Busca índice
    indice_valor = IndiceFGTSService.buscar_indice(
        competencia=competencia_date,
        data_pagamento=data_pagamento
    )
    
    if indice_valor is None:
        return HttpResponse(
            f'Índice FGTS não encontrado para a competência {competencia_str} '
            f'na data de pagamento {data_pagamento.strftime("%d/%m/%Y")}. '
            'O download da memória de cálculo não pode ser gerado sem o índice correto.',
            status=400
        )
    
    # Ajuste plano econômico legado (multiplica e divide conforme VB6)
    valor_fgts_ajustado, fator_mult, fator_div, fator_liquido = aplicar_plano_economico_legacy(
        lancamento.valor_fgts,
        competencia_date,
    )

    # JAM composto até a data de pagamento
    valor_jam, _detalhes_jam, meses_sem_coef = calcular_jam_ate_pagamento(
        valor_fgts=valor_fgts_ajustado,
        competencia=competencia_date,
        data_pagamento=data_pagamento,
    )

    # Usar a mesma função do relatório para garantir valores idênticos
    from empresas.models_grupo import get_aliquota_fgts
    aliquota_vinculo = get_aliquota_fgts(lancamento.vinculo if lancamento.vinculo_id else None)
    calc = calcular_fgts_atualizado(
        valor_fgts=valor_fgts_ajustado,
        competencia=competencia_date,
        pagamento=data_pagamento,
        indice=indice_valor,
        jam_coef=None,
        valor_jam_override=valor_jam,
        aplicar_plano_economico=False,
        fator_plano_info=(fator_mult, fator_div, fator_liquido),
        valor_fgts_base=lancamento.base_fgts,
        aliquota=aliquota_vinculo,
    )

    valor_deposito_fgts = calc['valor_deposito_fgts']
    valor_corrigido = calc['valor_corrigido']
    total = calc['total']
    
    # Formata data de admissão para competência
    data_admissao_mes = funcionario.data_admissao.strftime('%m/%Y')
    
    # Gera memória de cálculo
    memoria = gerar_memoria_calculo(
        funcionario_nome=funcionario.nome,
        funcionario_cpf=funcionario.cpf,
        data_admissao=funcionario.data_admissao,
        valor_fgts=valor_fgts_ajustado,
        competencia_str=competencia_str,
        data_pagamento=data_pagamento,
        indice=indice_valor,
        valor_jam=valor_jam,
        valor_corrigido=valor_corrigido,
        total=total,
        data_admissao_mes=data_admissao_mes,
        salario_colaborador=lancamento.base_fgts,
        valor_deposito_fgts=valor_deposito_fgts,
        fator_plano_economico=fator_liquido,
        fator_plano_mult=fator_mult,
        fator_plano_div=fator_div,
        aliquota=aliquota_vinculo,
    )
    
    # Retorna arquivo para download
    response = HttpResponse(memoria, content_type='text/plain; charset=utf-8')
    filename = f"memoria_calculo_{funcionario.nome.replace(' ', '_')}_{competencia_str.replace('/', '_')}.txt"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


class LancamentoDownloadTemplateView(LoginRequiredMixin, View):
    """View para download do template XLSX de importação de lançamentos"""
    
    def get(self, request):
        try:
            # Gerar arquivo template
            xlsx_bytes = LancamentoImportService.generate_template_xlsx()
            
            # Retornar como download
            response = HttpResponse(
                xlsx_bytes,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="template_lancamentos_fgts.xlsx"'
            return response
            
        except Exception as e:
            messages.error(request, f'❌ Erro ao gerar template: {str(e)}')
            return redirect('lancamento-list')


class LancamentoImportView(LoginRequiredMixin, EmpresaScopeMixin, View):
    """View para importação de lançamentos via XLSX"""
    template_name = 'lancamentos/lancamento_import.html'
    
    def get(self, request, *args, **kwargs):
        """Renderizar página de importação"""
        # Listar empresas permitidas
        empresa_ids = get_allowed_empresa_ids(request.user)
        if empresa_ids is None:
            empresas = Empresa.objects.all()
        else:
            empresas = Empresa.objects.filter(codigo__in=empresa_ids)
        empresas = empresas.order_by('nome')
        
        context = {
            'empresas': empresas,
        }
        return render(request, self.template_name, context)
    
    def post(self, request, *args, **kwargs):
        # Validar arquivo
        if 'file' not in request.FILES:
            messages.error(request, '❌ Nenhum arquivo foi enviado.')
            return redirect('lancamento-import')

        file = request.FILES['file']

        # Validar extensão
        if not file.name.endswith('.xlsx'):
            messages.error(request, '❌ Apenas arquivos .xlsx são permitidos.')
            return redirect('lancamento-import')

        # Empresa selecionada é opcional quando o XLSX traz a coluna EMPRESA por linha.
        empresa = None
        empresa_codigo = request.POST.get('empresa')
        if empresa_codigo:
            try:
                empresa = Empresa.objects.get(codigo=empresa_codigo)
            except Empresa.DoesNotExist:
                messages.error(request, '❌ Empresa não encontrada.')
                return redirect('lancamento-import')

            # Validar permissões (quando empresa foi selecionada)
            if not is_empresa_allowed(request.user, empresa.codigo):
                return HttpResponseForbidden('Você não tem permissão para importar lançamentos para esta empresa.')

        # Ler opções de cálculo do formulário
        recalcular_fgts = request.POST.get('recalcular_fgts', 'recalcular') != 'manter'
        aplicar_jam = request.POST.get('aplicar_jam') in ('on', '1', 'true', 'True')
        extrato_analitico = request.POST.get('extrato_analitico') in ('on', '1', 'true', 'True')
        data_referencia_jam_raw = request.POST.get('data_referencia_jam', '').strip()
        data_referencia_jam = None
        if data_referencia_jam_raw:
            try:
                from datetime import date as _date
                data_referencia_jam = datetime.strptime(data_referencia_jam_raw, '%Y-%m-%d').date()
            except Exception:
                pass

        # Criar registro com status='preview' e processar amostra síncrona
        importacao = ImportacaoLancamento.objects.create(
            usuario=request.user,
            empresa=empresa,
            arquivo=file,
            nome_arquivo=file.name,
            status='preview',
            recalcular_fgts=recalcular_fgts,
            aplicar_jam=aplicar_jam,
            data_referencia_jam=data_referencia_jam,
            extrato_analitico=extrato_analitico,
        )

        try:
            with open(importacao.arquivo.path, 'rb') as f:
                preview = LancamentoImportService.preview_lancamentos_from_file(
                    f, empresa, request.user,
                    recalcular_fgts=recalcular_fgts,
                    aplicar_jam=aplicar_jam,
                    data_referencia_jam=data_referencia_jam,
                )
            importacao.preview_resultado = preview
            importacao.save(update_fields=['preview_resultado', 'atualizado_em'])
        except Exception as exc:
            importacao.delete()
            messages.error(request, str(exc))
            return redirect('lancamento-import')

        return redirect('lancamento-import-preview', pk=importacao.pk)


class LancamentoImportStatusView(LoginRequiredMixin, View):
    """Página de acompanhamento de importação assíncrona."""
    template_name = 'lancamentos/lancamento_import_status.html'

    def get(self, request, pk):
        importacao = get_object_or_404(ImportacaoLancamento, pk=pk, usuario=request.user)
        return render(request, self.template_name, {'importacao': importacao})


@login_required
def lancamento_import_status_json(request, pk):
    """Endpoint JSON para polling do status da importação."""
    importacao = get_object_or_404(ImportacaoLancamento, pk=pk, usuario=request.user)
    return JsonResponse({
        'status': importacao.status,
        'resultado': importacao.resultado_json,
        'erro': importacao.mensagem_erro,
        'linhas_total': importacao.linhas_total,
        'linhas_processadas': importacao.linhas_processadas,
    })


class RelatorioTaskStatusView(LoginRequiredMixin, View):
    """Página de acompanhamento de relatório assíncrono."""
    template_name = 'lancamentos/relatorio_task_status.html'

    def get(self, request, pk):
        from .models_relatorio import RelatorioTask
        task = get_object_or_404(RelatorioTask, pk=pk, usuario=request.user)
        return render(request, self.template_name, {'task': task})


@login_required
def relatorio_task_status_json(request, pk):
    """Endpoint JSON para polling do status do relatório assíncrono."""
    from .models_relatorio import RelatorioTask
    task = get_object_or_404(RelatorioTask, pk=pk, usuario=request.user)

    tipo = (task.parametros_json or {}).get('tipo', '')
    if tipo == 'posicao':
        redirect_url = reverse_lazy('relatorio-posicao-resultado', kwargs={'pk': pk})
    else:
        redirect_url = reverse_lazy('relatorio-task-resultado', kwargs={'pk': pk})

    return JsonResponse({
        'status': task.status,
        'erro': task.mensagem_erro,
        'total_lancamentos': task.total_lancamentos,
        'redirect_url': str(redirect_url),
    })


class RelatorioTaskResultadoView(LoginRequiredMixin, View):
    """Exibe o relatório calculado em background, renderizado a partir de resultado_json."""
    template_name = 'lancamentos/relatorio_competencia.html'

    def get(self, request, pk):
        from .models_relatorio import RelatorioTask
        from .services.relatorio_service import deserializar_resultado
        from billing.services.features import feature_block_context
        task = get_object_or_404(RelatorioTask, pk=pk, usuario=request.user)
        if task.status != 'done':
            return redirect('relatorio-task-status', pk=pk)
        agrupamento_override = request.GET.get('agrupamento') or None
        contexto = deserializar_resultado(task.resultado_json, agrupamento_override)
        empresa = contexto.get('empresa')
        contexto.update(feature_block_context('custom_reports', user=request.user, empresa=empresa))
        contexto['exibir_indice'] = request.session.get('exibir_indice', False)
        contexto['exibir_jam'] = request.session.get('exibir_jam', True)
        contexto['exibir_correcao'] = request.session.get('exibir_correcao', True)
        contexto['task_id'] = pk
        return render(request, self.template_name, contexto)


class LancamentoImportDownloadRelatorioView(LoginRequiredMixin, View):
    """Gera e baixa o relatório XLSX de uma importação de lançamentos."""

    def get(self, request, pk):
        from .services.import_report_service import gerar_relatorio_lancamentos
        importacao = get_object_or_404(ImportacaoLancamento, pk=pk, usuario=request.user)
        if importacao.status != 'done':
            return HttpResponse('Relatório disponível apenas após o processamento.', status=400)
        xlsx_bytes = gerar_relatorio_lancamentos(importacao)
        response = HttpResponse(
            xlsx_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="relatorio_lancamentos_{pk}.xlsx"'
        )
        return response


class LancamentoImportPreviewView(LoginRequiredMixin, View):
    """Página de pré-visualização da amostra antes de confirmar o import."""

    def get(self, request, pk):
        importacao = get_object_or_404(ImportacaoLancamento, pk=pk, usuario=request.user, status='preview')
        confirm_form = ImportacaoConfirmacaoForm()
        return render(request, 'lancamentos/lancamento_import_preview.html', {
            'importacao': importacao,
            'confirm_form': confirm_form,
        })


class LancamentoImportConfirmView(LoginRequiredMixin, View):
    """Confirma o import e dispara o processamento assíncrono."""

    def post(self, request, pk):
        importacao = get_object_or_404(ImportacaoLancamento, pk=pk, usuario=request.user, status='preview')

        confirm_form = ImportacaoConfirmacaoForm(request.POST)
        if not confirm_form.is_valid():
            messages.error(request, '❌ Você precisa aceitar a responsabilidade antes de confirmar a importação.')
            return render(request, 'lancamentos/lancamento_import_preview.html', {
                'importacao': importacao,
                'confirm_form': confirm_form,
            })

        # Montar texto dos termos exibidos (rastreabilidade legal)
        from django.utils import timezone as tz
        opcao_fgts = 'RECALCULAR (8% da base)' if importacao.recalcular_fgts else 'MANTER valor do arquivo'
        opcao_jam = (
            f'APLICAR até {importacao.data_referencia_jam or "hoje"}'
            if importacao.aplicar_jam else 'NÃO APLICAR'
        )
        texto_termos = (
            f"Importação confirmada em {tz.now().strftime('%d/%m/%Y %H:%M:%S')} "
            f"pelo usuário {request.user.username} (ID {request.user.pk}).\n"
            f"Arquivo: {importacao.nome_arquivo}\n"
            f"Empresa: {importacao.empresa}\n"
            f"Opção FGTS: {opcao_fgts}\n"
            f"Opção JAM: {opcao_jam}\n"
            f"O usuário declarou-se responsável pela exatidão e adequação dos dados importados."
        )

        # Salvar registro de responsabilidade
        ip = (request.META.get('HTTP_X_FORWARDED_FOR') or '').split(',')[0].strip() or request.META.get('REMOTE_ADDR')
        ImportacaoResponsabilidade.objects.create(
            importacao=importacao,
            usuario=request.user,
            recalcular_fgts_escolha=importacao.recalcular_fgts,
            aplicar_jam_escolha=importacao.aplicar_jam,
            data_referencia_jam_escolha=importacao.data_referencia_jam,
            aceite_responsabilidade=True,
            texto_termos=texto_termos,
            ip_address=ip or None,
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
        )

        # Audit log genérico
        try:
            from audit_logs.models import AuditLog
            AuditLog.objects.create(
                user=request.user,
                action='IMPORT',
                module='lancamentos',
                view_name='LancamentoImportConfirmView',
                url_path=request.path,
                ip_address=ip or None,
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:500],
                method='POST',
                status_code=302,
                object_id=importacao.pk,
                object_repr=f"Importação #{importacao.pk}: {importacao.nome_arquivo}",
                description=(
                    f"Importação confirmada. "
                    f"recalcular_fgts={importacao.recalcular_fgts}, "
                    f"aplicar_jam={importacao.aplicar_jam}"
                ),
                new_values={
                    'recalcular_fgts': importacao.recalcular_fgts,
                    'aplicar_jam': importacao.aplicar_jam,
                    'data_referencia_jam': str(importacao.data_referencia_jam or ''),
                    'aceite_responsabilidade': True,
                },
            )
        except Exception:
            pass  # audit log é não-crítico

        importacao.status = 'pending'
        importacao.save(update_fields=['status', 'atualizado_em'])
        threading.Thread(target=_process_importacao, args=(importacao.id,), daemon=True).start()
        return redirect('lancamento-import-status', pk=importacao.pk)


class LegacyImportView(LoginRequiredMixin, FormView):
    """View para importar dados históricos do sistema legado (VB6)"""
    form_class = LegacyImportForm
    template_name = 'lancamentos/legacy_import.html'
    success_url = reverse_lazy('legacy-import')
    
    def get_form_kwargs(self):
        """Passa o usuário ao formulário"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        """Processa o arquivo CSV selecionado"""
        from .services.legacy_importer import LegacyDataImporter
        from django.core.files.storage import default_storage
        import tempfile
        import os
        
        try:
            # Obtém dados do formulário
            csv_file = form.cleaned_data['csv_file']
            import_type = form.cleaned_data['import_type']
            empresa = form.cleaned_data.get('empresa')
            skip_duplicates = form.cleaned_data.get('skip_duplicates', True)
            
            # Valida permissões de empresa
            if import_type in ['funcionarios', 'lancamentos'] and empresa:
                if not is_empresa_allowed(self.request.user, empresa.codigo):
                    messages.error(self.request, '❌ Você não tem permissão para importar dados nesta empresa.')
                    return self.form_invalid(form)
            
            # Salva arquivo temporário
            with tempfile.NamedTemporaryFile(mode='wb', suffix='.csv', delete=False) as tmp_file:
                for chunk in csv_file.chunks():
                    tmp_file.write(chunk)
                tmp_path = tmp_file.name
            
            try:
                # Instancia o importador
                importer = LegacyDataImporter()
                
                # Processa a importação conforme o tipo
                registros_criados = 0
                erros = []
                avisos = []
                
                if import_type == 'empresas':
                    registros_criados, erros = importer.importar_empresas(tmp_path)
                
                elif import_type == 'funcionarios' and empresa:
                    registros_criados, erros = importer.importar_funcionarios(tmp_path, empresa_id=empresa.pk)
                
                elif import_type == 'lancamentos' and empresa:
                    registros_criados, erros = importer.importar_lancamentos(tmp_path, empresa_id=empresa.pk)
                
                # Obtém relatório completo
                relatorio = importer.relatorio()
                
                # Armazena relatório na sessão para exibição
                self.request.session['last_import_report'] = {
                    'import_type': import_type,
                    'linhas_processadas': relatorio.get('linhas_processadas', 0),
                    'registros_criados': registros_criados,
                    'registros_duplicados': relatorio.get('registros_duplicados', 0),
                    'erros': erros[:20],  # Limita a 20 erros para exibição
                    'avisos': relatorio.get('avisos', [])[:20],
                    'total_erros': len(erros),
                    'total_avisos': len(relatorio.get('avisos', [])),
                }
                
                # Mensagem de sucesso
                tipo_desc = {
                    'empresas': 'empresas',
                    'funcionarios': 'funcionários',
                    'lancamentos': 'lançamentos'
                }.get(import_type, 'registros')
                
                mensagem = f'✅ Importação de {tipo_desc} realizada com sucesso! '
                mensagem += f'{registros_criados} registros criados'
                
                if relatorio.get('registros_duplicados', 0) > 0:
                    mensagem += f', {relatorio.get("registros_duplicados")} duplicados ignorados'
                
                if erros:
                    mensagem += f', {len(erros)} erros'
                
                messages.success(self.request, mensagem + '.')
                
                # Redireciona para página de resultado
                return redirect('legacy-import-result')
            
            finally:
                # Remove arquivo temporário
                try:
                    os.unlink(tmp_path)
                except:
                    pass
        
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erro na importação legada: {str(e)}", exc_info=True)
            messages.error(self.request, f'❌ Erro ao processar importação: {str(e)}')
            return self.form_invalid(form)
    
    def get_context_data(self, **kwargs):
        """Adiciona dados do último relatório ao contexto"""
        context = super().get_context_data(**kwargs)
        context['last_import_report'] = self.request.session.get('last_import_report')
        context['page_title'] = 'Importar Dados Legados'
        context['page_description'] = 'Importe dados históricos do sistema legado (VB6) para o novo sistema'
        return context


class LegacyImportResultView(LoginRequiredMixin, View):
    """Exibe resultado detalhado da importação legada"""
    template_name = 'lancamentos/legacy_import_result.html'
    
    def get(self, request, *args, **kwargs):
        """Exibe o relatório da importação"""
        from django.shortcuts import render
        
        relatorio = request.session.get('last_import_report')
        
        if not relatorio:
            messages.warning(request, '⚠️ Nenhuma importação anterior encontrada.')
            return redirect('legacy-import')
        
        context = {
            'relatorio': relatorio,
            'page_title': 'Resultado da Importação',
            'success': len(relatorio.get('erros', [])) == 0,
        }
        
        return render(request, self.template_name, context)


# ===== CONFERÊNCIA DE LANÇAMENTOS =====

class ConferenciaListView(LoginRequiredMixin, EmpresaScopeMixin, ListView):
    """Lista lançamentos para conferência"""
    model = ConferenciaLancamento
    template_name = 'lancamentos/conferencia_list.html'
    context_object_name = 'conferencias'
    paginate_by = 50
    
    def get_queryset(self):
        """Retorna conferências filtradas"""
        empresa_id = self.kwargs.get('empresa_id')
        empresa = get_object_or_404(Empresa, pk=empresa_id)
        
        # Verifica permissão
        if not is_empresa_allowed(self.request.user, empresa.codigo):
            return ConferenciaLancamento.objects.none()
        
        queryset = ConferenciaLancamento.objects.filter(
            lancamento__empresa=empresa
        ).select_related(
            'lancamento__funcionario',
            'lancamento__empresa',
            'conferido_por'
        ).order_by('-criado_em')
        
        # Filtros
        status_filtro = self.request.GET.get('status', 'PENDENTE')
        competencia = self.request.GET.get('competencia')
        funcionario_id = self.request.GET.get('funcionario')
        
        if status_filtro and status_filtro != 'TODOS':
            queryset = queryset.filter(status=status_filtro)
        
        if competencia:
            queryset = queryset.filter(lancamento__competencia=competencia)
        
        if funcionario_id:
            queryset = queryset.filter(lancamento__funcionario_id=funcionario_id)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        empresa_id = self.kwargs.get('empresa_id')
        empresa = get_object_or_404(Empresa, pk=empresa_id)
        
        # Formulário de filtros
        context['filtro_form'] = FiltroConferenciaForm(
            self.request.GET or None,
            empresa=empresa
        )
        
        # Relatório de conferências
        competencia = self.request.GET.get('competencia')
        context['relatorio'] = ConferenciaLancamento.gerar_relatorio_conferencia(
            empresa,
            competencia
        )
        
        context['empresa'] = empresa
        context['competencia_filtro'] = competencia
        context['status_filtro'] = self.request.GET.get('status', 'PENDENTE')
        context['page_title'] = 'Conferência de Lançamentos'
        
        return context


class ConferenciaDetailView(LoginRequiredMixin, EmpresaScopeMixin, DetailView):
    """Exibe detalhes de uma conferência específica"""
    model = ConferenciaLancamento
    template_name = 'lancamentos/conferencia_detail.html'
    context_object_name = 'conferencia'
    pk_url_kwarg = 'conferencia_id'
    
    def get_object(self):
        obj = super().get_object()
        
        # Verifica permissão
        if not is_empresa_allowed(self.request.user, obj.lancamento.empresa.codigo):
            raise HttpResponseForbidden('Sem permissão para acessar esta conferência')
        
        return obj
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Conferência #{self.object.id}'
        context['lancamento'] = self.object.lancamento
        
        # Executar validações
        context['problemas'] = self.object._validar()
        
        return context


class ConferenciaConferirView(LoginRequiredMixin, EmpresaScopeMixin, FormView):
    """Formulário para conferir um lançamento"""
    template_name = 'lancamentos/conferencia_conferir.html'
    form_class = ConferenciaLancamentoForm
    
    def dispatch(self, request, *args, **kwargs):
        self.conferencia = get_object_or_404(
            ConferenciaLancamento,
            pk=kwargs.get('conferencia_id')
        )
        
        # Verifica permissão
        if not is_empresa_allowed(request.user, self.conferencia.lancamento.empresa.codigo):
            return HttpResponseForbidden('Sem permissão para conferir este lançamento')
        
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['conferencia'] = self.conferencia
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['conferencia'] = self.conferencia
        context['lancamento'] = self.conferencia.lancamento
        context['page_title'] = 'Conferir Lançamento'
        
        # Executar validações
        context['problemas'] = self.conferencia._validar()
        
        return context
    
    def form_valid(self, form):
        valor_conferido = form.cleaned_data.get('valor_conferido')
        observacoes = form.cleaned_data.get('observacoes', '')
        
        # Conferir lançamento
        valido = self.conferencia.conferir(
            self.request.user,
            valor_conferido,
            observacoes
        )
        
        if valido:
            messages.success(
                self.request,
                f'✅ Lançamento conferido com sucesso! Status: CONFERIDO'
            )
        else:
            messages.warning(
                self.request,
                f'⚠️ Lançamento conferido COM PROBLEMAS. Revise as validações automáticas.'
            )
        
        return redirect('conferencia-list', empresa_id=self.conferencia.lancamento.empresa_id)
    
    def get_success_url(self):
        return reverse_lazy(
            'conferencia-list',
            kwargs={'empresa_id': self.conferencia.lancamento.empresa_id}
        )


class ConferenciaRejeitarView(LoginRequiredMixin, EmpresaScopeMixin, FormView):
    """Formulário para rejeitar um lançamento"""
    template_name = 'lancamentos/conferencia_rejeitar.html'
    form_class = RejeicaoLancamentoForm
    
    def dispatch(self, request, *args, **kwargs):
        self.conferencia = get_object_or_404(
            ConferenciaLancamento,
            pk=kwargs.get('conferencia_id')
        )
        
        # Verifica permissão
        if not is_empresa_allowed(request.user, self.conferencia.lancamento.empresa.codigo):
            return HttpResponseForbidden('Sem permissão para rejeitar este lançamento')
        
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['conferencia'] = self.conferencia
        context['lancamento'] = self.conferencia.lancamento
        context['page_title'] = 'Rejeitar Lançamento'
        return context
    
    def form_valid(self, form):
        motivo_padrao = form.cleaned_data.get('motivo_padrao')
        motivo_detalhado = form.cleaned_data.get('motivo_detalhado')
        
        # Monta motivo completo
        motivos_dict = dict(RejeicaoLancamentoForm.MOTIVOS_REJEICAO)
        motivo_completo = f"{motivos_dict[motivo_padrao]}: {motivo_detalhado}"
        
        # Rejeita lançamento
        self.conferencia.rejeitar(self.request.user, motivo_completo)
        
        messages.error(
            self.request,
            f'❌ Lançamento rejeitado. Motivo: {motivos_dict[motivo_padrao]}'
        )
        
        return redirect('conferencia-list', empresa_id=self.conferencia.lancamento.empresa_id)
    
    def get_success_url(self):
        return reverse_lazy(
            'conferencia-list',
            kwargs={'empresa_id': self.conferencia.lancamento.empresa_id}
        )


class ConferenciaRelatorioView(LoginRequiredMixin, EmpresaScopeMixin, View):
    """Exibe relatório consolidado de conferências"""
    template_name = 'lancamentos/conferencia_relatorio.html'
    
    def get(self, request, empresa_id):
        empresa = get_object_or_404(Empresa, pk=empresa_id)
        
        # Verifica permissão
        if not is_empresa_allowed(request.user, empresa.codigo):
            return HttpResponseForbidden('Sem permissão para acessar este relatório')
        
        competencia = request.GET.get('competencia')
        
        # Relatório geral
        relatorio = ConferenciaLancamento.gerar_relatorio_conferencia(empresa, competencia)
        
        # Detalhes por status
        conferencias_por_status = {
            'PENDENTE': ConferenciaLancamento.objects.filter(
                lancamento__empresa=empresa,
                status='PENDENTE'
            ).select_related('lancamento__funcionario'),
            'CONFERIDO': ConferenciaLancamento.objects.filter(
                lancamento__empresa=empresa,
                status='CONFERIDO'
            ).select_related('lancamento__funcionario', 'conferido_por'),
            'PROBLEMA': ConferenciaLancamento.objects.filter(
                lancamento__empresa=empresa,
                status='PROBLEMA'
            ).select_related('lancamento__funcionario', 'conferido_por'),
            'REJEITADO': ConferenciaLancamento.objects.filter(
                lancamento__empresa=empresa,
                status='REJEITADO'
            ).select_related('lancamento__funcionario', 'conferido_por'),
        }
        
        if competencia:
            for status, qs in conferencias_por_status.items():
                conferencias_por_status[status] = qs.filter(
                    lancamento__competencia=competencia
                )
        
        # Verificar se pode consolidar
        pode_consolidar, msg_consolidacao = False, ''
        if competencia:
            pode_consolidar, msg_consolidacao = ConferenciaLancamento.pode_consolidar_competencia(
                empresa,
                competencia
            )
        
        context = {
            'empresa': empresa,
            'competencia': competencia,
            'relatorio': relatorio,
            'conferencias_por_status': conferencias_por_status,
            'pode_consolidar': pode_consolidar,
            'msg_consolidacao': msg_consolidacao,
            'page_title': 'Relatório de Conferências',
        }

        return render(request, self.template_name, context)


# ===== IMPORTAÇÃO SEFIP.RE =====

class SefipImportView(LoginRequiredMixin, FormView):
    """
    Recebe o arquivo SEFIP.RE, executa o parser e cria Lancamentos.
    Accessível via GET (formulário) e POST (processamento).
    """
    form_class = SefipImportForm
    template_name = 'lancamentos/sefip_import.html'
    success_url = reverse_lazy('sefip-import')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        from .services.sefip_importer import SefipImporter, SefipImportError

        empresa = form.cleaned_data['empresa']
        arquivo = form.cleaned_data['arquivo_re']

        if not is_empresa_allowed(self.request.user, empresa.codigo):
            messages.error(self.request, 'Você não tem permissão para importar dados nesta empresa.')
            return self.form_invalid(form)

        try:
            arquivo_bytes = arquivo.read()
        except Exception as exc:
            messages.error(self.request, f'Erro ao ler o arquivo: {exc}')
            return self.form_invalid(form)

        try:
            importer = SefipImporter()
            resultado = importer.importar(arquivo_bytes, empresa)
        except SefipImportError as exc:
            messages.error(self.request, f'Erro no arquivo: {exc}')
            return self.form_invalid(form)
        except Exception as exc:
            import logging
            logging.getLogger(__name__).error('Erro na importação SEFIP.RE', exc_info=True)
            messages.error(self.request, f'Erro inesperado ao processar o arquivo: {exc}')
            return self.form_invalid(form)

        self.request.session['sefip_import_result'] = {
            'empresa_nome': empresa.nome,
            'criados': resultado['criados'],
            'ignorados': resultado['ignorados'],
            'competencias': resultado['competencias'],
            'erros': resultado['erros'][:30],
            'avisos': resultado['avisos'][:30],
            'total_erros': len(resultado['erros']),
            'total_avisos': len(resultado['avisos']),
        }

        competencias_str = ', '.join(resultado['competencias']) or '—'
        msg = (
            f"{resultado['criados']} lançamento(s) importado(s) "
            f"(competência: {competencias_str})"
        )
        if resultado['ignorados']:
            msg += f', {resultado["ignorados"]} ignorado(s)'
        if resultado['erros']:
            msg += f', {len(resultado["erros"])} erro(s)'

        if not resultado['erros']:
            messages.success(self.request, f'Importação concluída. {msg}.')
        else:
            messages.warning(self.request, f'Importação concluída com avisos. {msg}.')

        return redirect('sefip-import-result')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['ultimo_resultado'] = self.request.session.get('sefip_import_result')
        return context


class SefipImportResultView(LoginRequiredMixin, View):
    """Exibe o relatório detalhado da última importação SEFIP.RE."""
    template_name = 'lancamentos/sefip_import.html'

    def get(self, request, *args, **kwargs):
        resultado = request.session.get('sefip_import_result')
        if not resultado:
            messages.warning(request, 'Nenhuma importação recente encontrada.')
            return redirect('sefip-import')

        form = SefipImportForm(user=request.user)
        context = {
            'form': form,
            'resultado': resultado,
            'ultimo_resultado': resultado,
        }
        return render(request, self.template_name, context)


# ---------------------------------------------------------------------------
# Helpers para o relatório de recolhimento por funcionário
# ---------------------------------------------------------------------------

def _gerar_competencias_no_periodo(inicio_str, fim_str):
    """Gera lista de strings 'MM/YYYY' do período [inicio, fim] inclusive."""
    from dateutil.relativedelta import relativedelta as rd
    fmt = '%m/%Y'
    dt = datetime.strptime(inicio_str, fmt).replace(day=1)
    dt_fim = datetime.strptime(fim_str, fmt).replace(day=1)
    competencias = []
    while dt <= dt_fim:
        competencias.append(dt.strftime('%m/%Y'))
        dt += rd(months=1)
    return competencias


def _valor_atualizado_lancamento(lanc, data_ref, indice_cache):
    """
    Valor atualizado de um único lançamento na data de referência (data_ref):
    - Pago: mantém o valor efetivamente pago (ou valor_fgts se valor_pago nulo).
    - Em aberto: valor de depósito corrigido (base_fgts × índice efetivo), sem somar JAM.
    """
    from empresas.models_grupo import get_aliquota_fgts
    if lanc.pago:
        return lanc.valor_pago if lanc.valor_pago is not None else (lanc.valor_fgts or Decimal('0'))
    if not data_ref or not lanc.competencia:
        return lanc.valor_fgts or Decimal('0')
    try:
        comp_str = lanc.competencia.replace('13/', '12/')
        mes_str, ano_str = comp_str.split('/')
        competencia_date = date(int(ano_str), int(mes_str), 1)
    except Exception:
        return lanc.valor_fgts or Decimal('0')

    if competencia_date not in indice_cache:
        indice_cache[competencia_date] = IndiceFGTSService.buscar_indice(
            competencia=competencia_date, data_pagamento=data_ref,
        )
    indice = indice_cache[competencia_date]
    if indice is None:
        return lanc.valor_fgts or Decimal('0')

    try:
        resultado = calcular_fgts_atualizado(
            valor_fgts=lanc.valor_fgts,
            competencia=competencia_date,
            pagamento=data_ref,
            indice=indice,
            jam_coef=Decimal('0'),
            valor_fgts_base=lanc.base_fgts,
            aliquota=get_aliquota_fgts(getattr(lanc, 'vinculo', None)),
            aplicar_plano_economico=True,
        )
        return resultado['valor_deposito_fgts']
    except Exception:
        return lanc.valor_fgts or Decimal('0')


def _filtrar_qs_por_funcionario(qs, funcionario_filtro):
    """
    Aplica o filtro de funcionário vindo de RelatorioRecolhimentoFuncionarioForm.clean_funcionario:
    - modo 'cpf': busca por CPF, cruzando todas as empresas do queryset (mesma pessoa pode ter
      registros de Funcionario distintos por empresa).
    - modo 'id': busca por Funcionario.pk exato.
    """
    if not funcionario_filtro:
        return qs
    if funcionario_filtro['modo'] == 'cpf':
        return qs.filter(funcionario__cpf=funcionario_filtro['cpf'])
    return qs.filter(funcionario_id=funcionario_filtro['funcionario_id'])


def _calcular_dados_recolhimento(lancamentos_qs, data_ref=None, indice_cache=None):
    """
    Dado um queryset de Lancamento, retorna dict com:
    - valor_a_recolher: soma de valor_fgts dos não pagos
    - total_recolhido:  soma de valor_pago dos pagos (ou valor_fgts se valor_pago nulo)
    - valor_atualizado: soma do valor atualizado (na data_ref) de cada lançamento
    """
    valor_a_recolher = Decimal('0')
    total_recolhido = Decimal('0')
    valor_atualizado = Decimal('0')
    if indice_cache is None:
        indice_cache = {}
    for lanc in lancamentos_qs:
        if lanc.pago:
            total_recolhido += (lanc.valor_pago or lanc.valor_fgts or Decimal('0'))
        else:
            valor_a_recolher += (lanc.valor_fgts or Decimal('0'))
            valor_atualizado += _valor_atualizado_lancamento(lanc, data_ref, indice_cache)
    return {
        'valor_a_recolher': valor_a_recolher,
        'total_recolhido': total_recolhido,
        'total_geral': valor_a_recolher + total_recolhido,
        'valor_atualizado': valor_atualizado,
    }


class RelatorioRecolhimentoFuncionarioView(LoginRequiredMixin, FormView):
    """Relatório: Listagem do Recolhimento por Funcionário"""
    template_name = 'lancamentos/relatorio_recolhimento_funcionario.html'
    form_class = RelatorioRecolhimentoFuncionarioForm

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        empresa = form.cleaned_data.get('empresa')
        funcionario_filtro = form.cleaned_data.get('funcionario')
        comp_inicio = form.cleaned_data['competencia_inicio']
        comp_fim = form.cleaned_data['competencia_fim']

        competencias_periodo = _gerar_competencias_no_periodo(comp_inicio, comp_fim)

        # Monta queryset base
        qs = Lancamento.objects.filter(competencia__in=competencias_periodo).select_related(
            'empresa', 'funcionario', 'vinculo', 'vinculo__tipo_vinculo'
        )
        allowed_ids = get_allowed_empresa_ids(self.request.user)
        if allowed_ids is not None:
            qs = qs.filter(empresa__codigo__in=allowed_ids)
        if empresa:
            qs = qs.filter(empresa=empresa)
        qs = _filtrar_qs_por_funcionario(qs, funcionario_filtro)

        # Agrupa por empresa → funcionário
        from collections import defaultdict
        empresas_dict = defaultdict(lambda: {'empresa': None, 'funcionarios': defaultdict(list)})

        for lanc in qs:
            emp_id = lanc.empresa_id
            empresas_dict[emp_id]['empresa'] = lanc.empresa
            empresas_dict[emp_id]['funcionarios'][lanc.funcionario_id].append(lanc)

        # Sempre exibe "Valor Atualizado" (FGTS em aberto corrigido pelo índice) como
        # terceira métrica, independente de empresa única ou grupo.
        modo_valor_atualizado = True
        data_ref = IndiceFGTSService.obter_ultima_data_base()
        indice_cache = {}

        # Monta estrutura final
        empresas_resultado = []
        total_geral_recolher = Decimal('0')
        total_geral_recolhido = Decimal('0')

        for emp_id, emp_data in sorted(empresas_dict.items()):
            emp_obj = emp_data['empresa']
            funcionarios_lista = []
            emp_recolher = Decimal('0')
            emp_recolhido = Decimal('0')
            emp_atualizado = Decimal('0')

            for func_id, lancamentos in sorted(
                emp_data['funcionarios'].items(),
                key=lambda x: x[1][0].funcionario.nome if x[1] else ''
            ):
                func_obj = lancamentos[0].funcionario
                dados = _calcular_dados_recolhimento(lancamentos, data_ref, indice_cache)

                # Dados de admissão/demissão via vínculo
                vinculo = getattr(lancamentos[0], 'vinculo', None) or func_obj.vinculo_atual()
                data_admissao = vinculo.data_admissao if vinculo else getattr(func_obj, 'data_admissao', None)
                data_demissao = vinculo.data_demissao if vinculo else getattr(func_obj, 'data_demissao', None)

                total_exibido = dados['valor_atualizado'] if modo_valor_atualizado else dados['total_geral']
                funcionarios_lista.append({
                    'funcionario': func_obj,
                    'matricula': vinculo.matricula if vinculo and vinculo.matricula else '',
                    'data_admissao': data_admissao,
                    'data_demissao': data_demissao,
                    'pis': func_obj.pis,
                    'valor_a_recolher': dados['valor_a_recolher'],
                    'total_recolhido': dados['total_recolhido'],
                    'total_geral': total_exibido,
                })
                emp_recolher += dados['valor_a_recolher']
                emp_recolhido += dados['total_recolhido']
                emp_atualizado += dados['valor_atualizado']

            emp_total_exibido = emp_atualizado if modo_valor_atualizado else (emp_recolher + emp_recolhido)
            empresas_resultado.append({
                'empresa': emp_obj,
                'funcionarios': funcionarios_lista,
                'total_a_recolher': emp_recolher,
                'total_recolhido': emp_recolhido,
                'total_geral': emp_total_exibido,
            })
            total_geral_recolher += emp_recolher
            total_geral_recolhido += emp_recolhido

        total_geral_exibido = (
            sum((e['total_geral'] for e in empresas_resultado), Decimal('0'))
            if modo_valor_atualizado else total_geral_recolher + total_geral_recolhido
        )

        return render(self.request, self.template_name, {
            'form': form,
            'empresas_resultado': empresas_resultado,
            'comp_inicio': comp_inicio,
            'comp_fim': comp_fim,
            'total_geral_recolher': total_geral_recolher,
            'total_geral_recolhido': total_geral_recolhido,
            'total_geral': total_geral_exibido,
            'modo_valor_atualizado': modo_valor_atualizado,
            'gerou_relatorio': True,
        })

    def form_invalid(self, form):
        return render(self.request, self.template_name, {'form': form})

    def get(self, request, *args, **kwargs):
        form = self.get_form()
        return render(request, self.template_name, {'form': form})


def export_recolhimento_funcionario_pdf(request):
    """Exporta o relatório de Recolhimento por Funcionário em PDF usando ReportLab."""
    if not request.user.is_authenticated:
        return redirect('login')

    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    form = RelatorioRecolhimentoFuncionarioForm(data=request.GET, user=request.user)
    if not form.is_valid():
        messages.error(request, 'Parâmetros inválidos para exportação.')
        return redirect('relatorio-recolhimento-funcionario')

    empresa_filtro = form.cleaned_data.get('empresa')
    funcionario_filtro = form.cleaned_data.get('funcionario')
    comp_inicio = form.cleaned_data['competencia_inicio']
    comp_fim = form.cleaned_data['competencia_fim']

    competencias_periodo = _gerar_competencias_no_periodo(comp_inicio, comp_fim)

    qs = Lancamento.objects.filter(competencia__in=competencias_periodo).select_related(
        'empresa', 'funcionario', 'vinculo', 'vinculo__tipo_vinculo'
    )
    allowed_ids = get_allowed_empresa_ids(request.user)
    if allowed_ids is not None:
        qs = qs.filter(empresa__codigo__in=allowed_ids)
    if empresa_filtro:
        qs = qs.filter(empresa=empresa_filtro)
    qs = _filtrar_qs_por_funcionario(qs, funcionario_filtro)

    from collections import defaultdict
    empresas_dict = defaultdict(lambda: {'empresa': None, 'funcionarios': defaultdict(list)})
    for lanc in qs:
        empresas_dict[lanc.empresa_id]['empresa'] = lanc.empresa
        empresas_dict[lanc.empresa_id]['funcionarios'][lanc.funcionario_id].append(lanc)

    modo_valor_atualizado = True
    data_ref = IndiceFGTSService.obter_ultima_data_base()
    indice_cache = {}
    label_total = 'Valor Atualizado'

    # Estilos
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle('titulo', parent=styles['Normal'],
        fontSize=10, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=2,
        textColor=colors.white)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'],
        fontSize=7.5, fontName='Helvetica', alignment=TA_CENTER, spaceAfter=1,
        textColor=colors.white)
    empresa_style = ParagraphStyle('empresa', parent=styles['Normal'],
        fontSize=8.5, fontName='Helvetica-Bold', spaceAfter=2)
    normal_style = ParagraphStyle('normal_s', parent=styles['Normal'],
        fontSize=7.5, fontName='Helvetica')
    normal_white_style = ParagraphStyle('normal_white', parent=styles['Normal'],
        fontSize=7.5, fontName='Helvetica-Bold', textColor=colors.white)
    header_style = ParagraphStyle('header_s', parent=styles['Normal'],
        fontSize=7.5, fontName='Helvetica-Bold', textColor=colors.white)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=12*mm, rightMargin=12*mm,
        topMargin=12*mm, bottomMargin=12*mm
    )

    story = []
    agora = datetime.now().strftime('%d/%m/%Y - %H:%M')
    cor_header = colors.HexColor('#003A78')
    cor_linha_par = colors.HexColor('#EBF3FB')
    col_widths = [25*mm, 75*mm, 25*mm, 25*mm, 35*mm, 35*mm, 35*mm, 35*mm]

    total_geral_recolher = Decimal('0')
    total_geral_recolhido = Decimal('0')
    total_geral_atualizado = Decimal('0')
    primeira_empresa = True

    for emp_id, emp_data in sorted(empresas_dict.items()):
        emp_obj = emp_data['empresa']
        if not primeira_empresa:
            story.append(PageBreak())
        primeira_empresa = False

        # Cabeçalho do relatório
        cab_data = [
            [
                Paragraph('LISTAGEM DO RECOLHIMENTO POR FUNCIONÁRIO', titulo_style),
                Paragraph(f'Período: {comp_inicio} Até {comp_fim}', sub_style),
                Paragraph(f'Data/Hora: {agora}', sub_style),
            ]
        ]
        cab_table = Table(cab_data, colWidths=[110*mm, 80*mm, 67*mm])
        cab_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), cor_header),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
            ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(cab_table)
        story.append(Spacer(1, 3*mm))

        # Identificação da empresa
        cnpj = emp_obj.cnpj or ''
        story.append(Paragraph(f'{emp_obj.nome}  —  CNPJ: {cnpj}', empresa_style))
        story.append(Spacer(1, 2*mm))

        # Cabeçalho da tabela de dados
        header_row = [
            Paragraph('Cod.', header_style),
            Paragraph('Nome do Funcionário', header_style),
            Paragraph('Admissão', header_style),
            Paragraph('Demissão', header_style),
            Paragraph('P I S', header_style),
            Paragraph('Valor a Recolher', header_style),
            Paragraph('Total Recolhido', header_style),
            Paragraph(label_total, header_style),
        ]

        table_data = [header_row]
        emp_recolher = Decimal('0')
        emp_recolhido = Decimal('0')
        emp_atualizado = Decimal('0')
        row_idx = 0

        for func_id, lancamentos in sorted(
            emp_data['funcionarios'].items(),
            key=lambda x: x[1][0].funcionario.nome if x[1] else ''
        ):
            func_obj = lancamentos[0].funcionario
            dados = _calcular_dados_recolhimento(lancamentos, data_ref, indice_cache)
            vinculo = getattr(lancamentos[0], 'vinculo', None) or func_obj.vinculo_atual()
            data_adm = vinculo.data_admissao if vinculo else getattr(func_obj, 'data_admissao', None)
            data_dem = vinculo.data_demissao if vinculo else getattr(func_obj, 'data_demissao', None)

            emp_recolher += dados['valor_a_recolher']
            emp_recolhido += dados['total_recolhido']
            emp_atualizado += dados['valor_atualizado']
            total_exibido_linha = dados['valor_atualizado'] if modo_valor_atualizado else dados['total_geral']

            def fmt_date(d):
                return d.strftime('%d/%m/%Y') if d else '—'

            def fmt_val(v):
                return f'{v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

            table_data.append([
                Paragraph(vinculo.matricula if vinculo and vinculo.matricula else '', normal_style),
                Paragraph(func_obj.nome, normal_style),
                Paragraph(fmt_date(data_adm), normal_style),
                Paragraph(fmt_date(data_dem), normal_style),
                Paragraph(func_obj.pis or '—', normal_style),
                Paragraph(fmt_val(dados['valor_a_recolher']), normal_style),
                Paragraph(fmt_val(dados['total_recolhido']), normal_style),
                Paragraph(fmt_val(total_exibido_linha), normal_style),
            ])
            row_idx += 1

        # Linha de total da empresa
        def fmt_val(v):
            return f'{v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')

        total_emp = emp_atualizado if modo_valor_atualizado else (emp_recolher + emp_recolhido)
        table_data.append([
            Paragraph('', normal_style),
            Paragraph('<b>EMPRESA: Totais:</b>', normal_style),
            Paragraph('', normal_style),
            Paragraph('', normal_style),
            Paragraph('', normal_style),
            Paragraph(f'<b>{fmt_val(emp_recolher)}</b>', normal_style),
            Paragraph(f'<b>{fmt_val(emp_recolhido)}</b>', normal_style),
            Paragraph(f'<b>{fmt_val(total_emp)}</b>', normal_style),
        ])
        total_geral_recolher += emp_recolher
        total_geral_recolhido += emp_recolhido
        total_geral_atualizado += emp_atualizado

        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        ts = [
            ('BACKGROUND', (0, 0), (-1, 0), cor_header),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('ALIGN', (5, 0), (7, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (4, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, cor_linha_par]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D0DFF0')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#B0C4D8')),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]
        tbl.setStyle(TableStyle(ts))
        story.append(tbl)

    # Total geral (todas as empresas)
    if len(empresas_dict) > 1:
        story.append(Spacer(1, 5*mm))
        total_rel = total_geral_atualizado if modo_valor_atualizado else (total_geral_recolher + total_geral_recolhido)
        def fmt_val(v):
            return f'{v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        total_data = [[
            Paragraph('TOTAL GERAL DO RELATÓRIO:', normal_white_style),
            Paragraph(fmt_val(total_geral_recolher), normal_white_style),
            Paragraph(fmt_val(total_geral_recolhido), normal_white_style),
            Paragraph(fmt_val(total_rel), normal_white_style),
        ]]
        total_tbl = Table(total_data, colWidths=[165*mm, 35*mm, 35*mm, 35*mm])
        total_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#003A78')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(total_tbl)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    resp = HttpResponse(content_type='application/pdf')
    resp['Content-Disposition'] = (
        f'attachment; filename="recolhimento_funcionario_{comp_inicio.replace("/","_")}_a_{comp_fim.replace("/","_")}.pdf"'
    )
    resp.write(pdf)
    return resp


def export_recolhimento_funcionario_xlsx(request):
    """Exporta o relatório de Recolhimento por Funcionário em XLSX."""
    if not request.user.is_authenticated:
        return redirect('login')

    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    form = RelatorioRecolhimentoFuncionarioForm(data=request.GET, user=request.user)
    if not form.is_valid():
        messages.error(request, 'Parâmetros inválidos para exportação.')
        return redirect('recolhimento-funcionario')

    empresa_filtro = form.cleaned_data.get('empresa')
    funcionario_filtro = form.cleaned_data.get('funcionario')
    comp_inicio = form.cleaned_data['competencia_inicio']
    comp_fim = form.cleaned_data['competencia_fim']

    competencias_periodo = _gerar_competencias_no_periodo(comp_inicio, comp_fim)

    qs = Lancamento.objects.filter(competencia__in=competencias_periodo).select_related(
        'empresa', 'funcionario', 'vinculo', 'vinculo__tipo_vinculo'
    )
    allowed_ids = get_allowed_empresa_ids(request.user)
    if allowed_ids is not None:
        qs = qs.filter(empresa__codigo__in=allowed_ids)
    if empresa_filtro:
        qs = qs.filter(empresa=empresa_filtro)
    qs = _filtrar_qs_por_funcionario(qs, funcionario_filtro)

    from collections import defaultdict
    empresas_dict = defaultdict(lambda: {'empresa': None, 'funcionarios': defaultdict(list)})
    for lanc in qs:
        empresas_dict[lanc.empresa_id]['empresa'] = lanc.empresa
        empresas_dict[lanc.empresa_id]['funcionarios'][lanc.funcionario_id].append(lanc)

    modo_valor_atualizado = True
    data_ref = IndiceFGTSService.obter_ultima_data_base()
    indice_cache = {}
    label_total = 'Valor Atualizado'

    def fmt_date(d):
        return d.strftime('%d/%m/%Y') if d else ''

    cor_header = '003A78'
    cor_total_emp = 'D0DFF0'
    cor_total_geral = '003A78'
    cor_linha_par = 'EBF3FB'

    thin_border = Border(
        left=Side(style='thin', color='B0C4D8'),
        right=Side(style='thin', color='B0C4D8'),
        top=Side(style='thin', color='B0C4D8'),
        bottom=Side(style='thin', color='B0C4D8'),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Recolhimento por Funcionário'

    agora = datetime.now().strftime('%d/%m/%Y %H:%M')
    ws.append([
        'LISTAGEM DO RECOLHIMENTO POR FUNCIONÁRIO',
        f'Período: {comp_inicio} até {comp_fim}',
        '', '', '', '', '',
        f'Gerado em: {agora}',
    ])
    titulo_row = ws.max_row
    for col in range(1, 9):
        cell = ws.cell(row=titulo_row, column=col)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(fill_type='solid', fgColor=cor_header)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[titulo_row].height = 22

    ws.append([])  # linha em branco

    colunas = ['Cod.', 'Nome do Funcionário', 'Admissão', 'Demissão', 'PIS',
               'Valor a Recolher', 'Total Recolhido', label_total]

    total_geral_recolher = Decimal('0')
    total_geral_recolhido = Decimal('0')
    total_geral_atualizado = Decimal('0')

    for emp_id, emp_data in sorted(empresas_dict.items()):
        emp_obj = emp_data['empresa']

        # Linha de identificação da empresa
        ws.append([f'{emp_obj.nome}  —  CNPJ: {emp_obj.cnpj or ""}'])
        emp_nome_row = ws.max_row
        cell = ws.cell(row=emp_nome_row, column=1)
        cell.font = Font(bold=True, size=10)
        ws.merge_cells(start_row=emp_nome_row, start_column=1,
                       end_row=emp_nome_row, end_column=8)

        # Cabeçalho das colunas
        ws.append(colunas)
        header_row = ws.max_row
        for col, _ in enumerate(colunas, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color='FFFFFF', size=9)
            cell.fill = PatternFill(fill_type='solid', fgColor=cor_header)
            cell.alignment = Alignment(horizontal='center' if col > 5 else 'left',
                                       vertical='center')
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 16

        emp_recolher = Decimal('0')
        emp_recolhido = Decimal('0')
        emp_atualizado = Decimal('0')
        row_idx = 0

        for func_id, lancamentos in sorted(
            emp_data['funcionarios'].items(),
            key=lambda x: x[1][0].funcionario.nome if x[1] else ''
        ):
            func_obj = lancamentos[0].funcionario
            dados = _calcular_dados_recolhimento(lancamentos, data_ref, indice_cache)
            vinculo = getattr(lancamentos[0], 'vinculo', None) or func_obj.vinculo_atual()
            data_adm = vinculo.data_admissao if vinculo else getattr(func_obj, 'data_admissao', None)
            data_dem = vinculo.data_demissao if vinculo else getattr(func_obj, 'data_demissao', None)

            emp_recolher += dados['valor_a_recolher']
            emp_recolhido += dados['total_recolhido']
            emp_atualizado += dados['valor_atualizado']

            total_exibido_linha = dados['valor_atualizado'] if modo_valor_atualizado else dados['total_geral']

            fill = PatternFill(fill_type='solid', fgColor=cor_linha_par) if row_idx % 2 == 1 else None

            ws.append([
                vinculo.matricula if vinculo and vinculo.matricula else '',
                func_obj.nome,
                fmt_date(data_adm),
                fmt_date(data_dem),
                func_obj.pis or '',
                float(dados['valor_a_recolher']),
                float(dados['total_recolhido']),
                float(total_exibido_linha),
            ])
            data_row = ws.max_row
            for col in range(1, 9):
                cell = ws.cell(row=data_row, column=col)
                cell.border = thin_border
                cell.font = Font(size=9)
                if fill:
                    cell.fill = fill
                if col >= 6:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
            row_idx += 1

        # Total da empresa
        total_emp = emp_atualizado if modo_valor_atualizado else (emp_recolher + emp_recolhido)
        ws.append([
            '', 'EMPRESA: Totais:', '', '', '',
            float(emp_recolher), float(emp_recolhido), float(total_emp)
        ])
        total_emp_row = ws.max_row
        for col in range(1, 9):
            cell = ws.cell(row=total_emp_row, column=col)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill(fill_type='solid', fgColor=cor_total_emp)
            cell.border = thin_border
            if col >= 6:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

        total_geral_recolher += emp_recolher
        total_geral_recolhido += emp_recolhido
        total_geral_atualizado += emp_atualizado

        ws.append([])  # separador entre empresas

    # Total geral (quando há > 1 empresa)
    if len(empresas_dict) > 1:
        total_rel = total_geral_atualizado if modo_valor_atualizado else (total_geral_recolher + total_geral_recolhido)
        ws.append([
            'TOTAL GERAL DO RELATÓRIO:', '', '', '', '',
            float(total_geral_recolher), float(total_geral_recolhido), float(total_rel)
        ])
        total_geral_row = ws.max_row
        ws.merge_cells(start_row=total_geral_row, start_column=1,
                       end_row=total_geral_row, end_column=5)
        for col in range(1, 9):
            cell = ws.cell(row=total_geral_row, column=col)
            cell.font = Font(bold=True, color='FFFFFF', size=9)
            cell.fill = PatternFill(fill_type='solid', fgColor=cor_total_geral)
            cell.border = thin_border
            if col >= 6:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')

    col_widths_xlsx = [10, 40, 13, 13, 18, 18, 18, 18]
    for i, width in enumerate(col_widths_xlsx, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    xlsx_bytes = buffer.getvalue()
    buffer.close()

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = (
        f'attachment; filename="recolhimento_funcionario_{comp_inicio.replace("/","_")}'
        f'_a_{comp_fim.replace("/","_")}.xlsx"'
    )
    resp.write(xlsx_bytes)
    return resp


# ---------------------------------------------------------------------------
# Relatório de Posição em Aberto (com Valor Atualizado)
# ---------------------------------------------------------------------------

class RelatorioStatusPosicaoView(LoginRequiredMixin, View):
    """Formulário de filtros para o relatório de posição em aberto."""
    template_name = 'lancamentos/relatorio_posicao.html'

    def _form(self, data=None):
        from .forms import RelatorioStatusPosicaoForm
        empresa_ids = get_allowed_empresa_ids(self.request.user)
        return RelatorioStatusPosicaoForm(data, empresa_ids=empresa_ids)

    def _check_feature(self, empresa):
        from empresas.models_feature import empresa_tem_recurso
        user = self.request.user
        if user.is_staff or user.is_superuser:
            return True
        return empresa_tem_recurso(empresa, 'relatorio_posicao_fgts')

    def get(self, request):
        return render(request, self.template_name, {'form': self._form()})

    def post(self, request):
        form = self._form(request.POST)
        if not form.is_valid():
            return render(request, self.template_name, {'form': form})

        empresa = form.cleaned_data['empresa']
        if not is_empresa_allowed(request.user, empresa.pk):
            return HttpResponseForbidden()
        if not self._check_feature(empresa):
            return HttpResponseForbidden()

        from .models_relatorio import RelatorioTask
        from .services.relatorio_service import processar_relatorio_posicao

        task = RelatorioTask.objects.create(
            empresa=empresa,
            usuario=request.user,
            status='pending',
            parametros_json={
                'tipo': 'posicao',
                'empresa_id': empresa.pk,
                'competencia_inicio': form.cleaned_data['competencia_inicio'],
                'competencia_fim': form.cleaned_data['competencia_fim'],
            },
        )
        threading.Thread(target=processar_relatorio_posicao, args=(task.id,), daemon=True).start()
        return redirect('relatorio-task-status', pk=task.pk)


class RelatorioStatusPosicaoResultadoView(LoginRequiredMixin, View):
    """Exibe resultado do relatório de posição após processamento assíncrono."""
    template_name = 'lancamentos/relatorio_posicao_resultado.html'

    def get(self, request, pk):
        from .models_relatorio import RelatorioTask
        from empresas.models_feature import empresa_tem_recurso
        task = get_object_or_404(RelatorioTask, pk=pk, usuario=request.user)
        if not request.user.is_staff and not request.user.is_superuser:
            if task.empresa and not empresa_tem_recurso(task.empresa, 'relatorio_posicao_fgts'):
                return HttpResponseForbidden()
        if task.status != 'done':
            return redirect('relatorio-task-status', pk=pk)
        resultado = task.resultado_json or {}
        linhas = resultado.get('linhas', [])
        return render(request, self.template_name, {
            'task': task,
            'resultado': resultado,
            'linhas_preview': linhas[:50],
            'total_linhas': len(linhas),
            'data_ref': resultado.get('data_ref'),
        })


@login_required
def export_relatorio_posicao_xlsx(request, pk):
    """Gera download XLSX do relatório de posição em aberto."""
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font
    from .models_relatorio import RelatorioTask
    from empresas.models_feature import empresa_tem_recurso

    task = get_object_or_404(RelatorioTask, pk=pk, usuario=request.user)
    if not request.user.is_staff and not request.user.is_superuser:
        if task.empresa and not empresa_tem_recurso(task.empresa, 'relatorio_posicao_fgts'):
            return HttpResponseForbidden()
    if task.status != 'done':
        return redirect('relatorio-task-status', pk=pk)

    resultado = task.resultado_json or {}
    linhas = resultado.get('linhas', [])
    data_ref = resultado.get('data_ref', '')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Posição FGTS'

    headers = [
        'Cod_Empresa', 'Empresa', 'CNPJ',
        'Funcionário', 'Tipo Vínculo', 'PIS', 'Matrícula', 'Cargo', 'CBO',
        'Admissão', 'Demissão', 'Ano', 'Competência',
        'Base FGTS', 'Valor FGTS', 'parcela_13', 'Status Pagamento', 'Data Pagamento',
        'Valor Pago', 'Fonte Confirmação', f'Valor Atualizado (ref. {data_ref})',
        'Status Vínculo', 'Motivo Desligamento',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    def _fmt_date(iso_str):
        if not iso_str:
            return ''
        try:
            from datetime import date
            d = date.fromisoformat(iso_str)
            return d.strftime('%d/%m/%Y')
        except Exception:
            return iso_str

    def _decimal(s):
        if s is None or s == '':
            return ''
        try:
            return float(s)
        except Exception:
            return s

    def _ano(l):
        ano = l.get('ano')
        if ano:
            return ano
        comp = l.get('competencia', '')
        return comp.split('/')[-1] if '/' in comp else ''

    for l in linhas:
        ws.append([
            l.get('cod_empresa', ''),
            l.get('empresa', ''),
            l.get('empresa_cnpj', ''),
            l.get('funcionario', ''),
            l.get('tipo_vinculo_descricao', 'CLT'),
            l.get('funcionario_pis', ''),
            l.get('matricula', ''),
            l.get('cargo', ''),
            l.get('cbo', ''),
            _fmt_date(l.get('data_admissao')),
            _fmt_date(l.get('data_demissao')),
            _ano(l),
            l.get('competencia', ''),
            _decimal(l.get('base_fgts')),
            _decimal(l.get('valor_fgts')),
            l.get('parcela_13') or '',
            l.get('status_pagamento', ''),
            _fmt_date(l.get('data_pagamento')),
            _decimal(l.get('valor_pago')),
            l.get('fonte_confirmacao_pagamento', ''),
            _decimal(l.get('valor_atualizado')),
            l.get('status_vinculo', ''),
            l.get('observacoes', ''),
        ])

    col_widths = [10, 12, 30, 18, 14, 30, 14, 12, 20, 10, 10,
                  8, 12, 12, 12, 14, 16, 14, 12, 18, 20, 12, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    comp_i = resultado.get('competencia_inicio', '').replace('/', '_')
    comp_f = resultado.get('competencia_fim', '').replace('/', '_')
    resp = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="posicao_fgts_{comp_i}_a_{comp_f}.xlsx"'
    return resp
