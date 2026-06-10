"""
Serviço de importação de Extrato Analítico FGTS (CEF).

Parseia o arquivo XLSX emitido pelo portal Conectividade Social e confirma
os depósitos na plataforma, marcando Lancamentos como pago=True com
fonte_confirmacao_pagamento='extrato_analitico'.

Suporta dois formatos do XLSX:
  - Aba "Tratada": tabela estruturada com empresa_id, matricula, competencia, valor, data_pg
  - Aba "Original": relatório em texto com blocos por trabalhador (CNPJ + PIS + HISTORICO)
"""

from __future__ import annotations

import re
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Callable, Dict, List, Optional, Tuple

from django.db import transaction

from empresas.models import Empresa
from empresas.models_grupo import FuncionarioVinculo
from funcionarios.models import Funcionario
from lancamentos.models import Lancamento


class ExtratoImportError(Exception):
    pass


MESES_PT = {
    'JANEIRO': '01', 'FEVEREIRO': '02', 'MARCO': '03', 'MARCO': '03',
    'ABRIL': '04', 'MAIO': '05', 'JUNHO': '06', 'JULHO': '07',
    'AGOSTO': '08', 'SETEMBRO': '09', 'OUTUBRO': '10',
    'NOVEMBRO': '11', 'DEZEMBRO': '12',
}

# Linha após o cabeçalho PIS/PASEP DTA.ADM DATA DE AFAST.:
# "12345678901  DD/MM/YYYY  DD/MM/YYYY"  (datas podem estar em células separadas)
_RE_PIS_LINHA = re.compile(
    r'(\d{3}\.?\d{3}\.?\d{3}-?\d{2}|\d{11})'   # PIS (formatado ou não)
    r'\D+(\d{2}/\d{2}/\d{4})'                    # DTA.ADM
    r'(?:\D+(\d{2}/\d{2}/\d{4}))?',              # DATA DE AFAST. (opcional)
)

# "DEPOSITO [EM ATRASO] NOMEMES/ANO  valor"
_RE_DEPOSITO = re.compile(
    r'^DEPOSITO(?:\s+EM\s+ATRASO)?\s+([A-Z]+)/(\d{4})\s+([\d.,]+)',
    re.IGNORECASE,
)

# Tolerância percentual para comparação de valores (cobre diferenças de arredondamento CEF)
_TOLERANCIA_VALOR = Decimal('0.01')


def _extrair_deposito(linha: str) -> Optional[Tuple[str, Decimal, Optional[date], bool]]:
    """Retorna (competencia, valor, data_pg, em_atraso) ou None."""
    data_pg = _parse_data(linha[:10])
    resto = linha[10:].strip()
    m = _RE_DEPOSITO.match(resto)
    if not m:
        return None
    em_atraso = bool(m.group(0).upper().startswith('DEPOSITO EM ATRASO') or
                     'EM ATRASO' in m.group(0).upper())
    mes_nome = m.group(1).upper().replace('Ç', 'C').replace('Ã', 'A').replace('Á', 'A').replace('É', 'E')
    mes_num = MESES_PT.get(mes_nome)
    if not mes_num:
        return None
    valor = _parse_valor(m.group(3))
    if valor <= 0:
        return None
    return f'{mes_num}/{m.group(2)}', valor, data_pg, em_atraso


class RegistroExtrato:
    __slots__ = (
        'empresa_ref', 'matricula', 'pis', 'cnpj', 'nome',
        'competencia', 'data_pagamento', 'valor', 'em_atraso',
        'data_admissao', 'data_demissao',
    )

    def __init__(
        self,
        competencia: str,
        data_pagamento: Optional[date],
        valor: Decimal,
        empresa_ref: Any = None,
        matricula: str = '',
        pis: str = '',
        cnpj: str = '',
        nome: str = '',
        em_atraso: bool = False,
        data_admissao: Optional[date] = None,
        data_demissao: Optional[date] = None,
    ):
        self.empresa_ref = empresa_ref
        self.matricula = matricula
        self.pis = pis
        self.cnpj = cnpj
        self.nome = nome
        self.competencia = competencia
        self.data_pagamento = data_pagamento
        self.valor = valor
        self.em_atraso = em_atraso
        self.data_admissao = data_admissao
        self.data_demissao = data_demissao


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------

def _parse_data(s: str) -> Optional[date]:
    """Parseia DD/MM/YYYY."""
    try:
        return datetime.strptime(s.strip(), '%d/%m/%Y').date()
    except Exception:
        return None


def _parse_valor(s: str) -> Decimal:
    """Converte string BR (1.234,56) em Decimal."""
    try:
        limpo = str(s).strip().replace('.', '').replace(',', '.')
        return Decimal(limpo)
    except (InvalidOperation, ValueError):
        return Decimal('0.00')


def parse_tratada(ws) -> Tuple[List[RegistroExtrato], List[str]]:
    """
    Parseia a aba 'Tratada' do XLSX — formato tabular pré-processado.

    Colunas esperadas (índices 0-based):
      0: ID, 1: ignore, 2: empresa_codigo(int), 3: matricula,
      4: nome, 5: tipo, 6: data_pg(DD/MM/YYYY), 7: comp_texto,
      8: competencia(MM/YYYY), 9: valor(float), 10: matricula
    """
    registros: List[RegistroExtrato] = []
    erros: List[str] = []
    primeira_linha = True

    for row in ws.iter_rows(values_only=True):
        if primeira_linha:
            primeira_linha = False
            continue  # pula cabeçalho

        if not row or row[0] is None:
            continue

        try:
            empresa_id = int(row[2]) if row[2] is not None else None
            matricula = str(int(row[3])) if isinstance(row[3], (int, float)) else str(row[3] or '').strip()
            nome = str(row[4] or '').strip()
            data_pg_str = str(row[6] or '').strip()
            comp_raw = str(row[8] or '').strip()
            valor_raw = row[9]
        except Exception as exc:
            erros.append(f'Linha inválida: {exc}')
            continue

        # Competência: " 04/2018         " → "04/2018"
        comp = comp_raw.strip()
        if not re.match(r'^\d{2}/\d{4}$', comp):
            erros.append(f'Competência inválida: "{comp_raw}" — ignorada.')
            continue

        data_pg = _parse_data(data_pg_str)
        valor = _parse_valor(str(valor_raw)) if valor_raw is not None else Decimal('0.00')

        if valor <= 0:
            continue  # ignora estornos / JAM / saques

        registros.append(RegistroExtrato(
            empresa_ref=empresa_id,
            matricula=matricula,
            nome=nome,
            competencia=comp,
            data_pagamento=data_pg,
            valor=valor,
        ))

    return registros, erros


def parse_original(ws) -> Tuple[List[RegistroExtrato], List[str]]:
    """
    Parseia a aba 'Original' do XLSX — relatório em texto da CEF.

    Cada linha do XLSX pode ter conteúdo espalhado em várias colunas; todas são
    unidas com dois espaços para reconstruir a linha visual do relatório.

    Blocos por trabalhador:
      NOME DO TRABALHADOR → PIS/PASEP → INSCRICAO EMPREGADOR → HISTORICO DOS LANCAMENTOS
    """
    registros: List[RegistroExtrato] = []
    erros: List[str] = []

    # Une todas as células não-vazias de cada linha do XLSX
    linhas: List[str] = []
    for row in ws.iter_rows(values_only=True):
        partes = [str(c).strip() for c in row if c is not None and str(c).strip()]
        linhas.append('  '.join(partes))

    pis_atual = ''
    cnpj_atual = ''
    nome_atual = ''
    data_adm_atual: Optional[date] = None
    data_dem_atual: Optional[date] = None
    em_historico = False

    i = 0
    while i < len(linhas):
        linha = linhas[i]

        if linha.startswith('NOME DO TRABALHADOR') and 'NUM.CONTA' in linha:
            em_historico = False
            nome_linha = linhas[i + 1] if (i + 1) < len(linhas) else ''
            m_nome = re.match(r'^(.+?)\s{2,}\d{4,}', nome_linha)
            nome_atual = m_nome.group(1).strip() if m_nome else nome_linha.strip()
            i += 1

        elif 'PIS/PASEP' in linha and 'DTA.ADM' in linha:
            pis_linha = linhas[i + 1] if (i + 1) < len(linhas) else ''
            m_pis = _RE_PIS_LINHA.search(pis_linha)
            if m_pis:
                pis_raw = re.sub(r'\D', '', m_pis.group(1))
                pis_atual = pis_raw.zfill(11)
                data_adm_atual = _parse_data(m_pis.group(2))
                dem_raw = m_pis.group(3)
                # 00/00/0000 indica sem demissão
                data_dem_atual = _parse_data(dem_raw) if dem_raw and dem_raw != '00/00/0000' else None
            else:
                # Fallback: extrai só o PIS (sem datas)
                m_so_pis = re.match(r'^(\d{11})', pis_linha.replace(' ', ''))
                if m_so_pis:
                    pis_atual = m_so_pis.group(1)
                data_adm_atual = None
                data_dem_atual = None
            i += 1

        elif 'INSCRICAO EMPREGADOR' in linha:
            emp_linha = linhas[i + 1] if (i + 1) < len(linhas) else ''
            m_cnpj = re.search(r'(\d{14})\s*$', emp_linha)
            if m_cnpj:
                cnpj_atual = m_cnpj.group(1)
            i += 1

        elif 'HISTORICO DOS LANCAMENTOS' in linha:
            em_historico = True

        elif em_historico and linha:
            # Encerra histórico em marcadores de fim de seção
            if linha.startswith('*') or linha.startswith('#') or linha.startswith('SALDO ATUAL'):
                em_historico = False

            elif re.match(r'^\d{2}/\d{2}/\d{4}', linha):
                resultado = _extrair_deposito(linha)
                if resultado and pis_atual and cnpj_atual:
                    comp, valor, data_pg, em_atraso = resultado
                    registros.append(RegistroExtrato(
                        pis=pis_atual,
                        cnpj=cnpj_atual,
                        nome=nome_atual,
                        competencia=comp,
                        data_pagamento=data_pg,
                        valor=valor,
                        em_atraso=em_atraso,
                        data_admissao=data_adm_atual,
                        data_demissao=data_dem_atual,
                    ))

        i += 1

    return registros, erros


# ---------------------------------------------------------------------------
# Serviço principal
# ---------------------------------------------------------------------------

class ExtratoAnaliticoService:
    """
    Importador de Extrato Analítico da CEF.

    Detecta automaticamente se o XLSX contém a aba 'Tratada' (formato
    estruturado) ou se usa apenas a aba 'Original' (formato de relatório).
    """

    def __init__(self):
        self._empresa_cache: Dict[Any, Optional[Empresa]] = {}
        self._vinculo_cache: Dict[str, Optional[FuncionarioVinculo]] = {}

    # ------------------------------------------------------------------
    # API pública
    # ------------------------------------------------------------------

    def preview(
        self,
        xlsx_bytes: bytes,
        max_rows: int = 15,
    ) -> Dict[str, Any]:
        """Processa os primeiros max_rows registros sem gravar no banco."""
        registros, parse_erros = self._parse_xlsx(xlsx_bytes)
        total = len(registros)
        amostra = registros[:max_rows]

        rows = []
        for reg in amostra:
            empresa, vinculo, func, erro = self._resolver(reg)
            lancamento = None
            if vinculo and not erro:
                lancamento = Lancamento.objects.filter(
                    vinculo=vinculo,
                    competencia=reg.competencia,
                ).first()
                if not lancamento and func and empresa:
                    lancamento = Lancamento.objects.filter(
                        funcionario=func,
                        empresa=empresa,
                        competencia=reg.competencia,
                        vinculo__isnull=True,
                    ).first()

            rows.append({
                'pis': reg.pis,
                'cnpj': reg.cnpj,
                'matricula': reg.matricula,
                'nome': reg.nome,
                'competencia': reg.competencia,
                'data_pagamento': reg.data_pagamento.strftime('%d/%m/%Y') if reg.data_pagamento else '',
                'valor': str(reg.valor),
                'empresa_nome': empresa.nome if empresa else None,
                'funcionario_nome': func.nome if func else None,
                'lancamento_id': lancamento.pk if lancamento else None,
                'ja_confirmado_cef': (
                    lancamento.fonte_confirmacao_pagamento == 'extrato_analitico'
                    if lancamento else False
                ),
                'status': 'ok' if (lancamento and not erro) else 'error',
                'erro': erro or ('' if lancamento else 'Lançamento não encontrado na plataforma'),
            })

        linhas_ok = sum(1 for r in rows if r['status'] == 'ok')
        linhas_erro = sum(1 for r in rows if r['status'] == 'error')

        return {
            'total_registros': total,
            'linhas_amostradas': len(amostra),
            'linhas_ok': linhas_ok,
            'linhas_erro': linhas_erro,
            'rows': rows,
            'parse_erros': parse_erros[:10],
        }

    def importar(
        self,
        xlsx_bytes: bytes,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> Dict[str, Any]:
        """
        Confirma pagamentos em todos os registros do extrato.
        Marca: pago=True, fonte_confirmacao_pagamento='extrato_analitico'.

        Quando uma competência possui dois lançamentos (mensal + 13º), compara
        o valor do extrato com a soma de ambos para decidir se dá baixa nos dois.
        """
        registros, parse_erros = self._parse_xlsx(xlsx_bytes)
        total = len(registros)

        if progress_callback:
            progress_callback(0, total)

        confirmados = 0
        confirmados_com_13 = 0
        nao_encontrados = 0
        ja_confirmados = 0
        erros: List[str] = list(parse_erros)
        avisos: List[str] = []
        rows: List[Dict[str, Any]] = []

        def _row_base(reg: RegistroExtrato, empresa=None) -> Dict[str, Any]:
            return {
                'empresa': empresa.nome if empresa else (reg.cnpj or ''),
                'cnpj': reg.cnpj,
                'funcionario': reg.nome,
                'pis': reg.pis,
                'matricula': reg.matricula,
                'competencia': reg.competencia,
                'valor': str(reg.valor),
                'tipo': 'DEPOSITO EM ATRASO' if reg.em_atraso else 'DEPOSITO',
            }

        for idx, reg in enumerate(registros, 1):
            empresa, vinculo, func, erro = self._resolver(reg)

            if erro or not empresa:
                msg = erro or "empresa não identificada"
                erros.append(f'Reg {idx} ({reg.nome} / {reg.competencia}): {msg}')
                rows.append({**_row_base(reg, empresa), 'status': 'erro', 'detalhe': msg})
                continue

            lancamentos = self._buscar_lancamentos(vinculo, func, empresa, reg.competencia)
            if not lancamentos:
                nao_encontrados += 1
                detalhe = f'Lançamento não encontrado para {reg.nome} / {reg.competencia}'
                avisos.append(f'Reg {idx}: sem lançamento para {reg.nome} / {reg.competencia} — ignorado.')
                rows.append({**_row_base(reg, empresa), 'status': 'nao_encontrado', 'detalhe': detalhe})
                continue

            # Separa lançamentos já confirmados dos pendentes
            pendentes = [l for l in lancamentos if l.fonte_confirmacao_pagamento != 'extrato_analitico']
            if not pendentes:
                ja_confirmados += 1
                rows.append({**_row_base(reg, empresa), 'status': 'ja_confirmado', 'detalhe': 'Já confirmado pelo extrato CEF'})
                continue

            # Caso com dois lançamentos: possível mensal + 13º, ou dois vínculos distintos
            if len(pendentes) == 2:
                mensal_list = [l for l in pendentes if l.parcela_13 is None]
                decimo_list = [l for l in pendentes if l.parcela_13 is not None]

                # Sub-caso: dois lançamentos mensais (vínculos distintos na mesma competência)
                # Seleciona o que corresponde ao vínculo resolvido para este bloco do extrato
                if len(mensal_list) == 2 and not decimo_list:
                    vinculo_pk = vinculo.pk if vinculo else None
                    alvo = next(
                        (l for l in mensal_list if l.vinculo_id == vinculo_pk),
                        mensal_list[0],
                    )
                    try:
                        with transaction.atomic():
                            Lancamento.objects.filter(pk=alvo.pk).update(
                                pago=True,
                                fonte_confirmacao_pagamento='extrato_analitico',
                                data_pagto=reg.data_pagamento or alvo.data_pagto,
                                valor_pago=reg.valor if reg.valor > 0 else alvo.valor_pago,
                            )
                        confirmados += 1
                        rows.append({**_row_base(reg, empresa), 'status': 'confirmado', 'detalhe': 'Confirmado (múltiplos vínculos)'})
                    except Exception as exc:
                        erros.append(f'Reg {idx}: erro ao salvar — {exc}')
                        rows.append({**_row_base(reg, empresa), 'status': 'erro', 'detalhe': str(exc)})
                    if progress_callback and idx % 50 == 0:
                        progress_callback(idx, total)
                    continue

                if mensal_list and decimo_list:
                    mensal = mensal_list[0]
                    decimo = decimo_list[0]
                    soma_ambos = (mensal.valor_fgts or Decimal('0')) + (decimo.valor_fgts or Decimal('0'))

                    if soma_ambos > 0 and abs(reg.valor - soma_ambos) / soma_ambos <= _TOLERANCIA_VALOR:
                        # Valor do extrato cobre mensal + 13º — baixa nos dois
                        try:
                            with transaction.atomic():
                                Lancamento.objects.filter(pk=mensal.pk).update(
                                    pago=True,
                                    fonte_confirmacao_pagamento='extrato_analitico',
                                    data_pagto=reg.data_pagamento or mensal.data_pagto,
                                    valor_pago=mensal.valor_fgts,
                                )
                                Lancamento.objects.filter(pk=decimo.pk).update(
                                    pago=True,
                                    fonte_confirmacao_pagamento='extrato_analitico',
                                    data_pagto=reg.data_pagamento or decimo.data_pagto,
                                    valor_pago=decimo.valor_fgts,
                                )
                            confirmados_com_13 += 1
                            rows.append({**_row_base(reg, empresa), 'status': 'confirmado', 'detalhe': 'Confirmado (mensal + 13º)'})
                        except Exception as exc:
                            erros.append(f'Reg {idx}: erro ao salvar mensal+13º — {exc}')
                            rows.append({**_row_base(reg, empresa), 'status': 'erro', 'detalhe': str(exc)})

                        if progress_callback and idx % 50 == 0:
                            progress_callback(idx, total)
                        continue

                    # Valor cobre apenas o mensal (13º separado ou pago em outra guia)
                    valor_mensal = mensal.valor_fgts or Decimal('0')
                    if valor_mensal > 0 and abs(reg.valor - valor_mensal) / valor_mensal <= _TOLERANCIA_VALOR:
                        try:
                            with transaction.atomic():
                                Lancamento.objects.filter(pk=mensal.pk).update(
                                    pago=True,
                                    fonte_confirmacao_pagamento='extrato_analitico',
                                    data_pagto=reg.data_pagamento or mensal.data_pagto,
                                    valor_pago=reg.valor,
                                )
                            confirmados += 1
                            rows.append({**_row_base(reg, empresa), 'status': 'confirmado', 'detalhe': 'Confirmado (mensal); 13º permanece em aberto'})
                        except Exception as exc:
                            erros.append(f'Reg {idx}: erro ao salvar mensal — {exc}')
                            rows.append({**_row_base(reg, empresa), 'status': 'erro', 'detalhe': str(exc)})
                        avisos.append(
                            f'Reg {idx}: {reg.nome} / {reg.competencia} — '
                            f'valor do extrato ({reg.valor}) corresponde só ao mensal; '
                            f'13º permanece em aberto.'
                        )
                        if progress_callback and idx % 50 == 0:
                            progress_callback(idx, total)
                        continue

                    # Valor não bate com nenhum padrão esperado
                    detalhe_aviso = (
                        f'Valor do extrato ({reg.valor}) não corresponde ao mensal '
                        f'({mensal.valor_fgts}) nem à soma mensal+13º ({soma_ambos}). '
                        f'Verifique manualmente.'
                    )
                    avisos.append(f'Reg {idx}: {reg.nome} / {reg.competencia} — {detalhe_aviso}')
                    nao_encontrados += 1
                    rows.append({**_row_base(reg, empresa), 'status': 'nao_encontrado', 'detalhe': detalhe_aviso})
                    if progress_callback and idx % 50 == 0:
                        progress_callback(idx, total)
                    continue

            # Caso padrão: um único lançamento pendente (ou múltiplos sem padrão mensal+13º)
            lancamento = pendentes[0]
            try:
                with transaction.atomic():
                    Lancamento.objects.filter(pk=lancamento.pk).update(
                        pago=True,
                        fonte_confirmacao_pagamento='extrato_analitico',
                        data_pagto=reg.data_pagamento or lancamento.data_pagto,
                        valor_pago=reg.valor if reg.valor > 0 else lancamento.valor_pago,
                    )
                confirmados += 1
                rows.append({**_row_base(reg, empresa), 'status': 'confirmado', 'detalhe': 'Confirmado'})
            except Exception as exc:
                erros.append(f'Reg {idx}: erro ao salvar — {exc}')
                rows.append({**_row_base(reg, empresa), 'status': 'erro', 'detalhe': str(exc)})

            if progress_callback and idx % 50 == 0:
                progress_callback(idx, total)

        if progress_callback:
            progress_callback(total, total)

        return {
            'confirmados': confirmados,
            'confirmados_com_13': confirmados_com_13,
            'nao_encontrados': nao_encontrados,
            'ja_confirmados': ja_confirmados,
            'erros': erros,
            'avisos': avisos,
            'total': total,
            'rows': rows,
        }

    # ------------------------------------------------------------------
    # Helpers internos
    # ------------------------------------------------------------------

    def _parse_xlsx(self, xlsx_bytes: bytes) -> Tuple[List[RegistroExtrato], List[str]]:
        try:
            import openpyxl
        except ImportError:
            raise ExtratoImportError('A biblioteca openpyxl não está instalada.')

        try:
            wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
        except Exception as exc:
            raise ExtratoImportError(f'Não foi possível abrir o arquivo XLSX: {exc}')

        if 'Tratada' in wb.sheetnames:
            return parse_tratada(wb['Tratada'])

        # Tenta a primeira aba (Original)
        ws = wb.active
        return parse_original(ws)

    def _resolver(
        self, reg: RegistroExtrato
    ) -> Tuple[Optional[Empresa], Optional[FuncionarioVinculo], Optional[Funcionario], str]:
        """Resolve Empresa + FuncionarioVinculo + Funcionario para um registro."""
        empresa = self._resolver_empresa(reg)
        if not empresa:
            return None, None, None, 'Empresa não encontrada na plataforma.'

        vinculo, func = self._resolver_vinculo(reg, empresa)
        if not func:
            return empresa, None, None, f'Funcionário não encontrado (PIS:{reg.pis} / mat:{reg.matricula}).'

        return empresa, vinculo, func, ''

    def _resolver_empresa(self, reg: RegistroExtrato) -> Optional[Empresa]:
        # Aba Tratada: usa empresa_ref (empresa.codigo)
        if reg.empresa_ref is not None:
            key = f'id:{reg.empresa_ref}'
            if key not in self._empresa_cache:
                self._empresa_cache[key] = Empresa.objects.filter(codigo=reg.empresa_ref).first()
            return self._empresa_cache[key]

        # Aba Original: usa CNPJ
        cnpj_limpo = re.sub(r'\D', '', reg.cnpj)
        if cnpj_limpo not in self._empresa_cache:
            self._empresa_cache[cnpj_limpo] = Empresa.objects.filter(cnpj=cnpj_limpo).first()
        return self._empresa_cache[cnpj_limpo]

    def _resolver_vinculo(
        self, reg: RegistroExtrato, empresa: Empresa
    ) -> Tuple[Optional[FuncionarioVinculo], Optional[Funcionario]]:
        # Aba Tratada: usa matrícula — sem ambiguidade
        if reg.matricula:
            key = f'{empresa.pk}:mat:{reg.matricula}'
            if key not in self._vinculo_cache:
                v = FuncionarioVinculo.objects.filter(
                    empresa=empresa,
                    matricula=reg.matricula,
                ).select_related('funcionario').first()
                self._vinculo_cache[key] = v
            vinculo = self._vinculo_cache[key]
            if vinculo:
                return vinculo, vinculo.funcionario

        # Aba Original: usa PIS + datas de admissão/demissão (chave composta)
        if reg.pis:
            pis_limpo = re.sub(r'\D', '', reg.pis).zfill(11)

            # Chave de cache: inclui datas quando disponíveis para suportar duplo vínculo
            adm_str = reg.data_admissao.isoformat() if reg.data_admissao else ''
            dem_str = reg.data_demissao.isoformat() if reg.data_demissao else ''
            key = f'{empresa.pk}:pis:{pis_limpo}:adm:{adm_str}:dem:{dem_str}'

            if key not in self._vinculo_cache:
                vinculos = list(
                    FuncionarioVinculo.objects.filter(
                        funcionario__pis=pis_limpo,
                        empresa=empresa,
                    ).select_related('funcionario').order_by('-data_admissao')
                )

                if not vinculos:
                    self._vinculo_cache[key] = None
                elif len(vinculos) == 1:
                    self._vinculo_cache[key] = vinculos[0]
                else:
                    # Duplo vínculo: tenta casar pela data de admissão do extrato
                    matched = self._casar_vinculo_por_datas(vinculos, reg.data_admissao, reg.data_demissao)
                    self._vinculo_cache[key] = matched

            v = self._vinculo_cache[key]
            if v:
                return v, v.funcionario

        return None, None

    @staticmethod
    def _casar_vinculo_por_datas(
        vinculos: List['FuncionarioVinculo'],
        data_adm: Optional[date],
        data_dem: Optional[date],
    ) -> Optional['FuncionarioVinculo']:
        """
        Dado uma lista de vínculos com duplo contrato, tenta identificar o correto
        usando a chave composta data_admissao + data_demissao do extrato.

        Estratégia:
        1. Match exato em ambas as datas
        2. Match só pela admissão (demissão pode estar ausente no extrato)
        3. Match por mês/ano da admissão (CEF pode usar dia diferente do cadastro)
        4. Vínculo cuja janela (adm → dem) contém a data de admissão do extrato
        5. Fallback: vínculo mais recente
        """
        if not data_adm:
            return vinculos[0]  # sem data, retorna o mais recente

        # Tentativa 1: match exato admissão + demissão
        if data_dem:
            for v in vinculos:
                if v.data_admissao == data_adm and v.data_demissao == data_dem:
                    return v

        # Tentativa 2: match só pela admissão (data exata)
        for v in vinculos:
            if v.data_admissao == data_adm:
                return v

        # Tentativa 3: match por mês/ano da admissão (CEF pode reportar dia diferente do cadastro)
        for v in vinculos:
            if (v.data_admissao and
                    v.data_admissao.year == data_adm.year and
                    v.data_admissao.month == data_adm.month):
                return v

        # Tentativa 4: vínculo cuja janela (adm → dem) contém a data de admissão do extrato
        for v in vinculos:
            if v.data_admissao and v.data_admissao <= data_adm:
                if v.data_demissao is None or v.data_demissao >= data_adm:
                    return v

        # Fallback: mais recente
        return vinculos[0]

    def _buscar_lancamentos(
        self,
        vinculo: Optional[FuncionarioVinculo],
        func: Optional[Funcionario],
        empresa: Empresa,
        competencia: str,
    ) -> List[Lancamento]:
        """Retorna todos os Lancamentos do funcionário na empresa/competência (qualquer vínculo).

        Busca primeiro por funcionário+empresa para garantir que lançamentos de
        todos os vínculos ativos sejam retornados — crítico quando o colaborador
        tem múltiplos contratos simultâneos na mesma empresa.
        """
        if func:
            lancs = list(Lancamento.objects.filter(
                funcionario=func,
                empresa=empresa,
                competencia=competencia,
            ))
            if lancs:
                return lancs

        # Fallback legado: lançamentos sem funcionário vinculado diretamente
        if vinculo:
            return list(Lancamento.objects.filter(
                vinculo=vinculo,
                competencia=competencia,
                funcionario__isnull=True,
            ))

        return []
