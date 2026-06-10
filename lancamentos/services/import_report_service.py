"""
Geração de relatórios XLSX on-demand para importações.

Cada função recebe uma instância de model de importação (já com resultado_json
preenchido) e retorna bytes do workbook para download direto pelo browser.
"""

from __future__ import annotations

from collections import Counter
from io import BytesIO
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Constantes de estilo
# ---------------------------------------------------------------------------

_FILL_HEADER = PatternFill('solid', fgColor='003A78')
_FILL_OK     = PatternFill('solid', fgColor='C6EFCE')
_FILL_WARN   = PatternFill('solid', fgColor='FFEB9C')
_FILL_ERR    = PatternFill('solid', fgColor='FFC7CE')
_FILL_SKIP   = PatternFill('solid', fgColor='EEEEEE')
_FILL_JA     = PatternFill('solid', fgColor='DDEBF7')  # já confirmado (azul claro)
_FILL_NAO    = PatternFill('solid', fgColor='FCE4D6')  # não encontrado (laranja claro)

_FONT_HEADER = Font(bold=True, color='FFFFFF')
_FONT_ERR    = Font(color='9C0006')
_FONT_OK     = Font(color='276221')
_FONT_WARN   = Font(color='9C6500')

_STATUS_FILL = {
    'ok': _FILL_OK,
    'criado': _FILL_OK,
    'atualizado': _FILL_OK,
    'confirmado': _FILL_OK,
    'ignorado': _FILL_SKIP,
    'ja_confirmado': _FILL_JA,
    'nao_encontrado': _FILL_NAO,
    'erro': _FILL_ERR,
}

_STATUS_LABEL = {
    'ok': 'OK',
    'criado': 'Criado',
    'atualizado': 'Atualizado',
    'confirmado': 'Confirmado',
    'ignorado': 'Ignorado',
    'ja_confirmado': 'Já Confirmado',
    'nao_encontrado': 'Não Encontrado',
    'erro': 'Erro',
}


def _header_row(ws, cols: List[str], row: int = 1):
    for c, title in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=title)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _auto_width(ws, min_width: int = 10, max_width: int = 60):
    for col_cells in ws.columns:
        length = max(
            len(str(cell.value or '')) for cell in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max(length + 2, min_width), max_width
        )


def _title_block(ws, titulo: str, subtitulo: str = ''):
    ws.cell(row=1, column=1, value=titulo).font = Font(bold=True, size=14)
    if subtitulo:
        ws.cell(row=2, column=1, value=subtitulo).font = Font(italic=True, color='555555')


# ---------------------------------------------------------------------------
# Relatório de Lançamentos
# ---------------------------------------------------------------------------

def gerar_relatorio_lancamentos(importacao) -> bytes:
    """
    Gera XLSX com abas Analítico e Sintético para uma ImportacaoLancamento.
    Retorna bytes prontos para HttpResponse.
    """
    wb = openpyxl.Workbook()
    resultado = importacao.resultado_json or {}
    rows: List[Dict[str, Any]] = resultado.get('rows', [])

    _build_analitico_lancamentos(wb.active, rows, importacao)
    _build_sintetico_lancamentos(wb.create_sheet('Sintético'), resultado, importacao)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_analitico_lancamentos(ws, rows: List[Dict], importacao):
    ws.title = 'Analítico'

    cols = ['Linha', 'Empresa', 'CPF', 'Nome', 'Competência', 'Base FGTS', 'Ação', 'Status', 'Detalhe']
    _header_row(ws, cols)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    for r, row in enumerate(rows, 2):
        status = row.get('status', 'ok')
        fill = _STATUS_FILL.get(status, _FILL_SKIP)

        vals = [
            row.get('linha', ''),
            row.get('empresa', ''),
            row.get('cpf', ''),
            row.get('nome', ''),
            row.get('competencia', ''),
            row.get('base_fgts', ''),
            row.get('acao', '').capitalize(),
            _STATUS_LABEL.get(status, status),
            row.get('detalhe', ''),
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = fill
            if status == 'erro':
                cell.font = _FONT_ERR

    _auto_width(ws)

    # Se não há rows (importação antiga sem dados por linha), avisa
    if not rows:
        ws.cell(row=2, column=1, value='Dados analíticos não disponíveis para esta importação.')


def _build_sintetico_lancamentos(ws, resultado: Dict, importacao):
    ws.title = 'Sintético'

    # Cabeçalho informativo
    ws.cell(row=1, column=1, value='RELATÓRIO SINTÉTICO — IMPORTAÇÃO DE LANÇAMENTOS').font = Font(bold=True, size=13)
    ws.merge_cells('A1:E1')

    info = [
        ('Arquivo', importacao.nome_arquivo),
        ('Empresa', importacao.empresa.nome if importacao.empresa else '(múltiplas)'),
        ('Usuário', str(importacao.usuario)),
        ('Data/Hora', importacao.atualizado_em.strftime('%d/%m/%Y %H:%M') if importacao.atualizado_em else ''),
    ]
    for i, (label, value) in enumerate(info, 3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    # Tabela de totais
    row_totais = 8
    ws.cell(row=row_totais, column=1, value='TOTAIS').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=row_totais, column=1).fill = _FILL_HEADER
    ws.cell(row=row_totais, column=2, value='Quantidade').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=row_totais, column=2).fill = _FILL_HEADER

    totais = [
        ('Criados', resultado.get('created', 0), _FILL_OK),
        ('Atualizados', resultado.get('updated', 0), _FILL_OK),
        ('Ignorados', resultado.get('skipped', 0), _FILL_SKIP),
        ('Erros', len(resultado.get('errors', [])), _FILL_ERR),
        ('Total de linhas', (resultado.get('success', 0) + resultado.get('skipped', 0) + len(resultado.get('errors', []))), _FILL_HEADER),
    ]
    for i, (label, qty, fill) in enumerate(totais, row_totais + 1):
        ws.cell(row=i, column=1, value=label).fill = fill
        ws.cell(row=i, column=2, value=qty).fill = fill
        ws.cell(row=i, column=2).alignment = Alignment(horizontal='center')

    # Erros agrupados
    errors = resultado.get('errors', [])
    if errors:
        row_erros = row_totais + len(totais) + 2
        ws.cell(row=row_erros, column=1, value='ERROS ENCONTRADOS').font = Font(bold=True, color='FFFFFF')
        ws.cell(row=row_erros, column=1).fill = PatternFill('solid', fgColor='9C0006')
        ws.merge_cells(f'A{row_erros}:E{row_erros}')

        ws.cell(row=row_erros + 1, column=1, value='Linha').font = Font(bold=True)
        ws.cell(row=row_erros + 1, column=2, value='Mensagem de Erro').font = Font(bold=True)

        for j, err in enumerate(errors, row_erros + 2):
            ws.cell(row=j, column=1, value=err.get('row', ''))
            cell = ws.cell(row=j, column=2, value=err.get('error', ''))
            cell.fill = _FILL_ERR
            cell.font = _FONT_ERR
            cell.alignment = Alignment(wrap_text=True)

        # Erros agrupados por mensagem
        row_agrup = row_erros + 2 + len(errors) + 2
        ws.cell(row=row_agrup, column=1, value='ERROS AGRUPADOS POR TIPO').font = Font(bold=True, color='FFFFFF')
        ws.cell(row=row_agrup, column=1).fill = PatternFill('solid', fgColor='9C0006')
        ws.merge_cells(f'A{row_agrup}:E{row_agrup}')
        ws.cell(row=row_agrup + 1, column=1, value='Mensagem').font = Font(bold=True)
        ws.cell(row=row_agrup + 1, column=2, value='Ocorrências').font = Font(bold=True)

        contagem = Counter(e.get('error', '') for e in errors)
        for k, (msg, cnt) in enumerate(contagem.most_common(), row_agrup + 2):
            ws.cell(row=k, column=1, value=msg).alignment = Alignment(wrap_text=True)
            ws.cell(row=k, column=2, value=cnt).alignment = Alignment(horizontal='center')

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Relatório de Extrato Analítico
# ---------------------------------------------------------------------------

def gerar_relatorio_extrato(importacao) -> bytes:
    """
    Gera XLSX com abas Analítico e Sintético para uma ImportacaoExtratoAnalitico.
    Retorna bytes prontos para HttpResponse.
    """
    wb = openpyxl.Workbook()
    resultado = importacao.resultado_json or {}
    rows: List[Dict[str, Any]] = resultado.get('rows', [])

    _build_analitico_extrato(wb.active, rows, importacao)
    _build_sintetico_extrato(wb.create_sheet('Sintético'), resultado, importacao)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_analitico_extrato(ws, rows: List[Dict], importacao):
    ws.title = 'Analítico'

    cols = ['Empresa', 'CNPJ', 'Funcionário', 'PIS', 'Matrícula', 'Competência', 'Valor (R$)', 'Tipo', 'Status', 'Detalhe']
    _header_row(ws, cols)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    for r, row in enumerate(rows, 2):
        status = row.get('status', '')
        fill = _STATUS_FILL.get(status, _FILL_SKIP)

        try:
            valor_fmt = float(row.get('valor', 0))
        except (TypeError, ValueError):
            valor_fmt = row.get('valor', '')

        vals = [
            row.get('empresa', ''),
            row.get('cnpj', ''),
            row.get('funcionario', ''),
            row.get('pis', ''),
            row.get('matricula', ''),
            row.get('competencia', ''),
            valor_fmt,
            row.get('tipo', ''),
            _STATUS_LABEL.get(status, status),
            row.get('detalhe', ''),
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = fill
            if status == 'erro':
                cell.font = _FONT_ERR
            if c == 7 and isinstance(val, float):
                cell.number_format = '#,##0.00'

    _auto_width(ws)

    if not rows:
        ws.cell(row=2, column=1, value='Dados analíticos não disponíveis para esta importação.')


def _build_sintetico_extrato(ws, resultado: Dict, importacao):
    ws.title = 'Sintético'

    ws.cell(row=1, column=1, value='RELATÓRIO SINTÉTICO — IMPORTAÇÃO EXTRATO ANALÍTICO CEF').font = Font(bold=True, size=13)
    ws.merge_cells('A1:E1')

    info = [
        ('Arquivo', importacao.nome_arquivo),
        ('Usuário', str(importacao.usuario)),
        ('Data/Hora', importacao.atualizado_em.strftime('%d/%m/%Y %H:%M') if importacao.atualizado_em else ''),
    ]
    for i, (label, value) in enumerate(info, 3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)

    row_totais = 7
    ws.cell(row=row_totais, column=1, value='TOTAIS').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=row_totais, column=1).fill = _FILL_HEADER
    ws.cell(row=row_totais, column=2, value='Quantidade').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=row_totais, column=2).fill = _FILL_HEADER

    erros_list = resultado.get('erros', [])
    n_erros = len(erros_list) if isinstance(erros_list, list) else int(erros_list or 0)

    totais = [
        ('Confirmados', resultado.get('confirmados', 0), _FILL_OK),
        ('Confirmados com 13º', resultado.get('confirmados_com_13', 0), _FILL_OK),
        ('Já Confirmados (duplicata)', resultado.get('ja_confirmados', 0), _FILL_JA),
        ('Não Encontrados', resultado.get('nao_encontrados', 0), _FILL_NAO),
        ('Erros', n_erros, _FILL_ERR),
        ('Total de registros', resultado.get('total', 0), _FILL_HEADER),
    ]
    for i, (label, qty, fill) in enumerate(totais, row_totais + 1):
        ws.cell(row=i, column=1, value=label).fill = fill
        ws.cell(row=i, column=2, value=qty).fill = fill
        ws.cell(row=i, column=2).alignment = Alignment(horizontal='center')

    # Funcionários não encontrados
    rows: List[Dict] = resultado.get('rows', [])
    nao_encontrados = [r for r in rows if r.get('status') == 'nao_encontrado']
    if nao_encontrados:
        row_nao = row_totais + len(totais) + 2
        ws.cell(row=row_nao, column=1, value='REGISTROS NÃO ENCONTRADOS').font = Font(bold=True, color='FFFFFF')
        ws.cell(row=row_nao, column=1).fill = PatternFill('solid', fgColor='C65B11')
        ws.merge_cells(f'A{row_nao}:E{row_nao}')

        sub_cols = ['Funcionário', 'PIS', 'Matrícula', 'Empresa', 'Competência', 'Detalhe']
        for c, col in enumerate(sub_cols, 1):
            ws.cell(row=row_nao + 1, column=c, value=col).font = Font(bold=True)

        for j, r in enumerate(nao_encontrados, row_nao + 2):
            ws.cell(row=j, column=1, value=r.get('funcionario', '')).fill = _FILL_NAO
            ws.cell(row=j, column=2, value=r.get('pis', '')).fill = _FILL_NAO
            ws.cell(row=j, column=3, value=r.get('matricula', '')).fill = _FILL_NAO
            ws.cell(row=j, column=4, value=r.get('empresa', '')).fill = _FILL_NAO
            ws.cell(row=j, column=5, value=r.get('competencia', '')).fill = _FILL_NAO
            ws.cell(row=j, column=6, value=r.get('detalhe', '')).fill = _FILL_NAO

    # Erros
    erros_texto = [e for e in erros_list] if isinstance(erros_list, list) else []
    if erros_texto:
        row_err = row_totais + len(totais) + 2 + (len(nao_encontrados) + 3 if nao_encontrados else 0)
        ws.cell(row=row_err, column=1, value='ERROS').font = Font(bold=True, color='FFFFFF')
        ws.cell(row=row_err, column=1).fill = PatternFill('solid', fgColor='9C0006')
        ws.merge_cells(f'A{row_err}:E{row_err}')
        for j, msg in enumerate(erros_texto, row_err + 1):
            cell = ws.cell(row=j, column=1, value=msg)
            cell.fill = _FILL_ERR
            cell.font = _FONT_ERR
            cell.alignment = Alignment(wrap_text=True)
            ws.merge_cells(f'A{j}:E{j}')

    _auto_width(ws)
