from datetime import datetime
from dateutil.relativedelta import relativedelta
from io import BytesIO
import re
import unicodedata

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.utils.timezone import make_aware
from django.utils import timezone
from django.core.exceptions import ValidationError
from decimal import Decimal, InvalidOperation
from lancamentos.models import Lancamento
from empresas.models import Empresa
from funcionarios.models import Funcionario
from billing.models import BillingCustomer
from fgtsweb.mixins import is_empresa_allowed, get_allowed_empresa_ids
from empresas.models_grupo import FuncionarioVinculo, get_aliquota_fgts


class LancamentoImportService:
    """Serviço para gerenciar importação e exportação de lançamentos FGTS em XLSX"""
    
    REQUIRED_COLUMNS = [
        'CPF_FUNCIONARIO', 'NOME_FUNCIONARIO', 'COMPETENCIA', 'BASE_FGTS'
    ]
    
    OPTIONAL_COLUMNS = [
        # EMPRESA é opcional, mas recomendado para grupos com múltiplos vínculos ativos.
        # Quando informado, deve ser o codigo folha da empresa.
        'EMPRESA',
        # MATRICULA é opcional, mas recomendado para identificar a "cadeira" sem depender de CPF.
        # Deve ser a matrícula do vínculo dentro da empresa (numérica; não reutilizável).
        'MATRICULA',
        'VALOR_FGTS', 'PAGO', 'DATA_PAGTO', 'VALOR_PAGO', 'PARCELA_13'
    ]

    @staticmethod
    def _normalize_header(header_value: object) -> str:
        if header_value is None:
            return ''

        text = str(header_value).strip()
        if not text:
            return ''

        # Remove acentos e normaliza separadores para facilitar a vida do usuário.
        text = unicodedata.normalize('NFKD', text)
        text = text.encode('ascii', 'ignore').decode('ascii')
        text = text.upper()
        text = re.sub(r'[^A-Z0-9]+', '_', text)
        text = text.strip('_')
        return text

    @staticmethod
    def _canonicalize_header(normalized_header: str) -> str:
        if not normalized_header:
            return ''

        aliases = {
            # EMPRESA / CodEmpresa
            'CODEMPRESA': 'EMPRESA',
            'COD_EMPRESA': 'EMPRESA',
            'CODIGO_EMPRESA': 'EMPRESA',
            # VINCULO / ID_VINCULO
            'ID_VINCULO': 'VINCULO',
            'VINCULO_ID': 'VINCULO',
            # PARCELA_13
            'PARCELA13': 'PARCELA_13',
            'PARCELA_13': 'PARCELA_13',
        }
        return aliases.get(normalized_header, normalized_header)

    @staticmethod
    def _resolve_empresa_from_identifier(value):
        if value is None:
            return None

        raw = str(value).strip()
        if not raw:
            return None

        if raw.endswith('.0') and raw.replace('.', '', 1).isdigit():
            raw = str(int(float(raw)))

        qs = Empresa.objects.filter(codigo_folha__iexact=raw)
        results = list(qs[:2])  # busca no máximo 2 para detectar duplicidade com 1 query
        if len(results) > 1:
            raise ValueError(f"Codigo Folha '{raw}' duplicado. Contate o administrador.")
        if results:
            return results[0]

        if raw.isdigit():
            try:
                return Empresa.objects.get(codigo=int(raw))
            except Empresa.DoesNotExist:
                return None

        return None
    
    @staticmethod
    def generate_template_xlsx():
        """Gera um arquivo XLSX com o modelo para importação de lançamentos"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lançamentos FGTS"
        
        # Definir estilos
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        example_fill = PatternFill(start_color="E7E9FF", end_color="E7E9FF", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers com colunas obrigatórias e opcionais
        # Observação: o import aceita aliases (ex: COD_EMPRESA -> EMPRESA).
        # Aqui preferimos exibir nomes mais “autoexplicativos” para reduzir dúvidas.
        all_columns_internal = LancamentoImportService.REQUIRED_COLUMNS + LancamentoImportService.OPTIONAL_COLUMNS
        last_col_letter = openpyxl.utils.get_column_letter(len(all_columns_internal))

        display_overrides = {
            'EMPRESA': 'COD_EMPRESA',
        }

        for col_idx, column_name in enumerate(all_columns_internal, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = display_overrides.get(column_name, column_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
        
        # Linha de exemplo
        example_data = [
            '',                      # CPF_FUNCIONARIO (opcional se MATRICULA estiver preenchida)
            'João da Silva',         # NOME_FUNCIONARIO
            '01/2026',              # COMPETENCIA (MM/YYYY)
            '3500.00',              # BASE_FGTS
            'CFABC1234',            # EMPRESA (codigo folha)
            '1001',                 # MATRICULA (do vínculo)
            '280.00',               # VALOR_FGTS (8% da base)
            'NÃO',                  # PAGO (SIM/NÃO)
            '',                     # DATA_PAGTO (dd/mm/yyyy)
            '',                     # VALOR_PAGO
            '',                     # PARCELA_13
        ]
        
        for col_idx, value in enumerate(example_data, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = value
            cell.fill = example_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Instruções (uma área única, editável, para reduzir confusão)
        ws.merge_cells(f'A4:{last_col_letter}4')
        title_cell = ws['A4']
        title_cell.value = "INSTRUÇÕES DE PREENCHIMENTO (LEIA ANTES DE IMPORTAR)"
        title_cell.font = Font(bold=True, size=12, color="667eea")
        title_cell.alignment = Alignment(horizontal='left', vertical='center')

        instructions_lines = [
            "Fluxo recomendado (evita vínculo ambíguo): preencha EMPRESA (Codigo Folha) + MATRICULA.",
            "- CPF_FUNCIONARIO: CPF do colaborador (apenas números). Pode ficar em branco se MATRICULA estiver preenchida.",
            "- NOME_FUNCIONARIO: Nome completo (apenas para conferência).",
            "- COMPETENCIA: MM/YYYY (ex: 01/2026).",
            "- BASE_FGTS: Base de cálculo (ex: 3500.00). Use ponto como separador decimal.",
            "- COD_EMPRESA (Codigo Folha): código da empresa do vínculo nesta linha (recomendado em grupos/múltiplas empresas).",
            "- MATRICULA: matrícula do vínculo (cadeira) na empresa. Se houver mais de um vínculo ativo na mesma competência, informe a MATRICULA.",
            "- PARCELA_13: opcional. Use 1 (1ª parcela) ou 2 (2ª parcela). Deixe em branco para competência normal.",
            "",
            "Observações:",
            "- O sistema aceita tanto EMPRESA quanto COD_EMPRESA (mesma coisa).",
            "- O colaborador e o vínculo precisam existir e estar ativos na competência.",
            "- Remova a linha de exemplo (linha 2) antes de importar.",
        ]

        instructions_text = "\n".join(instructions_lines)
        instructions_start_row = 5
        instructions_end_row = 20
        ws.merge_cells(f'A{instructions_start_row}:{last_col_letter}{instructions_end_row}')
        body_cell = ws.cell(row=instructions_start_row, column=1)
        body_cell.value = instructions_text
        body_cell.font = Font(size=10)
        body_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        for r in range(instructions_start_row, instructions_end_row + 1):
            ws.row_dimensions[r].height = 18
        
        # Retornar bytes do arquivo
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    @staticmethod
    def preview_lancamentos_from_file(file, empresa, user, max_rows: int = 15, *,
                                      recalcular_fgts: bool = True,
                                      aplicar_jam: bool = False,
                                      data_referencia_jam=None) -> dict:
        """
        Lê as primeiras max_rows linhas do arquivo, valida cada uma sem salvar,
        e retorna um dict para o usuário revisar antes de confirmar o import.

        Returns:
            dict com total_linhas_arquivo, linhas_amostradas, linhas_ok, linhas_erro e rows.
        """
        try:
            wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
            ws = wb.active
        except openpyxl.utils.exceptions.InvalidFileException:
            raise ValueError("❌ Arquivo inválido. Por favor, envie um arquivo XLSX válido.")
        except Exception as e:
            raise ValueError(f"❌ Erro ao ler arquivo: {str(e)}.")

        if not ws.max_row or ws.max_row < 2:
            raise ValueError("❌ Arquivo vazio. O arquivo deve conter pelo menos uma linha de dados além do cabeçalho.")

        total_linhas_arquivo = max((ws.max_row or 2) - 1, 0)

        rows_iter = ws.iter_rows(values_only=True)
        try:
            raw_headers = list(next(rows_iter))
        except StopIteration:
            raise ValueError("❌ Arquivo vazio.")

        headers_upper = [
            LancamentoImportService._canonicalize_header(LancamentoImportService._normalize_header(h))
            for h in raw_headers
        ]

        has_empresa_column = 'EMPRESA' in headers_upper

        if not has_empresa_column and not empresa:
            raise ValueError("❌ Selecione uma empresa para importar ou preencha a coluna EMPRESA no arquivo.")

        missing_columns = [col for col in LancamentoImportService.REQUIRED_COLUMNS if col not in headers_upper]
        if missing_columns:
            raise ValueError(
                f"❌ Colunas obrigatórias faltando: {', '.join(missing_columns)}. "
                f"Baixe o modelo atualizado e preencha corretamente."
            )

        column_indices = {col: headers_upper.index(col) for col in LancamentoImportService.REQUIRED_COLUMNS}
        for col in LancamentoImportService.OPTIONAL_COLUMNS:
            if col in headers_upper:
                column_indices[col] = headers_upper.index(col)

        billing_cache = {}
        empresa_cache = {}
        vinculo_cache = {}
        allowed_empresa_ids = get_allowed_empresa_ids(user)

        jam_coef_cache: dict = {}
        if aplicar_jam:
            try:
                from datetime import date as _date
                from coefjam.models import CoefJam
                for coef in CoefJam.objects.order_by('-data_pagamento'):
                    comp_raw = str(coef.competencia).strip()
                    try:
                        if '/' in comp_raw:
                            m, y = comp_raw.split('/')
                        else:
                            y, m = comp_raw[:7].split('-')
                        jam_coef_cache.setdefault(_date(int(y), int(m), 1), Decimal(str(coef.valor)))
                    except Exception:
                        continue
            except Exception:
                pass

        def _jam_coef_lookup(comp_date):
            return jam_coef_cache.get(comp_date) if jam_coef_cache else None

        nome_idx = column_indices.get('NOME_FUNCIONARIO')
        cpf_idx = column_indices.get('CPF_FUNCIONARIO')
        empresa_idx = column_indices.get('EMPRESA')

        rows_preview = []
        linhas_amostradas = 0
        linhas_ok = 0
        linhas_erro = 0

        for row_idx, row in enumerate(rows_iter, start=2):
            if not any(row):
                continue
            if linhas_amostradas >= max_rows:
                break

            linhas_amostradas += 1

            raw_nome = str(row[nome_idx]).strip() if nome_idx is not None and row[nome_idx] else ''
            raw_cpf_val = row[cpf_idx] if cpf_idx is not None else ''
            raw_cpf = ''.join(filter(str.isdigit, str(raw_cpf_val))) if raw_cpf_val else ''
            raw_cpf_display = f"{raw_cpf[:3]}.{raw_cpf[3:6]}.{raw_cpf[6:9]}-{raw_cpf[9:]}" if len(raw_cpf) == 11 else raw_cpf

            raw_empresa_val = row[empresa_idx] if empresa_idx is not None and empresa_idx < len(row) else None
            raw_empresa_display = str(raw_empresa_val).strip() if raw_empresa_val not in [None, ''] else (empresa.codigo_folha if empresa else '')

            entry = {
                'row_idx': row_idx,
                'raw_nome': raw_nome,
                'raw_cpf': raw_cpf_display,
                'raw_empresa': raw_empresa_display,
            }

            try:
                lancamento_data = LancamentoImportService._process_row(
                    row, column_indices, empresa, user, row_idx,
                    billing_cache, empresa_cache, allowed_empresa_ids, vinculo_cache,
                    recalcular_fgts=recalcular_fgts,
                    aplicar_jam=aplicar_jam,
                    data_referencia_jam=data_referencia_jam,
                    jam_coef_lookup=_jam_coef_lookup,
                )
                vinculo = lancamento_data.get('vinculo')
                entry.update({
                    'status': 'ok',
                    'funcionario_nome': vinculo.funcionario.nome if vinculo else '',
                    'empresa_nome': lancamento_data['empresa'].nome,
                    'competencia': lancamento_data['competencia'],
                    'base_fgts': f"{lancamento_data['base_fgts']:.2f}",
                    'valor_fgts': f"{lancamento_data['valor_fgts']:.2f}",
                    'parcela_13': lancamento_data.get('parcela_13'),
                    'jam_aplicado': lancamento_data.get('_jam_aplicado', False),
                    'valor_fgts_modo': 'arquivo' if not recalcular_fgts else 'calculado',
                })
                linhas_ok += 1
            except Exception as exc:
                entry.update({
                    'status': 'error',
                    'error': str(exc),
                })
                linhas_erro += 1

            rows_preview.append(entry)

        try:
            wb.close()
        except Exception:
            pass

        return {
            'total_linhas_arquivo': total_linhas_arquivo,
            'linhas_amostradas': linhas_amostradas,
            'linhas_ok': linhas_ok,
            'linhas_erro': linhas_erro,
            'rows': rows_preview,
        }

    @staticmethod
    def import_lancamentos_from_file(file, empresa, user, progress_callback=None, *,
                                     recalcular_fgts: bool = True,
                                     aplicar_jam: bool = False,
                                     data_referencia_jam=None,
                                     extrato_analitico: bool = False):
        """
        Importa lançamentos de um arquivo XLSX para uma empresa específica

        Args:
            file: Arquivo XLSX
            empresa: Instância de Empresa
            user: Usuário que está fazendo a importação
            progress_callback: Callable opcional (linhas_processadas: int, linhas_total: int)

        Returns:
            dict: Resultado da importação com estatísticas e erros
        """

        # Nota: empresa pode ser None quando o XLSX traz a coluna EMPRESA por linha.

        # Processar arquivo
        try:
            # read_only=True faz streaming linha a linha — sem carregar o arquivo inteiro
            # na memória. Essencial para arquivos com 10k+ linhas.
            wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
            ws = wb.active
        except openpyxl.utils.exceptions.InvalidFileException:
            raise ValueError("❌ Arquivo inválido. Por favor, envie um arquivo XLSX válido.")
        except Exception as e:
            raise ValueError(f"❌ Erro ao ler arquivo: {str(e)}. Verifique se o arquivo não está corrompido.")

        # Validar se planilha tem dados
        if not ws.max_row or ws.max_row < 2:
            raise ValueError("❌ Arquivo vazio. O arquivo deve conter pelo menos uma linha de dados além do cabeçalho.")

        # max_row vem do atributo <dimension> do XML — pode estar ausente em alguns arquivos.
        total_linhas = max((ws.max_row or 2) - 1, 0)

        # Leitura em única passagem: cabeçalho na 1ª linha, dados continuam do mesmo iterador.
        rows_iter = ws.iter_rows(values_only=True)
        try:
            raw_headers = list(next(rows_iter))
        except StopIteration:
            raise ValueError("❌ Arquivo vazio. O arquivo deve conter pelo menos uma linha de dados além do cabeçalho.")
        
        # Validar headers (usa raw_headers extraído da primeira linha do iterador)
        headers_upper = [
            LancamentoImportService._canonicalize_header(LancamentoImportService._normalize_header(h))
            for h in raw_headers
        ]

        has_empresa_column = 'EMPRESA' in headers_upper

        if not has_empresa_column and not empresa:
            raise ValueError(
                "❌ Selecione uma empresa para importar ou preencha a coluna EMPRESA no arquivo."
            )

        # Se não houver coluna EMPRESA, preserva o comportamento antigo (importação para uma empresa específica)
        if not has_empresa_column and empresa:
            if not is_empresa_allowed(user, empresa.codigo):
                raise ValueError(
                    f"Empresa '{empresa.nome}' não possui permissão para importar lançamentos. "
                    f"Verifique o status do plano."
                )

            try:
                billing_customer = BillingCustomer.objects.get(empresa=empresa)
            except BillingCustomer.DoesNotExist:
                raise ValueError(
                    f"Empresa '{empresa.nome}' não possui billing configurado. "
                    f"Entre em contato com o administrador."
                )

            if billing_customer.status not in ['active', 'trial']:
                raise ValueError(
                    f"Empresa '{empresa.nome}' não possui plano ativo. "
                    f"Status atual: {billing_customer.get_status_display()}"
                )

            if not billing_customer.plan:
                raise ValueError(
                    f"Empresa '{empresa.nome}' não possui plano configurado. "
                    f"Entre em contato com o administrador."
                )
        
        missing_columns = [col for col in LancamentoImportService.REQUIRED_COLUMNS if col not in headers_upper]
        if missing_columns:
            raise ValueError(
                f"❌ Colunas obrigatórias faltando no arquivo: {', '.join(missing_columns)}. "
                f"Por favor, baixe o modelo atualizado e preencha corretamente."
            )
        
        # Criar mapeamento de índices
        column_indices = {col: headers_upper.index(col) for col in LancamentoImportService.REQUIRED_COLUMNS}
        for col in LancamentoImportService.OPTIONAL_COLUMNS:
            if col in headers_upper:
                column_indices[col] = headers_upper.index(col)
        
        # Processar linhas
        result = {
            'success': 0,
            'errors': [],
            'warnings': [],
            'created': 0,
            'updated': 0,
            'skipped': 0,
            'linhas_valor_do_arquivo': 0,
            'linhas_jam_aplicado': 0,
            'rows': [],
        }
        
        linhas_processadas = 0

        # Caches que persistem entre linhas para evitar queries repetidas
        billing_cache = {}   # str(empresa.codigo) → BillingCustomer | str(erro)
        empresa_cache = {}   # raw_value → Empresa | None
        vinculo_cache = {}   # "c:{cpf}:{empresa_pk}" → list[FuncionarioVinculo]
                             # "m:{matricula}:{empresa_pk}" → FuncionarioVinculo | None
        # get_allowed_empresa_ids faz ~5 queries (user.empresa, grupo, roles…);
        # computar uma única vez antes do loop e reutilizar em toda a importação.
        allowed_empresa_ids = get_allowed_empresa_ids(user)  # None = superuser irrestrito

        # Cache de coeficientes JAM: pré-carrega TODOS de uma vez quando aplicar_jam=True.
        # Sem o cache, calcular_jam_ate_pagamento faz 1 query/mês/linha
        # (ex: 74 meses × 1.000 linhas = 74.000 queries desnecessárias).
        jam_coef_cache: dict = {}
        if aplicar_jam:
            try:
                from datetime import date as _date
                from coefjam.models import CoefJam
                # order_by desc → setdefault mantém o registro mais recente por competência
                for coef in CoefJam.objects.order_by('-data_pagamento'):
                    comp_raw = str(coef.competencia).strip()
                    try:
                        if '/' in comp_raw:
                            m, y = comp_raw.split('/')
                        else:
                            y, m = comp_raw[:7].split('-')
                        jam_coef_cache.setdefault(_date(int(y), int(m), 1), Decimal(str(coef.valor)))
                    except Exception:
                        continue
            except Exception:
                pass  # se falhar, o JAM usa a busca normal (1 query/mês)

        def _jam_coef_lookup(comp_date):
            return jam_coef_cache.get(comp_date) if jam_coef_cache else None

        if progress_callback:
            progress_callback(0, total_linhas)

        for row_idx, row in enumerate(rows_iter, start=2):
            # Pular linhas vazias
            if not any(row):
                continue

            linhas_processadas += 1

            if progress_callback and linhas_processadas % 200 == 0:
                progress_callback(linhas_processadas, total_linhas)

            # Captura prévia dos campos-chave para o relatório analítico
            def _raw(col):
                idx = column_indices.get(col)
                return str(row[idx]).strip() if idx is not None and row[idx] not in (None, '') else ''

            raw_cpf = _raw('CPF_FUNCIONARIO')
            raw_nome = _raw('NOME_FUNCIONARIO')
            raw_comp = _raw('COMPETENCIA')
            raw_base = _raw('BASE_FGTS')
            raw_emp = _raw('EMPRESA') or (empresa.nome if empresa else '')

            try:
                lancamento_data = LancamentoImportService._process_row(
                    row, column_indices, empresa, user, row_idx,
                    billing_cache, empresa_cache, allowed_empresa_ids, vinculo_cache,
                    recalcular_fgts=recalcular_fgts,
                    aplicar_jam=aplicar_jam,
                    data_referencia_jam=data_referencia_jam,
                    jam_coef_lookup=_jam_coef_lookup,
                    extrato_analitico=extrato_analitico,
                )

                if lancamento_data:
                    # Extrair flags internas antes de criar o objeto
                    _jam_aplicado = lancamento_data.pop('_jam_aplicado', False)
                    desired_valor_fgts = lancamento_data['valor_fgts']
                    _vinculo = lancamento_data.get('vinculo')
                    _aliquota = get_aliquota_fgts(_vinculo)
                    computed_expected = (lancamento_data['base_fgts'] * _aliquota).quantize(Decimal('0.01'))
                    # needs_restore: o valor foi validado em _process_row, mas Lancamento.save()
                    # pode recalcular; restauramos se o valor final difere do esperado
                    needs_restore = (not recalcular_fgts) and (desired_valor_fgts != computed_expected)

                    # Verificar se já existe lançamento para esta competência/parcela
                    existing_qs = Lancamento.objects.filter(
                        empresa=lancamento_data['empresa'],
                        competencia=lancamento_data['competencia'],
                        parcela_13=lancamento_data.get('parcela_13')
                    )

                    if lancamento_data.get('vinculo') is not None:
                        existing_qs = existing_qs.filter(vinculo=lancamento_data['vinculo'])
                    else:
                        existing_qs = existing_qs.filter(funcionario=lancamento_data['funcionario'], vinculo__isnull=True)

                    existing_list = list(existing_qs[:2])
                    if len(existing_list) > 1:
                        parcela_label = f" (13º {lancamento_data.get('parcela_13')}ª parcela)" if lancamento_data.get('parcela_13') else ""
                        raise ValueError(
                            f"Duplicidade detectada: já existem múltiplos lançamentos para esta competência{parcela_label}."
                        )

                    existing = existing_list[0] if existing_list else None

                    if existing:
                        # Atualizar
                        for key, value in lancamento_data.items():
                            if key not in ['funcionario', 'vinculo']:
                                setattr(existing, key, value)
                        existing.save()
                        if needs_restore:
                            Lancamento.objects.filter(pk=existing.pk).update(
                                valor_fgts=desired_valor_fgts,
                                atualizado_em=timezone.now(),
                            )
                        result['updated'] += 1
                        acao = 'atualizado'
                    else:
                        # Criar novo
                        novo = Lancamento(**lancamento_data)
                        novo.save()
                        if needs_restore:
                            Lancamento.objects.filter(pk=novo.pk).update(
                                valor_fgts=desired_valor_fgts,
                                atualizado_em=timezone.now(),
                            )
                        result['created'] += 1
                        acao = 'criado'

                    result['success'] += 1
                    if needs_restore:
                        result['linhas_valor_do_arquivo'] += 1
                    if _jam_aplicado:
                        result['linhas_jam_aplicado'] += 1

                    detalhe = acao.capitalize()
                    if needs_restore:
                        detalhe += ' (valor mantido do arquivo)'
                    if _jam_aplicado:
                        detalhe += ' (JAM aplicado)'
                    result['rows'].append({
                        'linha': row_idx, 'cpf': raw_cpf, 'nome': raw_nome,
                        'competencia': raw_comp, 'base_fgts': raw_base, 'empresa': raw_emp,
                        'acao': acao, 'status': 'ok', 'detalhe': detalhe,
                    })
                else:
                    result['skipped'] += 1
                    result['rows'].append({
                        'linha': row_idx, 'cpf': raw_cpf, 'nome': raw_nome,
                        'competencia': raw_comp, 'base_fgts': raw_base, 'empresa': raw_emp,
                        'acao': 'ignorado', 'status': 'ignorado', 'detalhe': 'Linha ignorada',
                    })

            except Exception as e:
                result['errors'].append({
                    'row': row_idx,
                    'error': str(e)
                })
                result['rows'].append({
                    'linha': row_idx, 'cpf': raw_cpf, 'nome': raw_nome,
                    'competencia': raw_comp, 'base_fgts': raw_base, 'empresa': raw_emp,
                    'acao': 'erro', 'status': 'erro', 'detalhe': str(e),
                })
        
        # Validar se alguma linha foi processada
        if linhas_processadas == 0:
            raise ValueError(
                "❌ Nenhuma linha de dados encontrada no arquivo. "
                "Verifique se você preencheu o arquivo corretamente e removeu a linha de exemplo."
            )

        try:
            wb.close()
        except Exception:
            # openpyxl pode lançar erro ao parsear metadados do AutoFilter ou
            # outros elementos não relacionados aos dados (ex: CustomFilter inválido).
            # O dado já foi processado; ignore erros de fechamento.
            pass
        return result
    
    @staticmethod
    def _process_row(row, column_indices, empresa, user, row_idx, billing_cache, empresa_cache, allowed_empresa_ids, vinculo_cache, *,
                     recalcular_fgts: bool = True,
                     aplicar_jam: bool = False,
                     data_referencia_jam=None,
                     jam_coef_lookup=None,
                     extrato_analitico: bool = False):
        """Processa uma linha do arquivo e retorna dados do lançamento"""

        # Extrair MATRÍCULA (opcional)
        matricula_idx = column_indices.get('MATRICULA')
        raw_matricula = row[matricula_idx] if matricula_idx is not None else None
        matricula = ''
        if raw_matricula not in [None, '']:
            if isinstance(raw_matricula, (int, float)):
                matricula = str(int(raw_matricula))
            else:
                matricula = str(raw_matricula).strip()
                if matricula.endswith('.0') and matricula.replace('.', '', 1).isdigit():
                    matricula = str(int(float(matricula)))
            matricula = ''.join(filter(str.isdigit, matricula))
        
        # Extrair CPF
        cpf_idx = column_indices.get('CPF_FUNCIONARIO')
        cpf = str(row[cpf_idx]).strip() if row[cpf_idx] else ''
        cpf = ''.join(filter(str.isdigit, cpf))  # Remover formatação

        # CPF pode ficar em branco quando MATRÍCULA estiver preenchida
        if not cpf and not matricula:
            raise ValueError("Informe CPF do funcionário ou a MATRÍCULA do vínculo")

        if cpf and len(cpf) != 11:
            raise ValueError(f"CPF inválido: {cpf}. O CPF deve conter 11 dígitos")
        
        # Resolver empresa (por linha) - se existir coluna EMPRESA, ela sobrescreve a seleção do formulário
        empresa_row = empresa
        empresa_idx = column_indices.get('EMPRESA')
        if empresa_idx is not None and row[empresa_idx] not in [None, '']:
            raw = row[empresa_idx]
            raw_key = str(raw).strip()
            if raw_key in empresa_cache:
                empresa_row = empresa_cache[raw_key]
            else:
                empresa_row = LancamentoImportService._resolve_empresa_from_identifier(raw)
                empresa_cache[raw_key] = empresa_row
            if not empresa_row:
                raise ValueError(f"Empresa '{raw}' não encontrada")

        if not empresa_row:
            raise ValueError('Empresa não informada. Selecione uma empresa ou preencha a coluna EMPRESA.')

        if allowed_empresa_ids is not None and empresa_row.codigo not in allowed_empresa_ids:
            raise ValueError('Você não tem permissão para importar lançamentos para a empresa informada.')

        billing_customer = LancamentoImportService._validate_billing_for_empresa(empresa_row, billing_cache)

        # Extrair competência
        competencia_idx = column_indices.get('COMPETENCIA')
        competencia = str(row[competencia_idx]).strip() if row[competencia_idx] else ''
        
        if not competencia:
            raise ValueError(f"Competência não informada. Use o formato MM/YYYY (ex: 01/2026)")
        
        # Validar formato MM/YYYY
        try:
            if '/' not in competencia:
                raise ValueError("Formato inválido")
            mes, ano = competencia.split('/')
            mes = int(mes)
            ano = int(ano)
            if mes < 1 or mes > 12:
                raise ValueError("Mês deve estar entre 01 e 12")
            if ano < 1900 or ano > 2100:
                raise ValueError("Ano inválido")
            competencia = f"{mes:02d}/{ano}"
        except ValueError as ve:
            raise ValueError(f"Competência inválida: '{competencia}'. Use o formato MM/YYYY (ex: 01/2026). {str(ve)}")
        except Exception:
            raise ValueError(f"Competência inválida: '{competencia}'. Use o formato MM/YYYY (ex: 01/2026)")

        # Validar limite de histórico (usa billing_customer já obtido do cache acima)
        try:
            max_history_months = billing_customer.get_effective_max_history_months()
            if max_history_months is not None and max_history_months > 0:
                competencia_date = datetime(ano, mes, 1).date()
                today = datetime.today().date()
                current_month = datetime(today.year, today.month, 1).date()
                min_date = current_month - relativedelta(months=max_history_months - 1)
                if competencia_date < min_date:
                    raise ValueError(
                        f"Competência {competencia} fora do limite do plano: "
                        f"máximo de {max_history_months} meses de histórico."
                    )
        except ValueError:
            raise
        except Exception:
            pass

        # Resolver vínculo (cadeira) por empresa + competência
        vinculo_idx = column_indices.get('VINCULO')
        raw_vinculo = row[vinculo_idx] if vinculo_idx is not None else None

        vinculo = LancamentoImportService._resolve_vinculo_for_empresa_competencia(
            cpf=cpf,
            empresa=empresa_row,
            competencia=competencia,
            raw_vinculo=raw_vinculo,
            raw_matricula=matricula,
            vinculo_cache=vinculo_cache,
        )

        funcionario = vinculo.funcionario
        
        # Extrair base FGTS
        base_fgts_idx = column_indices.get('BASE_FGTS')
        base_fgts_value = row[base_fgts_idx]
        
        try:
            if isinstance(base_fgts_value, (int, float)):
                base_fgts = Decimal(str(base_fgts_value))
            else:
                base_fgts_str = str(base_fgts_value).strip().replace(',', '.')
                base_fgts = Decimal(base_fgts_str)
        except (InvalidOperation, ValueError, AttributeError):
            raise ValueError(f"Base FGTS inválida: '{base_fgts_value}'. Use valores numéricos com ponto decimal (ex: 3500.00)")
        
        if base_fgts < 0:
            raise ValueError(f"Base FGTS não pode ser negativa: {base_fgts}")
        
        # Calcular ou extrair valor FGTS conforme opção do usuário
        valor_fgts_idx = column_indices.get('VALOR_FGTS')
        valor_fgts_arquivo = None

        if valor_fgts_idx is not None and row[valor_fgts_idx]:
            try:
                valor_fgts_value = row[valor_fgts_idx]
                if isinstance(valor_fgts_value, (int, float)):
                    valor_fgts_arquivo = Decimal(str(valor_fgts_value))
                else:
                    valor_fgts_str = str(valor_fgts_value).strip().replace(',', '.')
                    valor_fgts_arquivo = Decimal(valor_fgts_str)
            except (InvalidOperation, ValueError):
                pass

        if recalcular_fgts or valor_fgts_arquivo is None:
            # Calcular usando o percentual do tipo de vínculo
            aliquota = get_aliquota_fgts(vinculo)
            valor_fgts = (base_fgts * aliquota).quantize(Decimal('0.01'))
        else:
            # Arquivo tem valor explícito → validar contra % do vínculo
            aliquota = get_aliquota_fgts(vinculo)
            esperado = (base_fgts * aliquota).quantize(Decimal('0.01'))
            if valor_fgts_arquivo != esperado:
                tipo_nome = vinculo.tipo_vinculo.descricao if vinculo.tipo_vinculo_id else 'CLT'
                pct = aliquota * 100
                raise ValueError(
                    f"Valor FGTS no arquivo (R$ {valor_fgts_arquivo:.2f}) não corresponde ao "
                    f"tipo de vínculo '{tipo_nome}' ({pct:.0f}%). "
                    f"Valor esperado: R$ {esperado:.2f}. "
                    f"Corrija o arquivo ou use a opção 'Recalcular FGTS'."
                )
            valor_fgts = valor_fgts_arquivo
        
        # Aplicar correção JAM se o usuário solicitou
        _jam_aplicado = False
        if aplicar_jam:
            try:
                from datetime import date as _date
                from lancamentos.services.calculo import calcular_jam_ate_pagamento
                mes_str, ano_str = competencia.split('/')
                comp_date = _date(int(ano_str), int(mes_str), 1)
                ref_date = data_referencia_jam or _date.today()
                if ref_date > comp_date:
                    jam_total, _, _ = calcular_jam_ate_pagamento(valor_fgts, comp_date, ref_date, coef_lookup=jam_coef_lookup)
                    if jam_total > Decimal('0.00'):
                        valor_fgts = (valor_fgts + jam_total).quantize(Decimal('0.01'))
                        _jam_aplicado = True
            except Exception:
                pass  # JAM não bloqueia a linha — importa sem a correção

        # Dados do lançamento
        lancamento_data = {
            'empresa': empresa_row,
            'funcionario': funcionario,
            'vinculo': vinculo,
            'competencia': competencia,
            'base_fgts': base_fgts,
            'valor_fgts': valor_fgts,
            'pago': False,
            'data_pagto': None,
            'valor_pago': None,
            'parcela_13': None,
            '_jam_aplicado': _jam_aplicado,  # flag interna, removida antes de salvar
        }
        
        # Processar campo PAGO (opcional)
        pago_idx = column_indices.get('PAGO')
        if pago_idx is not None and row[pago_idx]:
            pago_value = str(row[pago_idx]).strip().upper()
            lancamento_data['pago'] = pago_value in ['SIM', 'S', 'TRUE', '1', 'YES']

        # Processar DATA_PAGTO (opcional)
        data_pagto_idx = column_indices.get('DATA_PAGTO')
        if data_pagto_idx is not None and row[data_pagto_idx]:
            try:
                data_value = row[data_pagto_idx]
                if isinstance(data_value, datetime):
                    lancamento_data['data_pagto'] = data_value.date()
                else:
                    # Tentar parsear dd/mm/yyyy
                    data_str = str(data_value).strip()
                    lancamento_data['data_pagto'] = datetime.strptime(data_str, '%d/%m/%Y').date()
            except Exception:
                pass  # Ignorar data inválida

        # Processar VALOR_PAGO (opcional)
        valor_pago_idx = column_indices.get('VALOR_PAGO')
        if valor_pago_idx is not None and row[valor_pago_idx]:
            try:
                valor_pago_value = row[valor_pago_idx]
                if isinstance(valor_pago_value, (int, float)):
                    lancamento_data['valor_pago'] = Decimal(str(valor_pago_value))
                else:
                    valor_pago_str = str(valor_pago_value).strip().replace(',', '.')
                    lancamento_data['valor_pago'] = Decimal(valor_pago_str)
            except (InvalidOperation, ValueError):
                pass  # Ignorar valor inválido

        # Validar consistência entre PAGO e DATA_PAGTO/VALOR_PAGO
        if not lancamento_data['pago'] and (lancamento_data['data_pagto'] or lancamento_data['valor_pago']):
            raise ValueError(
                "❌ Divergência: PAGO está 'NÃO'/vazio, mas DATA_PAGTO ou VALOR_PAGO estão "
                "preenchidos. Marque PAGO=SIM ou remova a data/valor de pagamento da planilha."
            )
        if lancamento_data['pago'] and data_pagto_idx is not None and not lancamento_data['data_pagto']:
            raise ValueError(
                "❌ Divergência: PAGO está 'SIM', mas a coluna DATA_PAGTO está em branco nesta linha. "
                "Preencha a data de pagamento ou marque PAGO=NÃO."
            )

        # Definir fonte_confirmacao_pagamento para lançamentos pagos
        if lancamento_data['pago']:
            lancamento_data['fonte_confirmacao_pagamento'] = 'extrato_analitico' if extrato_analitico else 'manual'

        # Se marcado como pago e o arquivo não possui a coluna DATA_PAGTO, usar data atual
        if lancamento_data['pago'] and not lancamento_data['data_pagto']:
            lancamento_data['data_pagto'] = datetime.today().date()

        # Processar PARCELA_13 (opcional)
        # Aceita valores: 1, 2, "1", "2", "SIM" (= 1), "PRIMEIRA" (= 1), "SEGUNDA" (= 2), etc.
        parcela_13_idx = column_indices.get('PARCELA_13')
        if parcela_13_idx is not None and row[parcela_13_idx]:
            try:
                parcela_value = str(row[parcela_13_idx]).strip().upper()
                if parcela_value in ['1', 'PRIMEIRA', 'ADIANTAMENTO', 'SIM']:
                    lancamento_data['parcela_13'] = 1
                elif parcela_value in ['2', 'SEGUNDA', 'DEZEMBRO']:
                    lancamento_data['parcela_13'] = 2
                # Caso contrário deixa None (competência normal)
            except Exception:
                pass  # Ignorar valor inválido
        
        return lancamento_data

    @staticmethod
    def _validate_billing_for_empresa(empresa: Empresa, billing_cache: dict) -> 'BillingCustomer':
        """Valida billing para uma empresa e retorna o BillingCustomer, com cache por código."""
        key = str(getattr(empresa, 'codigo', empresa.pk))
        cached = billing_cache.get(key)
        if cached is not None:
            if isinstance(cached, str):
                raise ValueError(cached)
            return cached  # BillingCustomer já validado

        try:
            billing_customer = BillingCustomer.objects.get(empresa=empresa)
        except BillingCustomer.DoesNotExist:
            msg = (
                f"Empresa '{empresa.nome}' não possui billing configurado. "
                f"Entre em contato com o administrador."
            )
            billing_cache[key] = msg
            raise ValueError(msg)

        if billing_customer.status not in ['active', 'trial']:
            msg = (
                f"Empresa '{empresa.nome}' não possui plano ativo. "
                f"Status atual: {billing_customer.get_status_display()}"
            )
            billing_cache[key] = msg
            raise ValueError(msg)

        if not billing_customer.plan:
            msg = (
                f"Empresa '{empresa.nome}' não possui plano configurado. "
                f"Entre em contato com o administrador."
            )
            billing_cache[key] = msg
            raise ValueError(msg)

        billing_cache[key] = billing_customer
        return billing_customer

    @staticmethod
    def _resolve_vinculo_for_empresa_competencia(*, cpf: str, empresa: Empresa, competencia: str, raw_vinculo, raw_matricula, vinculo_cache: dict = None) -> FuncionarioVinculo:
        """Resolve qual vínculo (cadeira) usar, considerando empresa e competência.

        - Se a coluna VINCULO (ID) for informada, ela prevalece.
        - Caso contrário, exige exatamente 1 vínculo ativo na competência.
          Se houver mais de um, pede VINCULO para desambiguar.

        vinculo_cache: dict mutable compartilhado entre rows para evitar queries repetidas
        para o mesmo (cpf/matricula + empresa). Chaves:
          "m:{matricula}:{empresa_pk}" → FuncionarioVinculo | None
          "c:{cpf}:{empresa_pk}"       → list[FuncionarioVinculo]
        """
        if vinculo_cache is None:
            vinculo_cache = {}

        cpf_fmt = f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}" if cpf and len(cpf) == 11 else (cpf or '—')

        if raw_vinculo not in [None, '']:
            try:
                vinculo_id = int(raw_vinculo) if not isinstance(raw_vinculo, str) else int(raw_vinculo.strip())
            except Exception:
                raise ValueError(f"VINCULO inválido: '{raw_vinculo}'. Informe o ID numérico do vínculo.")

            try:
                vinculo = FuncionarioVinculo.objects.select_related('funcionario', 'empresa').get(pk=vinculo_id)
            except FuncionarioVinculo.DoesNotExist:
                raise ValueError(f"Vínculo (VINCULO={vinculo_id}) não encontrado.")

            if getattr(vinculo.funcionario, 'cpf', None) != cpf:
                raise ValueError(
                    f"Vínculo (VINCULO={vinculo_id}) não pertence ao CPF {cpf_fmt}. "
                    f"Verifique a coluna VINCULO."
                )

            if vinculo.empresa_id != empresa.id:
                raise ValueError(
                    f"Vínculo (VINCULO={vinculo_id}) pertence à empresa '{vinculo.empresa.nome}', "
                    f"mas a linha está para a empresa '{empresa.nome}'."
                )

            if not vinculo.is_ativo_em_competencia(competencia):
                raise ValueError(
                    f"Vínculo (VINCULO={vinculo_id}) não está ativo na competência {competencia}."
                )

            return vinculo

        # Resolver por MATRÍCULA do vínculo (recomendado)
        if raw_matricula not in [None, '']:
            matricula = str(raw_matricula).strip()
            matricula = ''.join(filter(str.isdigit, matricula))
            if not matricula:
                raise ValueError(f"MATRICULA inválida: '{raw_matricula}'. Informe apenas números.")

            cache_key_m = f"m:{matricula}:{empresa.pk}"
            if cache_key_m in vinculo_cache:
                vinculo = vinculo_cache[cache_key_m]
            else:
                vinculo = FuncionarioVinculo.objects.select_related('funcionario', 'empresa').filter(
                    empresa=empresa,
                    matricula=matricula,
                ).order_by('-data_admissao', '-id').first()
                vinculo_cache[cache_key_m] = vinculo

            if not vinculo:
                raise ValueError(
                    f"Vínculo não encontrado para a MATRÍCULA {matricula} na empresa '{empresa.nome}'."
                )

            if cpf and getattr(vinculo.funcionario, 'cpf', None) != cpf:
                raise ValueError(
                    f"MATRICULA {matricula} pertence a outro CPF ({vinculo.funcionario.cpf}). "
                    f"Verifique CPF/MATRICULA."
                )

            if not vinculo.is_ativo_em_competencia(competencia):
                raise ValueError(
                    f"Vínculo da MATRÍCULA {matricula} não está ativo na competência {competencia}."
                )

            return vinculo

        # Resolver por CPF — busca todos os vínculos uma única vez e cacheia
        cache_key_c = f"c:{cpf}:{empresa.pk}"
        if cache_key_c in vinculo_cache:
            vinculos_list = vinculo_cache[cache_key_c]
        else:
            vinculos_list = list(
                FuncionarioVinculo.objects.filter(
                    empresa=empresa,
                    funcionario__cpf=cpf,
                ).select_related('funcionario').order_by('-data_admissao', '-id')
            )
            vinculo_cache[cache_key_c] = vinculos_list

        if not vinculos_list:
            raise ValueError(
                f"Colaborador com CPF {cpf_fmt} não encontrado na empresa '{empresa.nome}'. "
                f"Certifique-se de que o colaborador está cadastrado e vinculado a esta empresa antes de importar seus lançamentos."
            )

        ativos = []
        for v in vinculos_list:
            try:
                if v.is_ativo_em_competencia(competencia):
                    ativos.append(v)
            except Exception:
                continue

        if len(ativos) == 1:
            return ativos[0]

        if len(ativos) == 0:
            raise ValueError(
                f"Nenhum vínculo ativo encontrado para o CPF {cpf_fmt} na empresa '{empresa.nome}' "
                f"na competência {competencia}. Verifique a competência, a empresa (coluna EMPRESA) e as datas de admissão/demissão do vínculo."
            )

        raise ValueError(
            f"Vínculo ambíguo: existem múltiplos vínculos ativos para o CPF {cpf_fmt} na empresa '{empresa.nome}' "
            f"na competência {competencia}. Preencha a coluna MATRICULA (recomendado) ou VINCULO (ID interno) para escolher a cadeira correta."
        )

    @staticmethod
    def _resolve_funcionario_for_empresa_competencia(*, cpf: str, empresa: Empresa, competencia: str) -> Funcionario:
        """Compat: mantém API antiga retornando Funcionario.

        A partir do suporte a múltiplos vínculos na mesma competência, a resolução correta é por vínculo.
        """

        vinculo = LancamentoImportService._resolve_vinculo_for_empresa_competencia(
            cpf=cpf,
            empresa=empresa,
            competencia=competencia,
            raw_vinculo=None,
            raw_matricula=None,
        )
        return vinculo.funcionario
