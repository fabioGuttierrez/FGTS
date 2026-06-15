from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.utils.timezone import make_aware
from django.core.exceptions import ValidationError
from django.db import transaction
from fgtsweb.utils.validators import digits_only, normalize_upper_ascii, validate_cpf
from .models import Funcionario
from empresas.models import Empresa
from empresas.models_grupo import FuncionarioVinculo
from billing.models import BillingCustomer
from fgtsweb.mixins import is_empresa_allowed


class FuncionarioImportService:
    """Serviço para gerenciar importação e exportação de funcionários em XLSX"""
    
    REQUIRED_COLUMNS = [
        'NOME', 'CPF', 'DATA_ADMISSAO', 'EMPRESA'
    ]
    
    OPTIONAL_COLUMNS = [
        'MATRICULA', 'PIS', 'CBO', 'CARTEIRA_PROFISSIONAL',
        'SERIE_CARTEIRA', 'DATA_NASCIMENTO', 'DATA_DEMISSAO', 'OBSERVACAO', 'SALARIO', 'CARGO'
    ]

    @staticmethod
    def _resolve_empresa_from_identifier(value):
        if value is None:
            return None

        raw = str(value).strip()
        if not raw:
            return None

        if raw.endswith('.0') and raw.replace('.', '', 1).isdigit():
            raw = str(int(float(raw)))

        # Tenta buscar por codigo_folha
        qs = Empresa.objects.filter(codigo_folha__iexact=raw)
        if qs.count() > 1:
            raise ValueError(f"Codigo Folha '{raw}' duplicado. Contate o administrador.")
        if qs.exists():
            return qs.first()

        # Tenta buscar por ID/codigo
        if raw.isdigit():
            try:
                return Empresa.objects.get(pk=int(raw))
            except Empresa.DoesNotExist:
                try:
                    return Empresa.objects.get(codigo=int(raw))
                except Empresa.DoesNotExist:
                    return None

        return None
    
    @staticmethod
    def generate_template_xlsx():
        """Gera um arquivo XLSX com o modelo para importação de funcionários"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Funcionários"
        
        # Definir estilos
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        example_fill = PatternFill(start_color="E7E9FF", end_color="E7E9FF", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers com colunas obrigatórias e opcionais
        all_columns = FuncionarioImportService.REQUIRED_COLUMNS + FuncionarioImportService.OPTIONAL_COLUMNS
        
        for col_idx, column_name in enumerate(all_columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = column_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Adicionar linha de exemplo
        example_data = [
            "João da Silva",  # NOME
            "123.456.789-00",  # CPF
            "2023-01-15",  # DATA_ADMISSAO
            "EMP001",  # EMPRESA (codigo folha)
            "1001",  # MATRICULA (do vínculo)
            "120.123.456-70",  # PIS
            "2110",  # CBO
            "AB123456",  # CARTEIRA_PROFISSIONAL
            "12",  # SERIE_CARTEIRA
            "1990-05-10",  # DATA_NASCIMENTO
            "",  # DATA_DEMISSAO
            "Informações adicionais",  # OBSERVACAO
            "3500.00",  # SALARIO
        ]
        
        for col_idx, value in enumerate(example_data, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = value
            cell.fill = example_fill
            cell.border = border
            header_name = all_columns[col_idx - 1]
            if header_name in ['DATA_ADMISSAO', 'DATA_NASCIMENTO', 'DATA_DEMISSAO']:
                cell.number_format = 'YYYY-MM-DD'
        
        # Adicionar informações
        info_row = 4
        ws.merge_cells(f'A{info_row}:D{info_row}')
        info_cell = ws[f'A{info_row}']
        info_cell.value = "⚠️ INSTRUÇÕES DE PREENCHIMENTO"
        info_cell.font = Font(bold=True, size=10, color="764ba2")
        
        instructions = [
            "• Campos obrigatórios: NOME, CPF, DATA_ADMISSAO, EMPRESA",
            "• Formato de datas: YYYY-MM-DD (ex: 2023-01-15)",
            "• CPF deve estar no formato XXX.XXX.XXX-XX",
            "• EMPRESA: use o CODIGO FOLHA da empresa (pode conter letras)",
            "• MATRICULA: opcional, mas recomendada para evitar vínculo ambíguo em importações futuras",
            "• Para ver o codigo folha, acesse a lista de empresas no sistema",
            "• PIS é opcional e pode conter CPF (conforme regra atual)",
            "• DATA_DEMISSAO deixar em branco se o funcionário está ativo",
            "• SALARIO: se preenchido, cria automaticamente o primeiro lançamento de FGTS",
            "• SALARIO formato: número com ponto (ex: 3500.00)",
            "• Não altere os nomes das colunas ou a ordem delas",
            "• Apague a linha de exemplo antes de importar seus dados",
        ]
        
        for idx, instruction in enumerate(instructions, 1):
            instr_row = info_row + idx
            ws.merge_cells(f'A{instr_row}:D{instr_row}')
            instr_cell = ws[f'A{instr_row}']
            instr_cell.value = instruction
            instr_cell.font = Font(size=9)
            instr_cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 30
        
        # Definir altura da linha de header
        ws.row_dimensions[1].height = 30
        
        return wb
    
    @staticmethod
    def parse_date(date_value):
        """Converte valor de data para objeto datetime"""
        if not date_value or str(date_value).strip() == "":
            return None
        
        if isinstance(date_value, datetime):
            return date_value.date()
        
        # Tentar diferentes formatos
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
            try:
                return datetime.strptime(str(date_value).strip(), fmt).date()
            except ValueError:
                continue
        
        raise ValueError(f"Formato de data inválido: {date_value}. Use YYYY-MM-DD ou DD/MM/YYYY")
    
    @staticmethod
    def import_funcionarios_from_file(file, empresa_id=None, user=None):
        """
        Importa funcionários de um arquivo XLSX
        
        Args:
            file: arquivo XLSX
            empresa_id: ID da empresa (opcional, pode estar no arquivo)
            user: usuário que está realizando a importação
            
        Returns:
            dict com estatísticas da importação
        """
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            # Obter headers
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[cell.value.upper()] = col_idx
            
            # Validar colunas obrigatórias
            missing_columns = [col for col in FuncionarioImportService.REQUIRED_COLUMNS if col not in headers]
            if missing_columns:
                raise ValueError(f"Colunas obrigatórias faltando: {', '.join(missing_columns)}")
            
            result = {
                'total': 0,
                'success': 0,
                'errors': [],
                'created_funcionarios': []
            }
            
            # Processar linhas
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                try:
                    row_data = {}
                    
                    # Extrair dados
                    for header, col_idx in headers.items():
                        cell = row[col_idx - 1]
                        row_data[header] = cell.value
                    
                    # Validar dados obrigatórios
                    if not str(row_data.get('NOME', '')).strip():
                        raise ValueError("Nome é obrigatório")
                    
                    if not str(row_data.get('CPF', '')).strip():
                        raise ValueError("CPF é obrigatório")
                    
                    if not row_data.get('DATA_ADMISSAO'):
                        raise ValueError("Data de admissão é obrigatória")
                    
                    # Validar e obter empresa
                    empresa_identifier = row_data.get('EMPRESA') or empresa_id
                    if not empresa_identifier:
                        raise ValueError("Empresa é obrigatória")
                    
                    empresa = FuncionarioImportService._resolve_empresa_from_identifier(empresa_identifier)
                    if not isinstance(empresa, Empresa):
                        raise ValueError(f"Empresa '{empresa_identifier}' não encontrada ou inválida")
                    
                    # VALIDAÇÃO 1: Verificar se usuário tem permissão para essa empresa
                    if user and not is_empresa_allowed(user, empresa.codigo):
                        raise ValueError(
                            "A empresa informada no arquivo não faz parte do seu grupo de empresas. "
                            "Verifique o código da empresa e tente novamente."
                        )
                    
                    # VALIDAÇÃO 2: Verificar se empresa tem billing ativo ou em trial
                    try:
                        billing_customer = empresa.billing_customer
                        if billing_customer.status not in ['active', 'trial']:
                            raise ValueError(
                                f"Empresa '{empresa.nome}' não possui assinatura ativa. "
                                f"Status atual: {billing_customer.get_status_display()}. "
                                f"Entre em contato com o administrador para regularizar."
                            )
                        
                        # ✅ EMPRESAS EM TRIAL TÊM LIMITE ILIMITADO!
                        if billing_customer.status == 'trial':
                            pass  # Pula validação de limite para empresas trial
                        # VALIDAÇÃO 3: Verificar limite de funcionários do plano (apenas para empresas active)
                        elif billing_customer.plan or billing_customer.override_max_employees is not None:
                            # Contar funcionários ativos da empresa
                            active_count = empresa.funcionariovinculo_set.filter(data_demissao__isnull=True).count()
                            
                            # Verificar se pode adicionar mais um
                            if not billing_customer.can_add_employee(active_count):
                                plan_name = billing_customer.plan.get_plan_type_display() if billing_customer.plan else 'Especial'
                                max_employees = billing_customer.get_effective_max_employees()
                                raise ValueError(
                                    f"Plano '{plan_name}' da empresa '{empresa.nome}' permite no máximo "
                                    f"{max_employees} colaboradores ativos. "
                                    f"Já existem {active_count} cadastrados. "
                                    f"Faça upgrade do plano para adicionar mais."
                                )
                        else:
                            raise ValueError(
                                f"Empresa '{empresa.nome}' não possui plano configurado. "
                                f"Entre em contato com o administrador."
                            )
                    except Empresa.billing_customer.RelatedObjectDoesNotExist:
                        raise ValueError(
                            f"Empresa '{empresa.nome}' não possui configuração de billing. "
                            f"Entre em contato com o administrador."
                        )
                    
                    # Preparar dados do funcionário
                    funcionario_data = {
                        'nome': normalize_upper_ascii(row_data['NOME'], allow_digits=False),
                        'cpf': validate_cpf(row_data['CPF']),
                    }

                    # Campos opcionais com validação/normalização
                    if row_data.get('PIS'):
                        # Regra atual permite CPF no lugar do PIS; não validamos DV.
                        pis = digits_only(row_data['PIS'])
                        if len(pis) > 15:
                            raise ValueError('PIS muito longo. Informe no máximo 15 dígitos.')
                        funcionario_data['pis'] = pis

                    if row_data.get('CBO'):
                        funcionario_data['cbo'] = normalize_upper_ascii(row_data['CBO'], allow_digits=True)

                    if row_data.get('CARTEIRA_PROFISSIONAL'):
                        funcionario_data['carteira_profissional'] = normalize_upper_ascii(row_data['CARTEIRA_PROFISSIONAL'], allow_digits=True)

                    if row_data.get('SERIE_CARTEIRA'):
                        funcionario_data['serie_carteira'] = normalize_upper_ascii(row_data['SERIE_CARTEIRA'], allow_digits=True)

                    if row_data.get('DATA_NASCIMENTO'):
                        funcionario_data['data_nascimento'] = FuncionarioImportService.parse_date(row_data['DATA_NASCIMENTO'])

                    if row_data.get('OBSERVACAO'):
                        funcionario_data['observacao'] = normalize_upper_ascii(row_data['OBSERVACAO'], allow_digits=True)

                    vinculo_matricula = None
                    if row_data.get('MATRICULA'):
                        vinculo_matricula = str(normalize_upper_ascii(row_data['MATRICULA'], allow_digits=True)).strip()
                        if not vinculo_matricula:
                            vinculo_matricula = None
                    
                    # Criar funcionário e vínculo de forma consistente
                    with transaction.atomic():
                        funcionario = Funcionario(**funcionario_data)
                        funcionario.save()

                        # Criar vínculo (empresa e datas vivem no vínculo, não no funcionário)
                        data_admissao = FuncionarioImportService.parse_date(row_data['DATA_ADMISSAO'])
                        data_demissao = None
                        if row_data.get('DATA_DEMISSAO'):
                            data_demissao = FuncionarioImportService.parse_date(row_data['DATA_DEMISSAO'])

                        FuncionarioVinculo.objects.create(
                            funcionario=funcionario,
                            empresa=empresa,
                            matricula=vinculo_matricula,
                            data_admissao=data_admissao,
                            data_demissao=data_demissao,
                            salario=str(row_data.get('SALARIO')).strip() if row_data.get('SALARIO') is not None else None,
                            cargo=str(row_data.get('CARGO')).strip() or None if row_data.get('CARGO') else None,
                        )

                        # Validar após vínculo existir
                        funcionario.full_clean()
                    
                    result['success'] += 1
                    result['created_funcionarios'].append(funcionario.id)
                    
                except Exception as e:
                    error_msg = f"Linha {row_idx}: {str(e)}"
                    result['errors'].append(error_msg)
                
                result['total'] += 1
            
            return result
            
        except Exception as e:
            raise Exception(f"Erro ao processar arquivo: {str(e)}")


class VinculoUpdateService:
    """Serviço para atualização em lote de vínculos existentes via XLSX."""

    MOTIVOS_VALIDOS = {
        'transferencia', 'pedido_demissao',
        'demissao_sem_justa_causa', 'demissao_justa_causa', 'outro',
    }

    UPDATE_COLUMNS = [
        'EMPRESA', 'MATRICULA', 'CPF',
        'CARGO', 'MOTIVO_SAIDA', 'DATA_DEMISSAO', 'SALARIO', 'OBSERVACOES',
    ]

    @staticmethod
    def generate_template_update_xlsx():
        """Gera XLSX modelo para atualização em lote de vínculos."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Atualizar Vínculos"

        header_fill = PatternFill(start_color="764ba2", end_color="764ba2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        example_fill = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        for col_idx, col_name in enumerate(VinculoUpdateService.UPDATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        example = [
            "CF001ABC",        # EMPRESA
            "1001",            # MATRICULA
            "",                # CPF (só se não tiver matrícula)
            "Analista FGTS",   # CARGO
            "pedido_demissao", # MOTIVO_SAIDA
            "2024-12-31",      # DATA_DEMISSAO
            "4500.00",         # SALARIO
            "Desligamento voluntário",  # OBSERVACOES
        ]
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = value
            cell.fill = example_fill
            cell.border = border

        info_row = 4
        ws.merge_cells(f'A{info_row}:H{info_row}')
        info_cell = ws[f'A{info_row}']
        info_cell.value = "⚠️ INSTRUÇÕES DE PREENCHIMENTO"
        info_cell.font = Font(bold=True, size=10, color="764ba2")

        instructions = [
            "• EMPRESA (obrigatório): código folha da empresa (ex: CF001ABC)",
            "• MATRICULA ou CPF (pelo menos um obrigatório): identifica o vínculo a atualizar",
            "• Campos em branco são IGNORADOS — apenas campos preenchidos serão atualizados",
            "• MOTIVO_SAIDA valores válidos: transferencia | pedido_demissao | demissao_sem_justa_causa | demissao_justa_causa | outro",
            "• DATA_DEMISSAO formato: YYYY-MM-DD (ex: 2024-12-31) ou DD/MM/YYYY",
            "• SALARIO formato: número com ponto decimal (ex: 4500.00)",
            "• Apague a linha de exemplo antes de importar seus dados",
        ]
        for idx, text in enumerate(instructions, 1):
            r = info_row + idx
            ws.merge_cells(f'A{r}:H{r}')
            c = ws[f'A{r}']
            c.value = text
            c.font = Font(size=9)
            c.alignment = Alignment(wrap_text=True, vertical='top')

        col_widths = [14, 14, 16, 20, 28, 16, 12, 30]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
        ws.row_dimensions[1].height = 30

        return wb

    @staticmethod
    def update_vinculos_from_file(file, user):
        """
        Atualiza vínculos existentes a partir de um arquivo XLSX.

        Identifica cada vínculo por (empresa + matricula) ou fallback (empresa + cpf).
        Atualiza apenas os campos não vazios: cargo, motivo_saida, data_demissao,
        salario, observacoes.

        Returns:
            dict com keys: total, success, errors (list[str])
        """
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception as e:
            raise Exception(f"Erro ao abrir arquivo: {e}")

        # Mapeia cabeçalhos
        headers = {}
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[str(cell.value).strip().upper()] = col_idx

        if 'EMPRESA' not in headers:
            raise ValueError("Coluna 'EMPRESA' não encontrada no arquivo.")
        if 'MATRICULA' not in headers and 'CPF' not in headers:
            raise ValueError("O arquivo precisa ter a coluna 'MATRICULA' ou 'CPF'.")

        result = {'total': 0, 'success': 0, 'errors': []}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            result['total'] += 1

            def cell(col):
                idx = headers.get(col)
                if idx is None:
                    return None
                v = row[idx - 1]
                return str(v).strip() if v is not None else None

            try:
                # 1. Resolve empresa
                empresa_raw = cell('EMPRESA')
                if not empresa_raw:
                    raise ValueError("Coluna EMPRESA está vazia.")

                empresa = FuncionarioImportService._resolve_empresa_from_identifier(empresa_raw)
                if not isinstance(empresa, Empresa):
                    raise ValueError(f"Empresa '{empresa_raw}' não encontrada.")

                if not is_empresa_allowed(user, empresa.codigo):
                    raise ValueError(f"Sem permissão para a empresa '{empresa.nome}'.")

                # 2. Localiza vínculo por matrícula ou CPF
                matricula = cell('MATRICULA')
                cpf_raw = cell('CPF')
                vinculo = None

                if matricula:
                    vinculo = FuncionarioVinculo.objects.filter(
                        empresa=empresa, matricula=matricula
                    ).order_by('-data_admissao').first()

                if vinculo is None and cpf_raw:
                    cpf_digits = ''.join(c for c in cpf_raw if c.isdigit())
                    vinculo = FuncionarioVinculo.objects.filter(
                        empresa=empresa,
                        funcionario__cpf__icontains=cpf_digits,
                    ).order_by('-data_admissao').first()

                if vinculo is None:
                    chave = matricula or cpf_raw or '(sem chave)'
                    raise ValueError(f"Vínculo não encontrado para '{chave}' na empresa '{empresa.nome}'.")

                # 3. Atualiza apenas campos fornecidos
                campos_atualizados = []

                cargo = cell('CARGO')
                if cargo:
                    vinculo.cargo = cargo
                    campos_atualizados.append('cargo')

                motivo = cell('MOTIVO_SAIDA')
                if motivo:
                    motivo_lower = motivo.lower().strip()
                    if motivo_lower not in VinculoUpdateService.MOTIVOS_VALIDOS:
                        raise ValueError(
                            f"MOTIVO_SAIDA inválido: '{motivo}'. "
                            f"Valores aceitos: {', '.join(sorted(VinculoUpdateService.MOTIVOS_VALIDOS))}"
                        )
                    vinculo.motivo_saida = motivo_lower
                    campos_atualizados.append('motivo_saida')

                data_demissao_raw = cell('DATA_DEMISSAO')
                if data_demissao_raw:
                    vinculo.data_demissao = FuncionarioImportService.parse_date(data_demissao_raw)
                    campos_atualizados.append('data_demissao')

                salario_raw = cell('SALARIO')
                if salario_raw:
                    from decimal import Decimal, InvalidOperation
                    try:
                        vinculo.salario = Decimal(salario_raw.replace(',', '.'))
                    except InvalidOperation:
                        raise ValueError(f"SALARIO inválido: '{salario_raw}'. Use formato numérico (ex: 3500.00).")
                    campos_atualizados.append('salario')

                observacoes = cell('OBSERVACOES')
                if observacoes:
                    vinculo.observacoes = observacoes
                    campos_atualizados.append('observacoes')

                if not campos_atualizados:
                    raise ValueError("Nenhum campo para atualizar nesta linha.")

                vinculo.save()
                result['success'] += 1

            except Exception as e:
                result['errors'].append(f"Linha {row_idx}: {e}")

        return result
